import os
import json
import calendar
import random
import re
import smtplib
import sqlite3
import string
import subprocess
import sys
import threading
import unicodedata
import time
import webbrowser
import socket
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from io import BytesIO
from functools import wraps, lru_cache
from html import escape as html_escape
from pathlib import Path
from typing import Any


def ensure_local_dependencies() -> None:
    """Permite abrir o app.py direto no Windows sem precisar passar pelo VS Code.

    Em produção/Render, as dependências continuam vindo do requirements.txt.
    Localmente, se o usuário der duplo clique no app.py e faltar algum pacote,
    o próprio script tenta instalar antes de continuar.
    """
    if os.getenv("RENDER") or os.getenv("DISABLE_AUTO_INSTALL", "false").lower() in {"1", "true", "yes", "sim"}:
        return

    required = {
        "flask": "Flask==3.0.3",
        "openpyxl": "openpyxl==3.1.5",
        "reportlab": "reportlab==4.2.5",
        "svglib": "svglib==1.5.1",
        "requests": "requests==2.32.3",
        "boto3": "boto3==1.35.99",
    }
    missing = []
    try:
        import importlib.util
        for module_name, package_name in required.items():
            if importlib.util.find_spec(module_name) is None:
                missing.append(package_name)
    except Exception:
        return

    if not missing:
        return

    print("\n[Portal de Insumos] Instalando dependências locais ausentes...")
    print("Pacotes:", ", ".join(missing))
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"])
    subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])


try:
    ensure_local_dependencies()
except Exception as exc:
    print("\nNão consegui instalar as dependências automaticamente.")
    print("Erro:", exc)
    print("\nRode o arquivo ABRIR_PORTAL.bat ou execute: pip install -r requirements.txt")
    input("\nPressione Enter para fechar...")
    raise

from flask import (
    Flask,
    abort,
    flash,
    g,
    has_request_context,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
import requests
from werkzeug.security import check_password_hash, generate_password_hash

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from svglib.svglib import svg2rlg

from cloudflare_d1 import cloudflare_d1_connect_from_env
from cloudflare_r2 import download_bytes_from_r2, upload_bytes_to_r2

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)
PRODUCT_IMAGE_DIR = INSTANCE_DIR / "product_images"
PRODUCT_IMAGE_DIR.mkdir(exist_ok=True)
PRODUCT_IMAGE_MAX_BYTES = 3 * 1024 * 1024
PRODUCT_IMAGE_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}

# Fonte CJK para PDFs. As fontes padrão do ReportLab (Helvetica) não
# renderizam caracteres chineses, por isso o PDF acabava mostrando
# quadradinhos/caixas pretas. STSong-Light é uma fonte CID nativa do
# ReportLab/Adobe para chinês simplificado e também segura para textos mistos.
PDF_TEXT_FONT = "Helvetica"
PDF_TEXT_FONT_BOLD = "Helvetica-Bold"
PDF_CJK_FONT = "Helvetica"
try:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    registerFontFamily(
        "STSong-Light",
        normal="STSong-Light",
        bold="STSong-Light",
        italic="STSong-Light",
        boldItalic="STSong-Light",
    )
    PDF_TEXT_FONT = "STSong-Light"
    PDF_TEXT_FONT_BOLD = "STSong-Light"
    PDF_CJK_FONT = "STSong-Light"
except Exception as exc:
    print(f"[PDF] Aviso: não foi possível registrar fonte chinesa STSong-Light: {exc}")

REQUEST_PDF_FONT = "Helvetica"
REQUEST_PDF_FONT_BOLD = "Helvetica-Bold"

app = Flask(__name__)

@app.route("/favicon.ico")
def favicon_ico():
    """Serve diretamente o SVG oficial da J&T também no caminho padrão."""
    return send_file(BASE_DIR / "static" / "img" / "browser-favicon.svg", mimetype="image/svg+xml")

app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-change-me")

DATABASE_DRIVER = os.getenv("DATABASE_DRIVER", "sqlite").strip().lower()
DATABASE_PATH = os.getenv("DATABASE_PATH", str(INSTANCE_DIR / "jt_insumos.db"))
ADMIN_CODE_MINUTES = int(os.getenv("ADMIN_CODE_MINUTES", "10"))

# SMTP opcional. Se não configurar, o código de confirmação de login aparece no terminal e em flash.
MAIL_SERVER = os.getenv("MAIL_SERVER", "").strip()
MAIL_PORT = int(os.getenv("MAIL_PORT", "587"))
MAIL_USE_TLS = os.getenv("MAIL_USE_TLS", "true").lower() == "true"
MAIL_USE_SSL = os.getenv("MAIL_USE_SSL", "false").lower() in {"1", "true", "yes", "sim"}
MAIL_USERNAME = os.getenv("MAIL_USERNAME", "").strip()
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD", "").strip()
MAIL_DEFAULT_SENDER = os.getenv("MAIL_DEFAULT_SENDER", MAIL_USERNAME or "no-reply@jt-insumos.local").strip()

FEISHU_STOCK_WEBHOOK_URL = os.getenv(
    "FEISHU_STOCK_WEBHOOK_URL",
    "https://open.feishu.cn/open-apis/bot/v2/hook/0759b1b1-b0ac-413b-8672-c113012c14db",
).strip()
PUBLIC_BASE_URL = os.getenv(
    "PUBLIC_BASE_URL",
    os.getenv("APP_PUBLIC_URL", os.getenv("RENDER_EXTERNAL_URL", "")),
).strip().rstrip("/")


# Lista oficial de bases/franquias para cadastro e administração de usuários.
BASE_FRANCHISE_OPTIONS_RAW = """ABC-MG
AET -MG
AFS -MG
AGF -MG
AMN -MG
AUI -MG
BCA -MG
BDA-MG
BHM-MG
BHZ 03-MG
BHZ 04-MG
BHZ 05-MG
BHZ 06-MG
BMN -MG
BTM 02-MG
BTM -MG
CAZ-MG
CGE -MG
CPA -MG
CRL -MG
CSP -MG
CTG -MG
CUV -MG
DIQ 02-MG
DIQ -MG
DMA -MG
ELD-MG
EXT -MG
F AAD-MG
F AAX-MG
F ADD-MG
F ALP 02-MG
F BCA-MG
F BCV-MG
F BDP-MG
F BHZ 08-MG
F BHZ 10-MG
F BHZ-MG
F CAR 02-MG
F CET-MG
F CGS-MG
F CLT-MG
F CMB-MG
F COG-MG
F CPA-MG
F CRT-MG
F DEL-MG
F DIV-MG
F EXT-MG
F FMG-MG
F GXP-MG
F IAJ-MG
F IAN-MG
F IMA-MG
F IRO-MG
F IRT-MG
F JDF 02-MG
F JDF 03-MG
F JDF 04-MG
F JDF-MG
F JPN-MG
F LGP-MG
F LPL-MG
F MNT-MG
F MPB-MG
F MRC-MG
F MTA-MG
F MTL-MG
F MTS-MG
F NSR-MG
F OPT-MG
F ORB-MG
F PDS-MG
F PLE-MG
F PMN-MG
F PPY 02-MG
F PSS-MG
F SAB-MG
F SGT-MG
F SJN-MG
F SSP-MG
F STD-MG
F STL-MG
F STS-MG
F TMN-MG
F TPT-MG
F TRC-MG
F UBE-MG
F UNI-MG
F VSR-MG
FAB-MG
FRT -MG
FUN-MG
GHE -MG
GVR -MG
IBR -MG
IGP-MG
IPN 02-MG
IPN -MG
IRM -MG
ITA -MG
ITU -MG
JDI-MG
JMV -MG
JNA -MG
JTB -MG
JUB -MG
LAS -MG
LGP -MG
LGS-MG
MNH -MG
MRE -MG
MTC -MG
NVL -MG
OLV -MG
OPT -MG
ORP-MG
PDE-MG
POJ -MG
POO -MG
PPR -MG
PPY -MG
PSS -MG
PTC -MG
PTN -MG
PTU -MG
QDF -MG
RDN 02-MG
RDN -MG
SAB -MG
SGF-MG
SJD -MG
SLN -MG
SON-MG
STL -MG
STZ -MG
SZD-MG
TFL -MG
UAB -MG
UBA 02-MG
UBA 03-MG
UBA -MG
UDI 02-MG
UDI -MG
VAG -MG
VCS -MG
VPS -MG
BDC-MG
JDF-MG
ARA-SP
ARU 02-SP
ARU-SP
AVR-SP
BAU 02-SP
BAU-SP
BBD-SP
BGI-SP
BTC-SP
CPE-SP
DCN SP
F AND-SP
F ASI-SP
F AUR-SP
F BAT-SP
F BDB-SP
F BRB-SP
F CTD-SP
F GNS-SP
F GPC-SP
F GRC-SP
F GRR-SP
F IBG-SP
F ITAP-SP
F ITV 02-SP
F JAU-SP
F JLS-SP
F JSB-SP
F LCP-SP
F LNS-SP
F MAT-SP
F MII-SP
F MRS-SP
F MTT-SP
F NVH-SP
F ORLN-SP
F OSW-SP
F OUR-SP
F PFT-SP
F PNP-SP
F PTR-SP
F QSC 03-SP
F RAO 02-SP
F RPD-SP
F SBV 02-SP
F SBV-SP
F SJB-SP
F STF-SP
F TPA-SP
F TQA-SP
F VTP-SP
FERN-SP
FRC 02-SP
FRC-SP
JBT-SP
MII-SP
PPB-SP
PRSS-SP
RAO 02-SP
RAO 03-SP
RAO-SP
SCL-SP
SJP 02-SP
SJP-SP
VGS-SP
"""
def _unit_sort_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).upper()


BASE_FRANCHISE_OPTIONS = sorted(
    list(dict.fromkeys(line.strip() for line in BASE_FRANCHISE_OPTIONS_RAW.splitlines() if line.strip())),
    key=_unit_sort_key,
)
# Franquias são as unidades com prefixo oficial "F " na lista.
FRANCHISE_UNIT_OPTIONS = [unit for unit in BASE_FRANCHISE_OPTIONS if unit.upper().startswith("F ")]
# Bases são todas as demais unidades da lista oficial.
BASE_UNIT_OPTIONS = [unit for unit in BASE_FRANCHISE_OPTIONS if not unit.upper().startswith("F ")]
BASE_FRANCHISE_OPTION_SET = set(BASE_FRANCHISE_OPTIONS)
BASE_UNIT_OPTION_SET = set(BASE_UNIT_OPTIONS)
FRANCHISE_UNIT_OPTION_SET = set(FRANCHISE_UNIT_OPTIONS)


def normalize_unit_lookup_key(value: Any) -> str:
    """Normaliza nomes de bases/franquias vindos de Excel ou formulário.

    A importação do Excel pode trazer NBSP, hífen diferente, espaço duplicado,
    célula numérica ou código sem espaço entre letras e números (ex.: RAO02-SP).
    Essa chave permite aceitar o nome correto mesmo com essas diferenças visuais.
    """
    text = str(value or "")
    text = text.replace("\u00a0", " ").replace("\u2007", " ").replace("\u202f", " ")
    text = text.replace("–", "-").replace("—", "-").replace("−", "-")
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"([A-Z])([0-9])", r"\1 \2", text)
    text = re.sub(r"([0-9])([A-Z])", r"\1 \2", text)
    return " ".join(text.split())


def build_unit_lookup(options: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for option in options:
        key = normalize_unit_lookup_key(option)
        if key and key not in lookup:
            lookup[key] = option
        compact_key = key.replace(" ", "")
        if compact_key and compact_key not in lookup:
            lookup[compact_key] = option
    return lookup


BASE_UNIT_OPTION_LOOKUP = build_unit_lookup(BASE_UNIT_OPTIONS)
FRANCHISE_UNIT_OPTION_LOOKUP = build_unit_lookup(FRANCHISE_UNIT_OPTIONS)
BASE_FRANCHISE_OPTION_LOOKUP = build_unit_lookup(BASE_FRANCHISE_OPTIONS)


def canonical_unit_option(value: Any, options: list[str], lookup: dict[str, str]) -> str:
    raw = str(value or "").strip().replace("\u00a0", " ").replace("\u2007", " ").replace("\u202f", " ")
    raw = re.sub(r"\s+", " ", raw).strip()
    if not raw:
        return ""
    if raw in set(options):
        return raw
    key = normalize_unit_lookup_key(raw)
    if key in lookup:
        return lookup[key]
    compact_key = key.replace(" ", "")
    if compact_key in lookup:
        return lookup[compact_key]
    # Fallback: comparação por containment apenas quando não gera ambiguidade.
    matches = [option for option in options if key and (key == normalize_unit_lookup_key(option) or key in normalize_unit_lookup_key(option) or normalize_unit_lookup_key(option) in key)]
    unique_matches = list(dict.fromkeys(matches))
    if len(unique_matches) == 1:
        return unique_matches[0]
    return ""
ASSET_SPECIAL_REGIONAL_OPTIONS = {"Matriz", "SC CGE", "SC RAO"}
ASSET_REGIONAL_OPTIONS = ["MG", "SPN", "Matriz", "SC CGE", "SC RAO"]
ASSET_REGIONAL_OPTION_SET = {option.upper() for option in ASSET_REGIONAL_OPTIONS}
ADMIN_ORGANIZATION_NAME = "ADMINISTRAÇÃO"
ADMIN_ORGANIZATION_OPTIONS = [ADMIN_ORGANIZATION_NAME]
DEV_ACCESS_PASSWORD_KEY = "dev_access_password"
DEFAULT_DEV_ACCESS_PASSWORD = os.getenv("DEV_ACCESS_PASSWORD", "DevJet2026")
SUPPLY_STOCK_TAG = "insumos"
ASSET_STOCK_TAG = "ativos"
DEFAULT_STOCK_TAG = SUPPLY_STOCK_TAG
DEFAULT_PRODUCT_REQUEST_BLOCK_MONTHS = 2
SYSTEM_STOCK_TAGS = [
    (SUPPLY_STOCK_TAG, "Insumos", "Estoque usado nas solicitacoes de insumos."),
    (ASSET_STOCK_TAG, "Ativos", "Estoque usado no cadastro e baixa de ativos."),
]


@dataclass
class User:
    id: int
    responsible_name: str
    organization_name: str
    franchise_name: str
    franchise_number: str
    cnpj: str
    username: str
    email: str
    password_hash: str
    role: str
    status: str
    created_at: datetime
    updated_at: datetime | None = None
    page_permissions_configured: bool = False
    action_permissions_configured: bool = False

    @property
    def is_dev(self) -> bool:
        return canonical_role_key(self.role, "") == "dev"

    @property
    def is_admin(self) -> bool:
        return role_is_admin_like(canonical_role_key(self.role, ""))

    @property
    def is_approved(self) -> bool:
        return self.status == "approved"

    @property
    def formatted_cnpj(self) -> str:
        digits = normalize_cnpj(self.cnpj)
        return format_cnpj(digits) if len(digits) == 14 else ""

    @property
    def formatted_phone(self) -> str:
        digits = normalize_phone_number(self.franchise_number)
        return format_phone_number(digits) if len(digits) in {10, 11} else ""


@dataclass
class AccessRoleType:
    role_key: str
    name: str
    description: str = ""
    permissions: list[str] = field(default_factory=list)
    action_permissions: list[str] = field(default_factory=list)
    editable_roles: list[str] = field(default_factory=list)
    editable_user_fields: list[str] = field(default_factory=list)
    created_at: datetime | None = None
    updated_at: datetime | None = None
    is_static: bool = False

    @property
    def is_admin_like(self) -> bool:
        admin_keys = {item["key"] for item in PAGE_PERMISSION_OPTIONS if item["admin_only"]}
        return bool(set(self.permissions) & admin_keys)


@dataclass
class StockTag:
    slug: str
    name: str
    description: str = ""
    active: bool = True
    system_key: bool = False
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime | None = None


@dataclass
class Product:
    id: int = 0
    name: str = ""
    category: str = ""
    category_emoji: str = ""
    image_name: str = ""
    image_key: str = ""
    image_content_type: str = ""
    unit_measure: str = "un"
    is_kit: bool = False
    kit_quantity: int = 1
    description: str = ""
    stock_quantity: int = 0
    price_cents: int = 0
    limit_base: int | None = None
    limit_franchise: int | None = None
    limit_block_days: int = 60
    min_order_quantity: int | None = None
    min_stock: int | None = None
    max_stock: int | None = None
    active: bool = True
    catalog_archived: bool = False
    visible_base: bool = True
    visible_franchise: bool = True
    internal: bool = False
    stock_tag: str = DEFAULT_STOCK_TAG
    stock_tag_name: str = "Insumos"
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime | None = None

    @property
    def price_brl(self) -> float:
        return self.price_cents / 100


@dataclass
class RequestItem:
    id: int
    request_id: int
    product_id: int
    product_name_snapshot: str
    quantity: int
    price_cents_snapshot: int
    product: Product | None = None


@dataclass
class SupplyRequest:
    id: int
    user_id: int
    status: str
    user_note: str
    admin_note: str
    people_count: int | None
    created_at: datetime
    reviewed_at: datetime | None = None
    reviewed_by_id: int | None = None
    user: User | None = None
    reviewed_by: User | None = None
    items: list[RequestItem] = field(default_factory=list)
    action_logs: list[dict[str, Any]] = field(default_factory=list)

    @property
    def total_cents(self) -> int:
        return sum((item.price_cents_snapshot or 0) * item.quantity for item in self.items)


@dataclass
class ProductRequestBlock:
    id: int
    user_id: int
    product_id: int
    product_name: str
    blocked_until: datetime
    reason: str = ""
    created_by_request_id: int | None = None
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime | None = None
    revoked_at: datetime | None = None
    updated_by_id: int | None = None
    source_request_status: str = ""

    @property
    def is_active(self) -> bool:
        if self.revoked_at is not None or self.blocked_until <= datetime.utcnow():
            return False
        if self.created_by_request_id is not None and self.source_request_status not in {"pending", "approved"}:
            return False
        return True

    @property
    def blocked_until_date_input(self) -> str:
        local_dt = to_sao_paulo_dt(self.blocked_until)
        return local_dt.strftime("%Y-%m-%d") if local_dt else ""

    @property
    def blocked_until_label(self) -> str:
        return format_sao_paulo_datetime(self.blocked_until)


@dataclass
class AssetItem:
    id: int
    asset_id: int
    item_name: str
    product_id: int | None = None
    quantity: int = 1
    serial_number: str = ""
    created_at: datetime = field(default_factory=datetime.utcnow)


@dataclass
class AssetRecord:
    id: int
    name: str
    base: str
    regional: str
    sector: str
    manager: str
    created_at: datetime
    updated_at: datetime | None = None
    created_by_id: int | None = None
    items: list[AssetItem] = field(default_factory=list)


@dataclass
class MaterialEntry:
    id: int
    product_id: int | None
    item_name: str
    quantity: int
    unit_measure: str
    unit_price_cents: int
    invoice_file_name: str = ""
    invoice_file_key: str = ""
    invoice_number: str = ""
    invoice_date: datetime | None = None
    invoice_value_cents: int = 0
    notes: str = ""
    created_by_id: int | None = None
    created_at: datetime = field(default_factory=datetime.utcnow)
    product: Product | None = None
    created_by_name: str = ""

    @property
    def total_cents(self) -> int:
        return int(self.quantity or 0) * int(self.unit_price_cents or 0)


# ---------- Banco SQLite sem SQLAlchemy ----------

def using_cloudflare_d1() -> bool:
    return DATABASE_DRIVER in {"cloudflare_d1", "d1"} or (
        bool(os.getenv("CLOUDFLARE_D1_DATABASE_ID")) and DATABASE_DRIVER not in {"sqlite", "local"}
    )


def db_connect():
    if using_cloudflare_d1():
        return cloudflare_d1_connect_from_env()
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ---------- Otimização Cloudflare D1 / limite de linhas lidas ----------

def low_row_read_mode() -> bool:
    """Reduz consultas grandes no Cloudflare D1 para preservar o limite mensal de Rows read."""
    flag = os.getenv("D1_LOW_ROW_READ", "").strip().lower()
    if flag in {"0", "false", "no", "nao", "não", "off"}:
        return False
    if flag in {"1", "true", "yes", "sim", "on"}:
        return True
    return using_cloudflare_d1()


DEFAULT_TABLE_PAGE_SIZE = 25
TABLE_PAGE_SIZE_OPTIONS = [25, 50, 100, 200, 300, 500]


def bounded_int(value: Any, default: int, minimum: int = 1, maximum: int = 500) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(minimum, min(maximum, parsed))


def list_page_limit(default: int = DEFAULT_TABLE_PAGE_SIZE, maximum: int = 300) -> int:
    """Limite das tabelas administrativas.

    O padrão visual precisa ser 25 linhas para evitar leituras grandes no D1.
    Variáveis antigas como D1_LIST_PAGE_SIZE não devem voltar a tela para 120
    quando o usuário não escolheu uma quantidade manualmente.
    """
    raw_limit = request.args.get("limit")
    if raw_limit is None or str(raw_limit).strip() == "":
        return bounded_int(default, 25, 25, maximum)
    return bounded_int(raw_limit, default, 25, maximum)


def api_page_limit(default: int = DEFAULT_TABLE_PAGE_SIZE, maximum: int = 250) -> int:
    """Limite padrão menor nas APIs para preservar Rows read do Cloudflare D1."""
    raw_limit = request.args.get("limit")
    if raw_limit is None or str(raw_limit).strip() == "":
        return bounded_int(default, 25, 25, maximum)
    return bounded_int(raw_limit, default, 25, maximum)


def exact_counts_enabled() -> bool:
    flag = os.getenv("D1_EXACT_COUNTS", "").strip().lower()
    if flag in {"1", "true", "yes", "sim", "on"}:
        return True
    if flag in {"0", "false", "no", "nao", "não", "off"}:
        return False
    return not low_row_read_mode()


def like_term(value: str) -> str:
    return f"%{str(value or '').strip()}%"


def get_cursor_lastrowid(cursor: Any) -> int | None:
    """Retorna o ID do último INSERT de forma segura para runtime e Pylance."""
    value = cursor.lastrowid
    return value if isinstance(value, int) else None


def open_browser_when_ready(port: int) -> None:
    """Abre o navegador automaticamente quando o site roda localmente."""
    if os.getenv("AUTO_OPEN_BROWSER", "true").lower() not in {"1", "true", "yes", "sim"}:
        return

    def _open() -> None:
        time.sleep(1.2)
        url = f"http://127.0.0.1:{port}"
        print(f"\nPortal aberto em: {url}\n")
        webbrowser.open(url)

    threading.Thread(target=_open, daemon=True).start()


def is_port_available(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.3)
        return sock.connect_ex(("127.0.0.1", port)) != 0


def resolve_port() -> int:
    """Usa PORT no Render. Localmente, evita falhar se a porta 5000 estiver ocupada."""
    env_port = os.getenv("PORT")
    if env_port:
        return int(env_port)

    preferred = 5000
    if is_port_available(preferred):
        return preferred

    for port in range(5001, 5015):
        if is_port_available(port):
            print(f"Porta 5000 ocupada. Usando porta {port}.")
            return port

    return preferred


def parse_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value)
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.utcnow()


def now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


SAO_PAULO_OFFSET = timedelta(hours=-3)
SAO_PAULO_TZ = timezone(SAO_PAULO_OFFSET, "GMT-3")


def sao_paulo_now() -> datetime:
    return datetime.utcnow() + SAO_PAULO_OFFSET


def to_sao_paulo_dt(value: Any | None = None) -> datetime | None:
    if value is None:
        return sao_paulo_now()
    parsed = parse_dt(value)
    if parsed is None:
        return None
    if parsed.tzinfo is not None:
        return parsed.astimezone(SAO_PAULO_TZ).replace(tzinfo=None)
    return parsed + SAO_PAULO_OFFSET


def format_sao_paulo_datetime(value: Any | None = None, fmt: str = "%d/%m/%Y %H:%M", suffix: str = "") -> str:
    local_dt = to_sao_paulo_dt(value)
    if local_dt is None:
        return ""
    return local_dt.strftime(fmt) + suffix


def sao_paulo_filename_timestamp() -> str:
    return sao_paulo_now().strftime("%Y%m%d_%H%M%S")


def sao_paulo_report_bounds_to_utc(start_local: datetime, end_local_date: datetime) -> tuple[datetime, datetime]:
    start_utc = start_local - SAO_PAULO_OFFSET
    end_local = end_local_date + timedelta(days=1) - timedelta(seconds=1)
    end_utc = end_local - SAO_PAULO_OFFSET
    return start_utc, end_utc


def normalize_username(value: str) -> str:
    value = (value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^a-z0-9._-]", "", value)
    return value


def valid_username(value: str) -> bool:
    username = normalize_username(value)
    return 3 <= len(username) <= 40 and bool(re.fullmatch(r"[a-z0-9][a-z0-9._-]*", username))


def synthetic_email_for_username(username: str) -> str:
    safe = normalize_username(username) or "usuario"
    return f"{safe}@usuario.local"




ROLE_KEY_ALIASES = {
    "base": "base",
    "unidade": "base",
    "franquia": "franchise",
    "franchise": "franchise",
    "admin": "admin",
    "administrador": "admin",
    "administradora": "admin",
    "dev": "dev",
    "desenvolvedor": "dev",
    "developer": "dev",
}


def canonical_role_key(value: Any, default: str = "base") -> str:
    raw = str(value or "").strip()
    if not raw:
        return default
    normalized = normalize_header(raw).replace("_", " ").strip()
    return ROLE_KEY_ALIASES.get(normalized) or raw.lower()


def normalize_user_role(value: Any, allow_admin: bool = True) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    normalized = normalize_header(raw).replace("_", " ").strip()
    role = ROLE_KEY_ALIASES.get(normalized)
    if role is None:
        raw_lower = raw.lower()
        if get_custom_access_role(raw_lower) is not None:
            role = raw_lower
        else:
            for custom in list_custom_access_roles():
                if normalized == normalize_header(custom.name).replace("_", " ").strip():
                    role = custom.role_key
                    break
    if role is None:
        return None
    if role in {"admin", "dev"} and not allow_admin:
        return None
    if role_is_admin_like(role) and not allow_admin:
        return None
    return role


def normalize_user_status(value: Any, default: str = "pending") -> str | None:
    raw = normalize_header(str(value or "")).replace("_", " ").strip()
    aliases = {
        "pendente": "pending",
        "pending": "pending",
        "aguardando": "pending",
        "aprovado": "approved",
        "aprovada": "approved",
        "approved": "approved",
        "ativo": "approved",
        "ativa": "approved",
        "recusado": "rejected",
        "recusada": "rejected",
        "rejected": "rejected",
        "reprovado": "rejected",
        "reprovada": "rejected",
    }
    if not raw:
        return default
    return aliases.get(raw, default)


def normalize_phone_number(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))[:11]


def normalize_cnpj(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))[:14]


def format_phone_number(value: Any) -> str:
    digits = normalize_phone_number(value)
    if not digits:
        return ""
    if len(digits) <= 2:
        return digits
    if len(digits) <= 6:
        return f"({digits[:2]}) {digits[2:]}"
    if len(digits) <= 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"


def format_cnpj(value: Any) -> str:
    digits = normalize_cnpj(value)
    if not digits:
        return ""
    if len(digits) <= 2:
        return digits
    if len(digits) <= 5:
        return f"{digits[:2]}.{digits[2:]}"
    if len(digits) <= 8:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:]}"
    if len(digits) <= 12:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:]}"
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def valid_organization_for_role(organization_name: str, role: str) -> bool:
    if role == "base":
        return organization_name in BASE_UNIT_OPTION_SET
    if role == "franchise":
        return organization_name in FRANCHISE_UNIT_OPTION_SET
    if role_is_admin_like(role):
        return bool(str(organization_name or "").strip())
    return True


def validate_user_profile_fields(
    role: str,
    organization_name: Any = "",
    franchise_name: Any = "",
    franchise_number: Any = "",
    cnpj: Any = "",
    strict_base: bool = True,
) -> tuple[str, str, str, str]:
    organization = str(organization_name or "").strip()
    franchise = str(franchise_name or "").strip()
    phone_digits = normalize_phone_number(franchise_number)
    cnpj_digits = normalize_cnpj(cnpj)

    if role == "base":
        canonical_base = canonical_unit_option(organization, BASE_UNIT_OPTIONS, BASE_UNIT_OPTION_LOOKUP)
        if canonical_base:
            return canonical_base, "", "", ""
        if not organization:
            raise ValueError("Informe o nome da base.")
        if strict_base:
            raise ValueError("Selecione uma base válida.")
        # Na importação em massa, não bloqueia bases oficiais ainda não cadastradas
        # na lista local do app. Salva o texto limpo exatamente como veio da planilha.
        return organization[:160], "", "", ""

    if role == "franchise":
        if not franchise:
            raise ValueError("Informe o nome da franquia.")
        if phone_digits and len(phone_digits) not in {10, 11}:
            raise ValueError("Informe um telefone válido com DDD, usando apenas números.")
        if cnpj_digits and len(cnpj_digits) != 14:
            raise ValueError("O CNPJ deve possuir 14 números.")
        franchise_clean = franchise[:160]
        return franchise_clean, franchise_clean, phone_digits, cnpj_digits

    if role in {"admin", "dev"} or get_custom_access_role(role) is not None:
        setor = organization[:160] if organization else ADMIN_ORGANIZATION_NAME
        return setor, "", phone_digits, cnpj_digits

    raise ValueError("Tipo de acesso inválido.")

def is_real_email(value: str | None) -> bool:
    value = (value or "").strip().lower()
    return "@" in value and not value.endswith("@usuario.local")


def row_to_user(row: Any | None) -> User | None:
    if row is None:
        return None
    raw_role = (row["role"] if "role" in row.keys() else "") or "base"
    role = normalize_user_role(raw_role, allow_admin=True) or canonical_role_key(raw_role, "base")
    return User(
        id=int(row["id"]),
        responsible_name=row["responsible_name"] or "",
        organization_name=row["organization_name"] or "",
        franchise_name=(row["franchise_name"] if "franchise_name" in row.keys() else "") or "",
        franchise_number=(row["franchise_number"] if "franchise_number" in row.keys() else "") or "",
        cnpj=(row["cnpj"] if "cnpj" in row.keys() else "") or "",
        username=(row["username"] if "username" in row.keys() else "") or normalize_username((row["email"] if "email" in row.keys() else "") or row["responsible_name"] or "usuario"),
        email=(row["email"] if "email" in row.keys() else "") or "",
        password_hash=row["password_hash"] or "",
        role=role,
        status=row["status"] or "pending",
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        updated_at=parse_dt(row["updated_at"]),
        page_permissions_configured=bool(row["page_permissions_configured"]) if "page_permissions_configured" in row.keys() else False,
        action_permissions_configured=bool(row["action_permissions_configured"]) if "action_permissions_configured" in row.keys() else False,
    )


def default_category_emoji(category: str) -> str:
    normalized = "".join(
        char
        for char in unicodedata.normalize("NFD", str(category or "").casefold())
        if unicodedata.category(char) != "Mn"
    )
    emoji_by_keyword = (
        (("administrativo", "escritorio", "papelaria"), "🗂️"),
        (("epi", "seguranca", "protecao"), "🦺"),
        (("embalagem", "envelope", "caixa"), "📦"),
        (("etiqueta", "label"), "🏷️"),
        (("limpeza", "higiene"), "🧹"),
        (("tecnologia", "informatica", "eletronico"), "💻"),
        (("uniforme", "vestuario"), "👕"),
        (("ferramenta", "manutencao", "operacional"), "🛠️"),
        (("impressao", "grafico"), "🖨️"),
    )
    for keywords, emoji in emoji_by_keyword:
        if any(keyword in normalized for keyword in keywords):
            return emoji
    return "📦"


def clean_category_emoji(value: Any, category: str = "") -> str:
    emoji = str(value or "").strip()
    return emoji[:16] if emoji else default_category_emoji(category)


def normalize_stock_tag_slug(value: Any, default: str = DEFAULT_STOCK_TAG) -> str:
    text_value = str(value or "").strip()
    if not text_value:
        return default
    text_value = "".join(ch for ch in unicodedata.normalize("NFD", text_value) if unicodedata.category(ch) != "Mn")
    text_value = re.sub(r"[^A-Za-z0-9]+", "-", text_value).strip("-").lower()
    return text_value or default


def default_stock_tag_name(slug: str) -> str:
    normalized = normalize_stock_tag_slug(slug)
    for system_slug, name, _description in SYSTEM_STOCK_TAGS:
        if system_slug == normalized:
            return name
    return normalized.replace("-", " ").title()


def can_manage_stock_tags(user: User | None = None) -> bool:
    user = user if user is not None else current_user()
    return bool(user and canonical_role_key(user.role, "") in {"admin", "dev"})


def row_to_stock_tag(row: Any | None) -> StockTag | None:
    if row is None:
        return None
    return StockTag(
        slug=normalize_stock_tag_slug(row["slug"]),
        name=(row["name"] or default_stock_tag_name(row["slug"])).strip(),
        description=(row["description"] if "description" in row.keys() else "") or "",
        active=bool(row["active"]) if "active" in row.keys() else True,
        system_key=bool(row["system_key"]) if "system_key" in row.keys() else False,
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        updated_at=parse_dt(row["updated_at"]) if "updated_at" in row.keys() else None,
    )


def list_stock_tags(active_only: bool = False, include_slug: str = "") -> list[StockTag]:
    clauses = []
    params: list[Any] = []
    if active_only:
        include_normalized = normalize_stock_tag_slug(include_slug, "") if include_slug else ""
        if include_normalized:
            clauses.append("(active = 1 OR slug = ?)")
            params.append(include_normalized)
        else:
            clauses.append("active = 1")
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    with db_connect() as conn:
        rows = conn.execute(
            f"SELECT * FROM stock_tags{where_sql} ORDER BY system_key DESC, name COLLATE NOCASE ASC",
            params,
        ).fetchall()
    return [tag for row in rows if (tag := row_to_stock_tag(row)) is not None]


def stock_tag_label(slug: str) -> str:
    normalized = normalize_stock_tag_slug(slug)
    for tag in list_stock_tags(active_only=False):
        if tag.slug == normalized:
            return tag.name
    return default_stock_tag_name(normalized)


def allowed_product_stock_tag(slug: str, current_slug: str = "") -> str:
    normalized = normalize_stock_tag_slug(slug)
    allowed = {tag.slug for tag in list_stock_tags(active_only=True, include_slug=current_slug)}
    return normalized if normalized in allowed else normalize_stock_tag_slug(current_slug or DEFAULT_STOCK_TAG)


def row_to_product(row: Any | None) -> Product | None:
    if row is None:
        return None
    category = row["category"] or ""
    stock_tag_slug = normalize_stock_tag_slug(row["stock_tag"] if "stock_tag" in row.keys() else DEFAULT_STOCK_TAG)
    return Product(
        id=int(row["id"]),
        name=row["name"] or "",
        category=category,
        category_emoji=clean_category_emoji(
            row["category_emoji"] if "category_emoji" in row.keys() else "",
            category,
        ),
        image_name=(row["image_name"] if "image_name" in row.keys() else "") or "",
        image_key=(row["image_key"] if "image_key" in row.keys() else "") or "",
        image_content_type=(row["image_content_type"] if "image_content_type" in row.keys() else "") or "",
        unit_measure=(row["unit_measure"] if "unit_measure" in row.keys() else None) or "un",
        is_kit=bool(row["is_kit"]) if "is_kit" in row.keys() else False,
        kit_quantity=max(1, int(row["kit_quantity"] or 1)) if "kit_quantity" in row.keys() else 1,
        description=row["description"] or "",
        stock_quantity=int(row["stock_quantity"] or 0),
        price_cents=int(row["price_cents"] or 0),
        limit_base=row["limit_base"] if "limit_base" in row.keys() and row["limit_base"] is not None else None,
        limit_franchise=row["limit_franchise"] if "limit_franchise" in row.keys() and row["limit_franchise"] is not None else None,
        limit_block_days=max(1, int(row["limit_block_days"] or 60)) if "limit_block_days" in row.keys() else 60,
        min_order_quantity=row["min_order_quantity"] if "min_order_quantity" in row.keys() and row["min_order_quantity"] is not None else None,
        min_stock=row["min_stock"] if "min_stock" in row.keys() and row["min_stock"] is not None else None,
        max_stock=row["max_stock"] if "max_stock" in row.keys() and row["max_stock"] is not None else None,
        active=bool(row["active"]),
        catalog_archived=bool(row["catalog_archived"]) if "catalog_archived" in row.keys() else False,
        visible_base=bool(row["visible_base"]) if "visible_base" in row.keys() else True,
        visible_franchise=bool(row["visible_franchise"]) if "visible_franchise" in row.keys() else True,
        internal=bool(row["internal"]) if "internal" in row.keys() else False,
        stock_tag=stock_tag_slug,
        stock_tag_name=(row["stock_tag_name"] if "stock_tag_name" in row.keys() and row["stock_tag_name"] else default_stock_tag_name(stock_tag_slug)),
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        updated_at=parse_dt(row["updated_at"]),
    )


def row_to_item(row: Any | None, load_product: bool = True) -> RequestItem | None:
    if row is None:
        return None
    item = RequestItem(
        id=int(row["id"]),
        request_id=int(row["request_id"]),
        product_id=int(row["product_id"]),
        product_name_snapshot=row["product_name_snapshot"] or "",
        quantity=int(row["quantity"] or 0),
        price_cents_snapshot=int(row["price_cents_snapshot"] or 0),
    )
    if load_product:
        item.product = get_product(item.product_id)
    return item


def row_to_supply_request(row: Any | None, include_user: bool = True, include_items: bool = True, include_actions: bool = False) -> SupplyRequest | None:
    if row is None:
        return None
    row_keys = set(row.keys()) if hasattr(row, "keys") else set()
    reviewed_by_id = row["reviewed_by_id"] if "reviewed_by_id" in row_keys else None
    req = SupplyRequest(
        id=int(row["id"]),
        user_id=int(row["user_id"]),
        status=row["status"] or "pending",
        user_note=row["user_note"] or "",
        admin_note=row["admin_note"] or "",
        people_count=int(row["people_count"] or 0) if "people_count" in row_keys and row["people_count"] is not None else None,
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        reviewed_at=parse_dt(row["reviewed_at"]) if "reviewed_at" in row_keys else None,
        reviewed_by_id=int(reviewed_by_id) if reviewed_by_id is not None else None,
    )
    if include_user:
        req.user = get_user(req.user_id)
    if req.reviewed_by_id is not None:
        req.reviewed_by = get_user(req.reviewed_by_id)
    if include_items:
        req.items = get_request_items(req.id)
    if include_actions:
        req.action_logs = list_request_action_logs(req.id)
    return req


def row_to_product_request_block(row: Any | None) -> ProductRequestBlock | None:
    if row is None:
        return None
    product_name = ""
    try:
        product_name = (row["product_name"] if "product_name" in row.keys() else "") or ""
    except Exception:
        product_name = ""
    return ProductRequestBlock(
        id=int(row["id"]),
        user_id=int(row["user_id"]),
        product_id=int(row["product_id"]),
        product_name=product_name or f"Produto #{int(row['product_id'])}",
        blocked_until=parse_dt(row["blocked_until"]) or datetime.utcnow(),
        reason=(row["reason"] if "reason" in row.keys() else "") or "",
        created_by_request_id=(int(row["created_by_request_id"]) if "created_by_request_id" in row.keys() and row["created_by_request_id"] is not None else None),
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        updated_at=parse_dt(row["updated_at"]) if "updated_at" in row.keys() else None,
        revoked_at=parse_dt(row["revoked_at"]) if "revoked_at" in row.keys() else None,
        updated_by_id=(int(row["updated_by_id"]) if "updated_by_id" in row.keys() and row["updated_by_id"] is not None else None),
        source_request_status=((row["source_request_status"] if "source_request_status" in row.keys() else "") or ""),
    )


def row_to_asset_item(row: Any | None) -> AssetItem | None:
    if row is None:
        return None
    row_keys = set(row.keys()) if hasattr(row, "keys") else set()
    product_id = row["product_id"] if "product_id" in row_keys else None
    quantity = row["quantity"] if "quantity" in row_keys else 1
    return AssetItem(
        id=int(row["id"]),
        asset_id=int(row["asset_id"]),
        item_name=row["item_name"] or "",
        product_id=int(product_id) if product_id else None,
        quantity=int(quantity or 1),
        serial_number=row["serial_number"] or "",
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
    )


def get_asset_items(conn: Any, asset_id: int) -> list[AssetItem]:
    rows = conn.execute(
        "SELECT * FROM asset_items WHERE asset_id = ? ORDER BY id ASC",
        (asset_id,),
    ).fetchall()
    return [item for row in rows if (item := row_to_asset_item(row)) is not None]


def row_to_asset(row: Any | None, conn: Any | None = None, include_items: bool = True) -> AssetRecord | None:
    if row is None:
        return None
    asset = AssetRecord(
        id=int(row["id"]),
        name=row["name"] or "",
        base=row["base"] or "",
        regional=row["regional"] or "",
        sector=row["sector"] or "",
        manager=row["manager"] or "",
        created_by_id=row["created_by_id"],
        created_at=parse_dt(row["created_at"]) or datetime.utcnow(),
        updated_at=parse_dt(row["updated_at"]),
    )
    if include_items:
        if conn is not None:
            asset.items = get_asset_items(conn, asset.id)
        else:
            with db_connect() as local_conn:
                asset.items = get_asset_items(local_conn, asset.id)
    return asset


def list_assets(base: str = "", regional: str = "") -> list[AssetRecord]:
    sql = "SELECT * FROM assets"
    params: list[Any] = []
    clauses: list[str] = []
    if base:
        clauses.append("base = ?")
        params.append(base)
    if regional:
        clauses.append("regional = ?")
        params.append(regional)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY created_at DESC, id DESC"
    with db_connect() as conn:
        rows = conn.execute(sql, params).fetchall()
        return [asset for row in rows if (asset := row_to_asset(row, conn=conn)) is not None]


def get_asset(asset_id: int) -> AssetRecord | None:
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM assets WHERE id = ?", (int(asset_id),)).fetchone()
        return row_to_asset(row, conn=conn)


def normalize_asset_regional(value: str) -> str:
    normalized = (value or "").strip().upper()
    if normalized == "MATRIZ":
        return "Matriz"
    if normalized in {"SC CGE", "SC-CGE", "SC_CGE", "SCCGE"}:
        return "SC CGE"
    if normalized in {"SC RAO", "SC-RAO", "SC_RAO", "SCRAO"}:
        return "SC RAO"
    if normalized in {"MG", "SPN"}:
        return normalized
    return ""


def is_special_asset_regional(regional: str) -> bool:
    return normalize_asset_regional(regional) in ASSET_SPECIAL_REGIONAL_OPTIONS


def asset_regional_for_base(base: str) -> str:
    normalized = re.sub(r"\s+", " ", (base or "").strip().upper())
    if "-MG" in normalized or normalized.endswith(" MG"):
        return "MG"
    if "-SP" in normalized or normalized.endswith(" SP"):
        return "SPN"
    return ""


def base_options_for_asset_regional(regional: str = "") -> list[str]:
    """Opções de unidade para Gestão de Ativos.

    Diferente do cadastro de usuários, a gestão de ativos deve listar bases e
    franquias. Mantemos o nome da função por compatibilidade com o restante do
    código, mas a origem correta aqui é BASE_FRANCHISE_OPTIONS.
    """
    normalized = normalize_asset_regional(regional)
    if not normalized:
        return BASE_FRANCHISE_OPTIONS
    if is_special_asset_regional(normalized):
        return []
    return [unit for unit in BASE_FRANCHISE_OPTIONS if asset_regional_for_base(unit) == normalized]


def base_unit_options_for_asset_regional(regional: str = "") -> list[str]:
    normalized = normalize_asset_regional(regional)
    options = BASE_UNIT_OPTIONS
    if normalized and not is_special_asset_regional(normalized):
        options = [unit for unit in options if asset_regional_for_base(unit) == normalized]
    elif is_special_asset_regional(normalized):
        options = []
    return options


def franchise_unit_options_for_asset_regional(regional: str = "") -> list[str]:
    normalized = normalize_asset_regional(regional)
    options = FRANCHISE_UNIT_OPTIONS
    if normalized and not is_special_asset_regional(normalized):
        options = [unit for unit in options if asset_regional_for_base(unit) == normalized]
    elif is_special_asset_regional(normalized):
        options = []
    return options


def validate_unit_selection(base: str, franchise: str, *, required: bool = False) -> tuple[str, str]:
    """Valida seleção separada de base/franquia.

    Retorna (unidade, tipo), onde tipo é "base", "franchise" ou "".
    """
    base = (base or "").strip()
    franchise = (franchise or "").strip()
    if base and franchise:
        raise ValueError("Selecione somente uma base ou uma franquia, não as duas.")
    if required and not base and not franchise:
        raise ValueError("Selecione obrigatoriamente uma base ou uma franquia para gerar o relatório.")
    if base:
        if base not in BASE_UNIT_OPTION_SET:
            raise ValueError("Base selecionada inválida.")
        return base, "base"
    if franchise:
        if franchise not in FRANCHISE_UNIT_OPTION_SET:
            raise ValueError("Franquia selecionada inválida.")
        return franchise, "franchise"
    return "", ""


def unit_kind_label(kind: str) -> str:
    if kind == "all":
        return "Todas as unidades"
    return "Franquia" if kind == "franchise" else "Base" if kind == "base" else "Unidade"


def init_db() -> None:
    with db_connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                responsible_name TEXT NOT NULL,
                organization_name TEXT NOT NULL,
                franchise_name TEXT NOT NULL DEFAULT '',
                franchise_number TEXT NOT NULL DEFAULT '',
                cnpj TEXT NOT NULL DEFAULT '',
                username TEXT NOT NULL UNIQUE,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'base',
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                updated_at TEXT,
                page_permissions_configured INTEGER NOT NULL DEFAULT 0,
                action_permissions_configured INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT,
                category_emoji TEXT,
                image_name TEXT,
                image_key TEXT,
                image_content_type TEXT,
                unit_measure TEXT NOT NULL DEFAULT 'un',
                is_kit INTEGER NOT NULL DEFAULT 0,
                kit_quantity INTEGER NOT NULL DEFAULT 1,
                description TEXT,
                stock_quantity INTEGER NOT NULL DEFAULT 0,
                price_cents INTEGER NOT NULL DEFAULT 0,
                limit_base INTEGER,
                limit_franchise INTEGER,
                limit_block_days INTEGER NOT NULL DEFAULT 60,
                min_order_quantity INTEGER,
                min_stock INTEGER,
                max_stock INTEGER,
                active INTEGER NOT NULL DEFAULT 1,
                catalog_archived INTEGER NOT NULL DEFAULT 0,
                visible_base INTEGER NOT NULL DEFAULT 1,
                visible_franchise INTEGER NOT NULL DEFAULT 1,
                internal INTEGER NOT NULL DEFAULT 0,
                stock_tag TEXT NOT NULL DEFAULT 'insumos',
                created_at TEXT NOT NULL,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS stock_tags (
                slug TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                system_key INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS supply_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                user_note TEXT,
                admin_note TEXT,
                people_count INTEGER,
                created_at TEXT NOT NULL,
                reviewed_at TEXT,
                reviewed_by_id INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(reviewed_by_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS request_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                product_name_snapshot TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                price_cents_snapshot INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(request_id) REFERENCES supply_requests(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS request_action_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                actor_user_id INTEGER,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(request_id) REFERENCES supply_requests(id) ON DELETE CASCADE,
                FOREIGN KEY(actor_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS product_request_blocks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                blocked_until TEXT NOT NULL,
                reason TEXT NOT NULL DEFAULT '',
                created_by_request_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                revoked_at TEXT,
                updated_by_id INTEGER,
                UNIQUE(user_id, product_id),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY(created_by_request_id) REFERENCES supply_requests(id),
                FOREIGN KEY(updated_by_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS admin_login_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                code_hash TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used_at TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                request_id INTEGER,
                created_by_id INTEGER,
                movement_type TEXT NOT NULL,
                quantity_delta INTEGER NOT NULL,
                stock_before INTEGER NOT NULL,
                stock_after INTEGER NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id),
                FOREIGN KEY(request_id) REFERENCES supply_requests(id),
                FOREIGN KEY(created_by_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                base TEXT NOT NULL,
                regional TEXT NOT NULL,
                sector TEXT NOT NULL,
                manager TEXT NOT NULL,
                created_by_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                FOREIGN KEY(created_by_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS asset_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL,
                product_id INTEGER,
                item_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                serial_number TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(asset_id) REFERENCES assets(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS material_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                item_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                unit_measure TEXT NOT NULL DEFAULT 'un',
                unit_price_cents INTEGER NOT NULL DEFAULT 0,
                invoice_file_name TEXT,
                invoice_file_key TEXT,
                invoice_number TEXT,
                invoice_date TEXT,
                invoice_value_cents INTEGER NOT NULL DEFAULT 0,
                notes TEXT,
                created_by_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id),
                FOREIGN KEY(created_by_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS user_page_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                page_key TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, page_key),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS user_action_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                action_key TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, action_key),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS access_role_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role_key TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                permissions_json TEXT NOT NULL DEFAULT '[]',
                action_permissions_json TEXT NOT NULL DEFAULT '[]',
                editable_roles_json TEXT NOT NULL DEFAULT '[]',
                editable_user_fields_json TEXT NOT NULL DEFAULT '[]',
                created_at TEXT NOT NULL,
                updated_at TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_access_role_types_name ON access_role_types(name COLLATE NOCASE);

            CREATE TABLE IF NOT EXISTS request_regional_admin_assignments (
                regional TEXT PRIMARY KEY,
                admin_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                updated_by_id INTEGER,
                FOREIGN KEY(admin_user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(updated_by_id) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_request_regional_admin_user ON request_regional_admin_assignments(admin_user_id);

            CREATE INDEX IF NOT EXISTS idx_users_status_created ON users(status, created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_users_role_status ON users(role, status);
            CREATE INDEX IF NOT EXISTS idx_users_username ON users(username);
            CREATE INDEX IF NOT EXISTS idx_users_responsible_name ON users(responsible_name COLLATE NOCASE);
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '',
                updated_at TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_stock_tags_name ON stock_tags(name COLLATE NOCASE);
            CREATE INDEX IF NOT EXISTS idx_supply_requests_status_created ON supply_requests(status, created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_supply_requests_user_created ON supply_requests(user_id, created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_request_items_request ON request_items(request_id);
            CREATE INDEX IF NOT EXISTS idx_request_action_logs_request ON request_action_logs(request_id, created_at DESC);
            CREATE INDEX IF NOT EXISTS idx_product_request_blocks_user_product ON product_request_blocks(user_id, product_id);
            CREATE INDEX IF NOT EXISTS idx_product_request_blocks_blocked_until ON product_request_blocks(blocked_until);
            CREATE INDEX IF NOT EXISTS idx_stock_movements_product ON stock_movements(product_id);
            """
        )
        user_columns = {row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
        if "page_permissions_configured" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN page_permissions_configured INTEGER NOT NULL DEFAULT 0")
            user_columns.add("page_permissions_configured")
        if "action_permissions_configured" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN action_permissions_configured INTEGER NOT NULL DEFAULT 0")
            user_columns.add("action_permissions_configured")

        access_role_columns = {row["name"] for row in conn.execute("PRAGMA table_info(access_role_types)").fetchall()}
        if "action_permissions_json" not in access_role_columns:
            conn.execute("ALTER TABLE access_role_types ADD COLUMN action_permissions_json TEXT NOT NULL DEFAULT '[]'")
        if "editable_roles_json" not in access_role_columns:
            conn.execute("ALTER TABLE access_role_types ADD COLUMN editable_roles_json TEXT NOT NULL DEFAULT '[]'")
            conn.execute(
                "UPDATE access_role_types SET editable_roles_json = ? WHERE role_key = 'admin' AND editable_roles_json = '[]'",
                (json.dumps(["base", "franchise", "admin"], ensure_ascii=False),),
            )
        if "editable_user_fields_json" not in access_role_columns:
            conn.execute("ALTER TABLE access_role_types ADD COLUMN editable_user_fields_json TEXT NOT NULL DEFAULT '[]'")
        admin_full_action_permissions = json.dumps([
            "products_create", "products_edit_basic", "products_edit_category", "products_edit_unit",
            "products_edit_price", "products_edit_stock", "products_edit_limits", "products_edit_visibility",
            "products_import", "products_export", "products_delete", "stock_material_entries",
            "stock_assets_create", "stock_reports", "requests_edit_items", "requests_approve_reject",
            "requests_delete", "users_create_edit", "users_import_export", "users_status_delete",
        ], ensure_ascii=False)
        admin_full_editable_roles = json.dumps(["base", "franchise", "admin"], ensure_ascii=False)
        admin_default_editable_fields = json.dumps(default_static_user_edit_fields("admin"), ensure_ascii=False)
        conn.execute(
            """
            INSERT INTO access_role_types (role_key, name, description, permissions_json, action_permissions_json, editable_roles_json, editable_user_fields_json, created_at, updated_at)
            VALUES ('admin', 'Administrador', ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(role_key) DO UPDATE SET
                action_permissions_json = excluded.action_permissions_json,
                editable_roles_json = excluded.editable_roles_json,
                updated_at = COALESCE(access_role_types.updated_at, excluded.updated_at)
            """,
            (
                STATIC_ROLE_DESCRIPTIONS.get("admin", "Tipo padrão administrativo."),
                json.dumps(default_static_page_permissions("admin"), ensure_ascii=False),
                admin_full_action_permissions,
                admin_full_editable_roles,
                admin_default_editable_fields,
                now_iso(),
                now_iso(),
            ),
        )
        admin_identity_lock_key = "migration_v203_admin_identity_fields_locked"
        if conn.execute("SELECT 1 FROM app_settings WHERE key = ?", (admin_identity_lock_key,)).fetchone() is None:
            old_admin_full_editable_fields = json.dumps([
                "responsible_name", "username", "password", "role", "status", "organization_name",
                "franchise_name", "franchise_number", "cnpj", "page_permissions",
            ], ensure_ascii=False)
            conn.execute(
                """
                UPDATE access_role_types
                   SET editable_user_fields_json = ?, updated_at = ?
                 WHERE role_key = 'admin'
                   AND (editable_user_fields_json = ? OR editable_user_fields_json = '[]' OR editable_user_fields_json IS NULL)
                """,
                (admin_default_editable_fields, now_iso(), old_admin_full_editable_fields),
            )
            conn.execute(
                "INSERT INTO app_settings (key, value, updated_at) VALUES (?, '1', ?)",
                (admin_identity_lock_key, now_iso()),
            )

        user_columns = {row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
        user_migrations = [
            ("franchise_name", "ALTER TABLE users ADD COLUMN franchise_name TEXT NOT NULL DEFAULT ''"),
            ("franchise_number", "ALTER TABLE users ADD COLUMN franchise_number TEXT NOT NULL DEFAULT ''"),
            ("cnpj", "ALTER TABLE users ADD COLUMN cnpj TEXT NOT NULL DEFAULT ''"),
        ]
        for column_name, migration_sql in user_migrations:
            if column_name not in user_columns:
                conn.execute(migration_sql)
        conn.execute(
            """
            UPDATE users
               SET franchise_name = CASE WHEN franchise_name = '' THEN organization_name ELSE franchise_name END
             WHERE role = 'franchise'
            """
        )

        # Limpa dados antigos inválidos de telefone/CNPJ em franquias.
        # Em versões anteriores, algumas franquias sem telefone acabavam recebendo
        # o nome da unidade no campo de telefone, exibindo valores como "10"
        # e bloqueando o salvamento da edição. Telefone e CNPJ são opcionais.
        legacy_contact_rows = conn.execute(
            "SELECT id, franchise_number, cnpj FROM users WHERE role = 'franchise'"
        ).fetchall()
        for legacy_row in legacy_contact_rows:
            clean_phone = normalize_phone_number(legacy_row["franchise_number"] if "franchise_number" in legacy_row.keys() else "")
            clean_cnpj = normalize_cnpj(legacy_row["cnpj"] if "cnpj" in legacy_row.keys() else "")
            normalized_phone = clean_phone if len(clean_phone) in {10, 11} else ""
            normalized_cnpj = clean_cnpj if len(clean_cnpj) == 14 else ""
            if normalized_phone != (legacy_row["franchise_number"] or "") or normalized_cnpj != (legacy_row["cnpj"] or ""):
                conn.execute(
                    "UPDATE users SET franchise_number = ?, cnpj = ? WHERE id = ?",
                    (normalized_phone, normalized_cnpj, legacy_row["id"]),
                )

        user_columns = {row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
        if "username" not in user_columns:
            conn.execute("ALTER TABLE users ADD COLUMN username TEXT")
            existing_users = conn.execute("SELECT id, responsible_name, email FROM users").fetchall()
            used_usernames: set[str] = set()
            for existing_user in existing_users:
                base_username = normalize_username((existing_user["email"] or "").split("@")[0] or existing_user["responsible_name"] or f"usuario{existing_user['id']}")
                if not base_username:
                    base_username = f"usuario{existing_user['id']}"
                candidate = base_username
                suffix = 2
                while candidate in used_usernames:
                    candidate = f"{base_username}{suffix}"
                    suffix += 1
                used_usernames.add(candidate)
                conn.execute("UPDATE users SET username = ? WHERE id = ?", (candidate, existing_user["id"]))
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_username_unique ON users(username)")

        supply_request_columns = {row["name"] for row in conn.execute("PRAGMA table_info(supply_requests)").fetchall()}
        if "people_count" not in supply_request_columns:
            conn.execute("ALTER TABLE supply_requests ADD COLUMN people_count INTEGER")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS product_request_blocks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                blocked_until TEXT NOT NULL,
                reason TEXT NOT NULL DEFAULT '',
                created_by_request_id INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                revoked_at TEXT,
                updated_by_id INTEGER,
                UNIQUE(user_id, product_id),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY(created_by_request_id) REFERENCES supply_requests(id),
                FOREIGN KEY(updated_by_id) REFERENCES users(id)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_product_request_blocks_user_product ON product_request_blocks(user_id, product_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_product_request_blocks_blocked_until ON product_request_blocks(blocked_until)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS request_regional_admin_assignments (
                regional TEXT PRIMARY KEY,
                admin_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                updated_by_id INTEGER,
                FOREIGN KEY(admin_user_id) REFERENCES users(id) ON DELETE CASCADE,
                FOREIGN KEY(updated_by_id) REFERENCES users(id)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_request_regional_admin_user ON request_regional_admin_assignments(admin_user_id)")

        product_columns = {row["name"] for row in conn.execute("PRAGMA table_info(products)").fetchall()}
        if "category_emoji" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN category_emoji TEXT")
        if "image_name" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN image_name TEXT")
        if "image_key" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN image_key TEXT")
        if "image_content_type" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN image_content_type TEXT")
        if "unit_measure" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN unit_measure TEXT NOT NULL DEFAULT 'un'")
        if "min_order_quantity" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN min_order_quantity INTEGER")
        if "min_stock" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN min_stock INTEGER")
        if "max_stock" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN max_stock INTEGER")
        if "catalog_archived" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN catalog_archived INTEGER NOT NULL DEFAULT 0")
        if "visible_base" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN visible_base INTEGER NOT NULL DEFAULT 1")
        if "visible_franchise" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN visible_franchise INTEGER NOT NULL DEFAULT 1")
        if "internal" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN internal INTEGER NOT NULL DEFAULT 0")
        if "is_kit" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN is_kit INTEGER NOT NULL DEFAULT 0")
        if "kit_quantity" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN kit_quantity INTEGER NOT NULL DEFAULT 1")
        if "stock_tag" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN stock_tag TEXT NOT NULL DEFAULT 'insumos'")
        if "limit_block_days" not in product_columns:
            conn.execute("ALTER TABLE products ADD COLUMN limit_block_days INTEGER NOT NULL DEFAULT 60")
        for stock_tag_slug, stock_tag_name, stock_tag_description in SYSTEM_STOCK_TAGS:
            conn.execute(
                """
                INSERT OR IGNORE INTO stock_tags (slug, name, description, active, system_key, created_at)
                VALUES (?, ?, ?, 1, 1, ?)
                """,
                (stock_tag_slug, stock_tag_name, stock_tag_description, now_iso()),
            )
            conn.execute(
                "UPDATE stock_tags SET system_key = 1, active = 1 WHERE slug = ?",
                (stock_tag_slug,),
            )
        conn.execute("UPDATE products SET stock_tag = ? WHERE stock_tag IS NULL OR TRIM(stock_tag) = ''", (DEFAULT_STOCK_TAG,))
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_catalog_active_name ON products(catalog_archived, active, name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_catalog_category ON products(catalog_archived, category)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_stock ON products(catalog_archived, stock_quantity)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_products_stock_tag ON products(stock_tag, catalog_archived, active)")
        asset_item_columns = {row["name"] for row in conn.execute("PRAGMA table_info(asset_items)").fetchall()}
        if "product_id" not in asset_item_columns:
            conn.execute("ALTER TABLE asset_items ADD COLUMN product_id INTEGER")
        if "quantity" not in asset_item_columns:
            conn.execute("ALTER TABLE asset_items ADD COLUMN quantity INTEGER NOT NULL DEFAULT 1")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_assets_base_regional ON assets(base, regional)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_asset_items_asset_id ON asset_items(asset_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_asset_items_product_id ON asset_items(product_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_material_entries_created_at ON material_entries(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_material_entries_product_id ON material_entries(product_id)")
        movement_count = conn.execute("SELECT COUNT(*) AS total FROM stock_movements").fetchone()["total"]
        if int(movement_count or 0) == 0:
            existing_products = conn.execute("SELECT id, stock_quantity FROM products WHERE stock_quantity > 0").fetchall()
            for product_row in existing_products:
                initial_stock = int(product_row["stock_quantity"] or 0)
                record_stock_movement(
                    conn,
                    int(product_row["id"]),
                    initial_stock,
                    0,
                    initial_stock,
                    "product_created",
                    "Registro inicial do estoque existente.",
                )
        conn.commit()


def seed_initial_data() -> None:
    admin_username = normalize_username(os.getenv("ADMIN_USERNAME", os.getenv("SEED_ADMIN_USERNAME", "admin")))
    admin_email = os.getenv("ADMIN_EMAIL", os.getenv("SEED_ADMIN_EMAIL", synthetic_email_for_username(admin_username))).strip().lower()
    admin_password = os.getenv("ADMIN_PASSWORD", os.getenv("SEED_ADMIN_PASSWORD", "Admin@123"))
    with db_connect() as conn:
        user_count = conn.execute("SELECT COUNT(*) AS total FROM users").fetchone()["total"]
        if user_count == 0:
            conn.execute(
                """
                INSERT INTO users (responsible_name, organization_name, username, email, password_hash, role, status, created_at)
                VALUES (?, ?, ?, ?, ?, 'admin', 'approved', ?)
                """,
                ("Administrador", "J&T Express", admin_username, admin_email, generate_password_hash(admin_password), now_iso()),
            )

        product_count = conn.execute("SELECT COUNT(*) AS total FROM products").fetchone()["total"]
        if product_count == 0:
            samples = [
                ("Envelope de segurança P", "Embalagens", "un", "Envelope pequeno para envios leves.", 500, 45, 100, 80, 120, 600),
                ("Envelope de segurança M", "Embalagens", "un", "Envelope médio para envios padrão.", 400, 65, 80, 60, 100, 500),
                ("Lacre plástico", "Operacional", "un", "Lacre numerado para controle interno.", 1000, 18, 200, 150, 250, 1200),
                ("Etiqueta térmica", "Etiquetas", "rolo", "Rolo de etiqueta para impressora térmica.", 120, 2500, 10, 8, 30, 180),
            ]
            for item in samples:
                cursor = conn.execute(
                    """
                    INSERT INTO products (name, category, unit_measure, description, stock_quantity, price_cents, limit_base, limit_franchise, min_stock, max_stock, active, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (*item, now_iso()),
                )
                row_id = get_cursor_lastrowid(cursor)
                if row_id is not None:
                    record_stock_movement(conn, row_id, int(item[4]), 0, int(item[4]), "product_created", "Estoque inicial do sistema.")
        conn.commit()


def setup_database() -> None:
    init_db()
    seed_initial_data()


# ---------- Consultas ----------

def get_user(user_id: int | None) -> User | None:
    if user_id is None:
        return None
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (int(user_id),)).fetchone()
    return row_to_user(row)


def get_user_by_username(username: str) -> User | None:
    username = normalize_username(username)
    if not username:
        return None
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM users WHERE lower(username) = lower(?)", (username,)).fetchone()
    return row_to_user(row)


def get_user_by_email(email: str) -> User | None:
    # Mantido apenas para compatibilidade com bancos antigos/rotas antigas.
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM users WHERE lower(email) = lower(?)", ((email or "").strip().lower(),)).fetchone()
    return row_to_user(row)


def get_product(product_id: int | None) -> Product | None:
    if product_id is None:
        return None
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM products WHERE id = ?", (int(product_id),)).fetchone()
    return row_to_product(row)




def get_product_by_name(name: str) -> Product | None:
    with db_connect() as conn:
        row = conn.execute(
            "SELECT * FROM products WHERE catalog_archived = 0 AND lower(name) = lower(?) LIMIT 1",
            (name.strip(),),
        ).fetchone()
    return row_to_product(row)


def get_request_items(request_id: int) -> list[RequestItem]:
    with db_connect() as conn:
        rows = conn.execute("SELECT * FROM request_items WHERE request_id = ? ORDER BY id", (request_id,)).fetchall()
    return [item for row in rows if (item := row_to_item(row)) is not None]


def get_supply_request(request_id: int) -> SupplyRequest | None:
    with db_connect() as conn:
        row = conn.execute("SELECT * FROM supply_requests WHERE id = ?", (request_id,)).fetchone()
    return row_to_supply_request(row, include_actions=True)


def request_action_label(action: str) -> str:
    labels = {
        "created": "Criada",
        "items_updated": "Itens editados",
        "approved": "Aprovada",
        "rejected": "Recusada",
        "deleted": "Excluída",
    }
    return labels.get((action or "").strip().lower(), action or "Ação")


def request_action_actor_name(actor: User | None, fallback: str = "Sistema") -> str:
    if actor is None:
        return fallback
    return actor.responsible_name or actor.username or fallback


def request_action_log_to_dict(row: Any) -> dict[str, Any]:
    actor_name = (row["actor_name"] if "actor_name" in row.keys() else "") or (row["actor_username"] if "actor_username" in row.keys() else "") or "Sistema"
    created_at = parse_dt(row["created_at"]) or datetime.utcnow()
    action = (row["action"] or "").strip().lower()
    return {
        "id": int(row["id"]),
        "request_id": int(row["request_id"]),
        "action": action,
        "action_label": request_action_label(action),
        "actor_user_id": int(row["actor_user_id"]) if row["actor_user_id"] is not None else None,
        "actor_name": actor_name,
        "note": row["note"] or "",
        "created_at": created_at,
    }


def list_request_action_logs(request_id: int) -> list[dict[str, Any]]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT ral.*, u.responsible_name AS actor_name, u.username AS actor_username
              FROM request_action_logs ral
              LEFT JOIN users u ON u.id = ral.actor_user_id
             WHERE ral.request_id = ?
             ORDER BY ral.created_at ASC, ral.id ASC
            """,
            (request_id,),
        ).fetchall()
    return [request_action_log_to_dict(row) for row in rows]


def record_request_action(conn: Any, request_id: int, action: str, actor_user_id: int | None = None, note: str = "") -> None:
    conn.execute(
        """
        INSERT INTO request_action_logs (request_id, action, actor_user_id, note, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (int(request_id), (action or "").strip().lower(), actor_user_id, compact_text(note, fallback="", limit=900), now_iso()),
    )


def request_action_summary(supply_request: SupplyRequest | None) -> str:
    if supply_request is None:
        return "-"
    terminal_actions = {"approved", "rejected", "deleted"}
    action_log = None
    for log in reversed(supply_request.action_logs or []):
        if log.get("action") in terminal_actions:
            action_log = log
            break
    if action_log is not None:
        date_text = format_sao_paulo_datetime(action_log.get("created_at"))
        return f"{action_log.get('action_label', 'Ação')} por {action_log.get('actor_name') or 'Sistema'} em {date_text}"
    if supply_request.status in terminal_actions:
        label = {"approved": "Aprovada", "rejected": "Recusada", "deleted": "Excluída"}.get(supply_request.status, status_label(supply_request.status))
        actor = request_action_actor_name(supply_request.reviewed_by, "Não registrado")
        if supply_request.reviewed_at:
            return f"{label} por {actor} em {format_sao_paulo_datetime(supply_request.reviewed_at)}"
        return f"{label} por {actor}"
    return "Aguardando análise"


def chunked_ids(values: list[int], chunk_size: int = 80) -> list[list[int]]:
    unique_values: list[int] = []
    seen: set[int] = set()
    for value in values:
        try:
            int_value = int(value)
        except (TypeError, ValueError):
            continue
        if int_value not in seen:
            seen.add(int_value)
            unique_values.append(int_value)
    return [unique_values[index:index + chunk_size] for index in range(0, len(unique_values), chunk_size)]


def execute_delete_ids_chunked(conn: Any, table: str, column: str, ids: list[int]) -> int:
    """Apaga IDs em blocos para funcionar bem no SQLite local e no Cloudflare D1."""
    total = 0
    allowed_tables = {"supply_requests", "request_items", "request_action_logs", "stock_movements", "admin_login_codes", "user_page_permissions", "user_action_permissions"}
    allowed_columns = {"id", "request_id", "product_id", "user_id", "created_by_id"}
    if table not in allowed_tables or column not in allowed_columns:
        raise ValueError("Tabela/coluna não autorizada para exclusão em bloco.")
    for chunk in chunked_ids(ids):
        placeholders = ",".join("?" for _ in chunk)
        cursor = conn.execute(f"DELETE FROM {table} WHERE {column} IN ({placeholders})", chunk)
        if isinstance(getattr(cursor, "rowcount", None), int) and cursor.rowcount > 0:
            total += int(cursor.rowcount)
    return total


def permanently_delete_supply_request(conn: Any, request_id: int) -> bool:
    """Remove uma solicitação e vínculos diretos do banco de dados."""
    row = conn.execute("SELECT id FROM supply_requests WHERE id = ?", (request_id,)).fetchone()
    if row is None:
        return False
    conn.execute("DELETE FROM product_request_blocks WHERE created_by_request_id = ?", (request_id,))
    conn.execute("DELETE FROM stock_movements WHERE request_id = ?", (request_id,))
    conn.execute("DELETE FROM request_action_logs WHERE request_id = ?", (request_id,))
    conn.execute("DELETE FROM request_items WHERE request_id = ?", (request_id,))
    conn.execute("DELETE FROM supply_requests WHERE id = ?", (request_id,))
    return True


def permanently_delete_empty_supply_requests(conn: Any, request_ids: list[int]) -> int:
    """Remove solicitações que ficaram sem itens depois da exclusão de produto."""
    empty_ids: list[int] = []
    for chunk in chunked_ids(request_ids):
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT sr.id
              FROM supply_requests sr
             WHERE sr.id IN ({placeholders})
               AND NOT EXISTS (SELECT 1 FROM request_items ri WHERE ri.request_id = sr.id)
            """,
            chunk,
        ).fetchall()
        empty_ids.extend(int(row["id"]) for row in rows)
    if not empty_ids:
        return 0
    execute_delete_ids_chunked(conn, "stock_movements", "request_id", empty_ids)
    execute_delete_ids_chunked(conn, "request_action_logs", "request_id", empty_ids)
    return execute_delete_ids_chunked(conn, "supply_requests", "id", empty_ids)


def permanently_delete_product(conn: Any, product_id: int) -> tuple[str, int, int]:
    """Remove definitivamente um produto, seus vínculos e movimentos de estoque.

    Retorna: (nome do produto, itens de solicitação removidos, solicitações vazias removidas).
    """
    product_row = conn.execute("SELECT id, name, image_key FROM products WHERE id = ?", (product_id,)).fetchone()
    if product_row is None:
        raise LookupError("Produto não encontrado.")

    product_name = product_row["name"] or f"Produto #{product_id}"
    request_rows = conn.execute("SELECT DISTINCT request_id FROM request_items WHERE product_id = ?", (product_id,)).fetchall()
    request_ids = [int(row["request_id"]) for row in request_rows]
    item_count_row = conn.execute("SELECT COUNT(*) AS total FROM request_items WHERE product_id = ?", (product_id,)).fetchone()
    removed_items = int(item_count_row["total"] or 0) if item_count_row is not None else 0

    conn.execute("DELETE FROM stock_movements WHERE product_id = ?", (product_id,))
    conn.execute("DELETE FROM request_items WHERE product_id = ?", (product_id,))
    conn.execute("DELETE FROM product_request_blocks WHERE product_id = ?", (product_id,))
    removed_empty_requests = permanently_delete_empty_supply_requests(conn, request_ids)
    conn.execute("UPDATE asset_items SET product_id = NULL WHERE product_id = ?", (product_id,))
    conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
    remove_local_product_image(product_row["image_key"] or "")
    return product_name, removed_items, removed_empty_requests


def permanently_delete_user(conn: Any, user_id: int) -> tuple[int, int]:
    """Remove definitivamente um usuário e as solicitações feitas por ele."""
    row = conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if row is None:
        raise LookupError("Usuário não encontrado.")

    request_rows = conn.execute("SELECT id FROM supply_requests WHERE user_id = ?", (user_id,)).fetchall()
    request_ids = [int(row["id"]) for row in request_rows]
    if request_ids:
        execute_delete_ids_chunked(conn, "stock_movements", "request_id", request_ids)
        execute_delete_ids_chunked(conn, "request_items", "request_id", request_ids)
        execute_delete_ids_chunked(conn, "supply_requests", "id", request_ids)

    conn.execute("DELETE FROM product_request_blocks WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM admin_login_codes WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM user_page_permissions WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM user_action_permissions WHERE user_id = ?", (user_id,))
    conn.execute("UPDATE assets SET created_by_id = NULL WHERE created_by_id = ?", (user_id,))
    conn.execute("UPDATE stock_movements SET created_by_id = NULL WHERE created_by_id = ?", (user_id,))
    conn.execute("UPDATE supply_requests SET reviewed_by_id = NULL WHERE reviewed_by_id = ?", (user_id,))
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    return len(request_ids), user_id



REQUEST_REGIONAL_OPTIONS = [
    {"value": "MG", "label": "MG"},
    {"value": "SP", "label": "SP"},
]
REQUEST_REGIONAL_VALUES = {item["value"] for item in REQUEST_REGIONAL_OPTIONS}
REQUEST_SORT_OPTIONS = [
    {"value": "newest", "label": "Mais recentes"},
    {"value": "oldest", "label": "Mais antigas"},
    {"value": "unit_asc", "label": "Unidade A-Z"},
    {"value": "unit_desc", "label": "Unidade Z-A"},
    {"value": "quantity_desc", "label": "Maior quantidade de materiais"},
    {"value": "quantity_asc", "label": "Menor quantidade de materiais"},
    {"value": "items_desc", "label": "Mais tipos de produtos"},
    {"value": "items_asc", "label": "Menos tipos de produtos"},
]
REQUEST_SORT_VALUES = {item["value"] for item in REQUEST_SORT_OPTIONS}
REQUEST_TYPE_OPTIONS = [
    {"value": "base", "label": "Base"},
    {"value": "franchise", "label": "Franquia"},
]
REQUEST_TYPE_VALUES = {item["value"] for item in REQUEST_TYPE_OPTIONS}
REQUEST_REGION_SQL = """
CASE
  WHEN UPPER(COALESCE(u.organization_name, '') || ' ' || COALESCE(u.franchise_name, '')) LIKE '%-MG%'
    OR UPPER(COALESCE(u.organization_name, '') || ' ' || COALESCE(u.franchise_name, '')) LIKE '% MG'
    THEN 'MG'
  WHEN UPPER(COALESCE(u.organization_name, '') || ' ' || COALESCE(u.franchise_name, '')) LIKE '%-SP%'
    OR UPPER(COALESCE(u.organization_name, '') || ' ' || COALESCE(u.franchise_name, '')) LIKE '% SP'
    THEN 'SP'
  ELSE ''
END
"""
REQUEST_UNIT_SQL = "COALESCE(NULLIF(u.organization_name, ''), NULLIF(u.franchise_name, ''), u.responsible_name, '')"
REQUEST_TOTAL_QUANTITY_SQL = "COALESCE((SELECT SUM(riq.quantity) FROM request_items riq WHERE riq.request_id = sr.id), 0)"
REQUEST_ITEM_COUNT_SQL = "COALESCE((SELECT COUNT(*) FROM request_items ric WHERE ric.request_id = sr.id), 0)"


def normalize_request_regional(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if raw in {"SP", "SPN", "SAO PAULO", "SÃO PAULO"}:
        return "SP"
    if raw in {"MG", "MINAS", "MINAS GERAIS"}:
        return "MG"
    return ""


def request_regional_label(value: str) -> str:
    regional = normalize_request_regional(value)
    return regional if regional else "Sem regional"


def request_regional_for_unit_name(unit_name: Any) -> str:
    regional = asset_regional_for_base(str(unit_name or ""))
    if regional == "SPN":
        return "SP"
    return regional if regional == "MG" else ""


def request_regional_for_user(user: User | None) -> str:
    if user is None:
        return ""
    return request_regional_for_unit_name(user.organization_name or user.franchise_name)


def normalize_request_filters_from_args(args: Any | None = None) -> dict[str, Any]:
    source = args if args is not None else request.args
    type_filter = (source.get("type", "") if hasattr(source, "get") else "") or ""
    type_filter = type_filter.strip().lower()
    if type_filter not in REQUEST_TYPE_VALUES:
        type_filter = ""

    regional_filter = normalize_request_regional(source.get("regional", "") if hasattr(source, "get") else "")

    product_id: int | None = None
    raw_product = (source.get("product_id", "") if hasattr(source, "get") else "") or ""
    try:
        parsed_product = int(raw_product)
        product_id = parsed_product if parsed_product > 0 else None
    except (TypeError, ValueError):
        product_id = None

    sort_filter = (source.get("sort", "newest") if hasattr(source, "get") else "newest") or "newest"
    sort_filter = str(sort_filter).strip().lower()
    if sort_filter not in REQUEST_SORT_VALUES:
        sort_filter = "newest"

    return {
        "type": type_filter,
        "regional": regional_filter,
        "product_id": product_id,
        "sort": sort_filter,
    }


def request_filters_to_query_args(filters: dict[str, Any] | None, *, include_empty: bool = False) -> dict[str, Any]:
    filters = filters or {}
    args: dict[str, Any] = {}
    if include_empty or filters.get("type"):
        args["type"] = filters.get("type", "") or ""
    if include_empty or filters.get("regional"):
        args["regional"] = filters.get("regional", "") or ""
    if include_empty or filters.get("product_id"):
        product_id = filters.get("product_id")
        args["product_id"] = int(product_id) if product_id else ""
    if include_empty or filters.get("sort", "newest") != "newest":
        args["sort"] = filters.get("sort", "newest") or "newest"
    return args


def list_products_for_request_filters(limit: int = 1000) -> list[Product]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT p.*
              FROM products p
              JOIN request_items ri ON ri.product_id = p.id
             ORDER BY p.name COLLATE NOCASE ASC
             LIMIT ?
            """,
            (bounded_int(limit, 1000, 50, 2000),),
        ).fetchall()
    return [product for row in rows if (product := row_to_product(row)) is not None]


def request_admin_assignment_admin_options() -> list[User]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT *
              FROM users
             WHERE status = 'approved'
               AND role <> 'dev'
             ORDER BY responsible_name COLLATE NOCASE ASC, organization_name COLLATE NOCASE ASC
            """
        ).fetchall()
    users = [user for row in rows if (user := row_to_user(row)) is not None]
    return [user for user in users if user.is_admin]


def list_request_regional_admin_assignments(conn: Any | None = None) -> dict[str, int]:
    sql = "SELECT regional, admin_user_id FROM request_regional_admin_assignments"
    if conn is not None:
        rows = conn.execute(sql).fetchall()
    else:
        with db_connect() as local_conn:
            rows = local_conn.execute(sql).fetchall()
    assignments: dict[str, int] = {}
    for row in rows:
        regional = normalize_request_regional(row["regional"])
        if regional:
            assignments[regional] = int(row["admin_user_id"])
    return assignments


def list_request_regional_admin_assignment_details() -> dict[str, dict[str, Any]]:
    assignments = list_request_regional_admin_assignments()
    details: dict[str, dict[str, Any]] = {}
    for option in REQUEST_REGIONAL_OPTIONS:
        regional = option["value"]
        admin_id = assignments.get(regional)
        details[regional] = {"regional": regional, "label": option["label"], "admin_id": admin_id, "admin": get_user(admin_id) if admin_id else None}
    return details


def set_request_regional_admin_assignment(conn: Any, regional: str, admin_user_id: int | None, updated_by_id: int | None = None) -> None:
    normalized = normalize_request_regional(regional)
    if normalized not in REQUEST_REGIONAL_VALUES:
        raise ValueError("Regional inválida.")
    if admin_user_id is None:
        conn.execute("DELETE FROM request_regional_admin_assignments WHERE regional = ?", (normalized,))
        return
    admin_row = conn.execute("SELECT * FROM users WHERE id = ?", (int(admin_user_id),)).fetchone()
    admin = row_to_user(admin_row)
    if admin is None or admin.status != "approved" or admin.is_dev or not admin.is_admin:
        raise ValueError("Selecione um Admin aprovado para receber os pedidos da regional.")
    existing = conn.execute("SELECT regional FROM request_regional_admin_assignments WHERE regional = ?", (normalized,)).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE request_regional_admin_assignments
               SET admin_user_id = ?, updated_at = ?, updated_by_id = ?
             WHERE regional = ?
            """,
            (int(admin_user_id), now_iso(), updated_by_id, normalized),
        )
    else:
        conn.execute(
            """
            INSERT INTO request_regional_admin_assignments (regional, admin_user_id, created_at, updated_at, updated_by_id)
            VALUES (?, ?, ?, ?, ?)
            """,
            (normalized, int(admin_user_id), now_iso(), now_iso(), updated_by_id),
        )


def request_assignment_visibility_condition(viewer: User | None) -> tuple[str, list[Any]]:
    if viewer is None or viewer.is_dev:
        return "", []
    return f"""
      AND (
            ({REQUEST_REGION_SQL}) = ''
         OR NOT EXISTS (
              SELECT 1
                FROM request_regional_admin_assignments raa_any
               WHERE raa_any.regional = ({REQUEST_REGION_SQL})
            )
         OR EXISTS (
              SELECT 1
                FROM request_regional_admin_assignments raa_viewer
               WHERE raa_viewer.regional = ({REQUEST_REGION_SQL})
                 AND raa_viewer.admin_user_id = ?
            )
      )
    """, [viewer.id]


def can_view_supply_request_by_assignment(supply_request: SupplyRequest | None, viewer: User | None = None) -> bool:
    if supply_request is None:
        return False
    viewer = viewer or current_user()
    if viewer is None or viewer.is_dev:
        return True
    requester = supply_request.user or get_user(supply_request.user_id)
    regional = request_regional_for_user(requester)
    if not regional:
        return True
    assignments = list_request_regional_admin_assignments()
    assigned_admin_id = assignments.get(regional)
    return assigned_admin_id is None or assigned_admin_id == viewer.id


def request_list_query_parts(status: str = "", user_id: int | None = None, filters: dict[str, Any] | None = None, viewer: User | None = None, apply_assignment_visibility: bool = False) -> tuple[str, list[Any], str]:
    filters = filters or {}
    clauses: list[str] = []
    params: list[Any] = []
    status = (status or "").strip().lower()
    if status:
        clauses.append("sr.status = ?")
        params.append(status)
    if user_id is not None:
        clauses.append("sr.user_id = ?")
        params.append(int(user_id))

    type_filter = (filters.get("type") or "").strip().lower()
    if type_filter in REQUEST_TYPE_VALUES:
        clauses.append("u.role = ?")
        params.append(type_filter)

    regional_filter = normalize_request_regional(filters.get("regional", ""))
    if regional_filter:
        clauses.append(f"({REQUEST_REGION_SQL}) = ?")
        params.append(regional_filter)

    product_id = filters.get("product_id")
    if product_id:
        clauses.append("EXISTS (SELECT 1 FROM request_items rif WHERE rif.request_id = sr.id AND rif.product_id = ?)")
        params.append(int(product_id))

    if apply_assignment_visibility:
        visibility_sql, visibility_params = request_assignment_visibility_condition(viewer)
        if visibility_sql:
            clauses.append(visibility_sql.strip()[4:].strip() if visibility_sql.strip().upper().startswith("AND ") else visibility_sql.strip())
            params.extend(visibility_params)

    where_sql = (" WHERE " + " AND ".join(f"({clause})" for clause in clauses)) if clauses else ""
    sort_filter = (filters.get("sort") or "newest").strip().lower()
    sort_map = {
        "newest": "sr.created_at DESC, sr.id DESC",
        "oldest": "sr.created_at ASC, sr.id ASC",
        "unit_asc": f"{REQUEST_UNIT_SQL} COLLATE NOCASE ASC, sr.created_at DESC, sr.id DESC",
        "unit_desc": f"{REQUEST_UNIT_SQL} COLLATE NOCASE DESC, sr.created_at DESC, sr.id DESC",
        "quantity_desc": f"{REQUEST_TOTAL_QUANTITY_SQL} DESC, sr.created_at DESC, sr.id DESC",
        "quantity_asc": f"{REQUEST_TOTAL_QUANTITY_SQL} ASC, sr.created_at DESC, sr.id DESC",
        "items_desc": f"{REQUEST_ITEM_COUNT_SQL} DESC, sr.created_at DESC, sr.id DESC",
        "items_asc": f"{REQUEST_ITEM_COUNT_SQL} ASC, sr.created_at DESC, sr.id DESC",
    }
    order_sql = sort_map.get(sort_filter, sort_map["newest"])
    return where_sql, params, order_sql


def count_supply_requests(status: str = "", user_id: int | None = None, filters: dict[str, Any] | None = None, viewer: User | None = None, apply_assignment_visibility: bool = False) -> int:
    where_sql, params, _order_sql = request_list_query_parts(status, user_id, filters, viewer, apply_assignment_visibility)
    with db_connect() as conn:
        row = conn.execute(f"SELECT COUNT(*) AS total FROM supply_requests sr JOIN users u ON u.id = sr.user_id {where_sql}", params).fetchone()
    return int((row["total"] if row else 0) or 0)


def list_supply_requests(status: str = "", user_id: int | None = None, limit: int | None = None, filters: dict[str, Any] | None = None, viewer: User | None = None, apply_assignment_visibility: bool = False) -> list[SupplyRequest]:
    if limit is None and low_row_read_mode():
        limit = bounded_int(os.getenv("D1_REQUEST_LIST_LIMIT"), DEFAULT_TABLE_PAGE_SIZE, DEFAULT_TABLE_PAGE_SIZE, 300)
    where_sql, params, order_sql = request_list_query_parts(status, user_id, filters, viewer, apply_assignment_visibility)
    sql = f"""
        SELECT sr.id, sr.user_id, sr.status, sr.user_note, sr.admin_note, sr.people_count,
               sr.created_at, sr.reviewed_at, sr.reviewed_by_id
          FROM supply_requests sr
          JOIN users u ON u.id = sr.user_id
          {where_sql}
         ORDER BY {order_sql}
    """
    if limit is not None:
        sql += " LIMIT ?"
        params.append(limit)
    with db_connect() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [req for row in rows if (req := row_to_supply_request(row)) is not None]


def build_pagination(endpoint: str, page: int, per_page: int, total: int, shown_count: int, extra_args: dict[str, Any] | None = None) -> dict[str, Any]:
    total_pages = max(1, (int(total or 0) + per_page - 1) // per_page)
    page = max(1, min(int(page or 1), total_pages))
    offset = (page - 1) * per_page
    extra_args = dict(extra_args or {})

    page_size_options = list(TABLE_PAGE_SIZE_OPTIONS)
    if per_page not in page_size_options:
        page_size_options.append(per_page)
        page_size_options.sort()

    def page_url(target_page: int, target_limit: int | None = None) -> str:
        args = dict(extra_args)
        args["page"] = max(1, int(target_page or 1))
        args["limit"] = target_limit or per_page
        return url_for(endpoint, **args)

    visible_page_numbers: set[int] = {1, total_pages}
    for number in range(page - 2, page + 3):
        if 1 <= number <= total_pages:
            visible_page_numbers.add(number)
    page_links: list[dict[str, Any]] = []
    previous_number = 0
    for number in sorted(visible_page_numbers):
        if previous_number and number - previous_number > 1:
            page_links.append({"ellipsis": True})
        page_links.append({"number": number, "active": number == page, "url": page_url(number)})
        previous_number = number

    start_item = offset + 1 if total else 0
    end_item = min(offset + int(shown_count or 0), int(total or 0))
    return {
        "page": page,
        "limit": per_page,
        "total": int(total or 0),
        "total_pages": total_pages,
        "start": start_item,
        "end": end_item,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "first_url": page_url(1),
        "prev_url": page_url(max(1, page - 1)),
        "next_url": page_url(min(total_pages, page + 1)),
        "last_url": page_url(total_pages),
        "page_links": page_links,
        "page_size_options": page_size_options,
    }


def list_supply_requests_page(status: str = "", user_id: int | None = None, page: int | None = None, limit: int | None = None, endpoint: str = "my_requests", extra_args: dict[str, Any] | None = None, filters: dict[str, Any] | None = None, viewer: User | None = None, apply_assignment_visibility: bool = False) -> tuple[list[SupplyRequest], dict[str, Any]]:
    per_page = limit if limit is not None else list_page_limit(default=DEFAULT_TABLE_PAGE_SIZE, maximum=500)
    current_page = bounded_int(page if page is not None else request.args.get("page"), 1, 1, 100000)
    where_sql, params, order_sql = request_list_query_parts(status, user_id, filters, viewer, apply_assignment_visibility)

    with db_connect() as conn:
        total_row = conn.execute(f"SELECT COUNT(*) AS total FROM supply_requests sr JOIN users u ON u.id = sr.user_id {where_sql}", params).fetchone()
        total_requests = int((total_row["total"] if total_row else 0) or 0)
        total_pages = max(1, (total_requests + per_page - 1) // per_page)
        if current_page > total_pages:
            current_page = total_pages
        offset = (current_page - 1) * per_page
        rows = conn.execute(
            f"""
            SELECT sr.id, sr.user_id, sr.status, sr.user_note, sr.admin_note, sr.people_count,
                   sr.created_at, sr.reviewed_at, sr.reviewed_by_id
              FROM supply_requests sr
              JOIN users u ON u.id = sr.user_id
              {where_sql}
             ORDER BY {order_sql}
             LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

    requests_list = [req for row in rows if (req := row_to_supply_request(row)) is not None]
    pagination = build_pagination(
        endpoint=endpoint,
        page=current_page,
        per_page=per_page,
        total=total_requests,
        shown_count=len(requests_list),
        extra_args=extra_args,
    )
    return requests_list, pagination


# ---------- Helpers ----------

def dev_user_exists(conn: Any | None = None) -> bool:
    sql = "SELECT 1 FROM users WHERE role = 'dev' AND status = 'approved' LIMIT 1"
    if conn is not None:
        return conn.execute(sql).fetchone() is not None
    with db_connect() as local_conn:
        return local_conn.execute(sql).fetchone() is not None

def get_dev_access_password(conn: Any | None = None) -> str:
    sql = "SELECT value FROM app_settings WHERE key = ?"
    try:
        if conn is not None:
            row = conn.execute(sql, (DEV_ACCESS_PASSWORD_KEY,)).fetchone()
        else:
            with db_connect() as local_conn:
                row = local_conn.execute(sql, (DEV_ACCESS_PASSWORD_KEY,)).fetchone()
        value = (row["value"] if row is not None and "value" in row.keys() else "") if row is not None else ""
        return str(value or DEFAULT_DEV_ACCESS_PASSWORD)
    except Exception:
        return DEFAULT_DEV_ACCESS_PASSWORD

def set_dev_access_password(conn: Any, value: str) -> None:
    clean = str(value or "").strip()
    if not clean:
        return
    conn.execute(
        """
        INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
        """,
        (DEV_ACCESS_PASSWORD_KEY, clean, now_iso()),
    )

def can_manage_dev_roles(user: User | None) -> bool:
    return bool(user and user.is_dev and user.is_approved)

def can_bootstrap_dev_role(user: User | None) -> bool:
    return bool(user and user.is_admin and user.is_approved and not dev_user_exists())

def dev_password_is_valid(value: Any, conn: Any | None = None) -> bool:
    return str(value or "").strip() == get_dev_access_password(conn)

def safe_local_redirect_target(value: Any, fallback_endpoint: str, **fallback_values: Any) -> str:
    fallback = url_for(fallback_endpoint, **fallback_values)
    target = str(value or "").strip()
    if not target:
        return fallback
    if target.startswith("//") or re.match(r"^[A-Za-z][A-Za-z0-9+.-]*:", target):
        return fallback
    if not target.startswith("/"):
        return fallback
    return target

def return_target(default_endpoint: str, **fallback_values: Any) -> str:
    target = request.form.get("return_to") or request.args.get("return_to")
    return safe_local_redirect_target(target, default_endpoint, **fallback_values)

def redirect_to_return(default_endpoint: str, **fallback_values: Any) -> Any:
    return redirect(return_target(default_endpoint, **fallback_values))

def can_change_admin_role(current: User | None, target: User | None = None) -> bool:
    if can_manage_dev_roles(current):
        return True
    if target is None:
        return False
    return target.role != "dev" and target.role in get_user_editable_roles(current)

def can_assign_dev_role(current: User | None, supplied_password: Any, conn: Any | None = None) -> bool:
    if can_manage_dev_roles(current):
        return True
    if not dev_password_is_valid(supplied_password, conn):
        return False
    return bool(current and current.is_admin and current.is_approved and not dev_user_exists(conn))

def allowed_role_options_for_editor(current: User | None, target: User | None = None) -> list[str]:
    if can_manage_dev_roles(current):
        options = all_role_options()
    else:
        options = [role for role in all_role_options() if role in get_user_editable_roles(current) and role != "dev"]
    if (can_manage_dev_roles(current) or not dev_user_exists()) and "dev" not in options:
        options.append("dev")
    # Mantém o tipo atual visível mesmo que o editor não possa escolher outros personalizados.
    if target is not None and target.role not in options:
        options.append(target.role)
    return list(dict.fromkeys(options))

def safe_role_for_update(current: User, target: User | None, requested_role: str, supplied_dev_password: Any, conn: Any | None = None) -> tuple[str, str | None]:
    normalized_requested = normalize_user_role(requested_role, allow_admin=True)
    requested = normalized_requested or (target.role if target is not None else "base")

    # Permite criar o primeiro Dev editando o próprio usuário admin, desde que a senha Dev esteja correta.
    # Fora desse bootstrap, o próprio usuário não consegue rebaixar/remover seu acesso por segurança.
    if target is not None and target.id == current.id:
        if requested == "dev" and target.role != "dev":
            if can_assign_dev_role(current, supplied_dev_password, conn):
                return "dev", None
            return current.role, "Para aplicar o acesso Dev, informe a senha Dev correta."
        return current.role, None

    if target is not None and target.is_admin and not can_change_admin_role(current, target):
        return target.role, "Somente usuários Dev podem alterar o tipo de acesso de administradores."
    if requested == "dev" and not can_assign_dev_role(current, supplied_dev_password, conn):
        fallback = target.role if target is not None else "admin"
        return fallback, "Para aplicar o acesso Dev, informe a senha Dev correta. Apenas Dev pode promover outros usuários quando já existir Dev aprovado."
    if not can_manage_dev_roles(current) and requested not in get_user_editable_roles(current):
        fallback = target.role if target is not None else "base"
        return fallback, "Seu tipo de acesso nao pode aplicar esse cargo."
    if requested not in STATIC_ROLE_KEYS and not can_manage_dev_roles(current) and requested not in get_user_editable_roles(current):
        fallback = target.role if target is not None else "base"
        return fallback, "Somente usuários Dev podem aplicar tipos de acesso personalizados."
    return requested, None

def can_edit_dev_password(current: User | None) -> bool:
    return can_manage_dev_roles(current)

def maybe_update_dev_password_from_form(conn: Any, current: User, form: Any, target: User | None = None) -> str | None:
    new_password = str(form.get("new_dev_password", "") or "").strip()
    if not new_password:
        return None
    if not can_edit_dev_password(current):
        return "Somente usuários Dev podem alterar a senha de autorização Dev."
    if target is None or target.id != current.id or not current.is_dev:
        return "A senha de autorização Dev só pode ser alterada pelo próprio usuário Dev."
    set_dev_access_password(conn, new_password)
    return None

def current_is_dev() -> bool:
    return bool(current_user() and current_user().is_dev)

def current_can_manage_dev_roles() -> bool:
    return can_manage_dev_roles(current_user())

def current_can_bootstrap_dev_role() -> bool:
    return can_bootstrap_dev_role(current_user())

def current_can_edit_dev_password() -> bool:
    return can_edit_dev_password(current_user())

def current_allowed_role_options(target: User | None = None) -> list[str]:
    return allowed_role_options_for_editor(current_user(), target)

def current_can_edit_admin_role(target: User | None = None) -> bool:
    return can_change_admin_role(current_user(), target)

def current_user() -> User | None:
    if has_request_context() and hasattr(g, "_current_user_cached"):
        return getattr(g, "_current_user_cached")
    uid = session.get("user_id")
    if uid is None:
        user = None
    else:
        try:
            user = get_user(int(uid))
        except (TypeError, ValueError):
            user = None
    if has_request_context():
        g._current_user_cached = user
    return user


def require_current_user() -> User:
    user = current_user()
    if user is None:
        abort(401)
    return user


@app.context_processor
def inject_globals():
    user = current_user()
    allowed_pages = get_user_page_permissions(user)
    return {
        "current_user": user,
        "format_brl": format_brl,
        "format_sao_paulo_datetime": format_sao_paulo_datetime,
        "status_label": status_label,
        "request_action_label": request_action_label,
        "request_action_summary": request_action_summary,
        "user_role_label": user_role_label,
        "stock_status_class": stock_status_class,
        "stock_status_label": stock_status_label,
        "can_access": lambda page_key: page_key in allowed_pages,
        "can_access_any": lambda page_keys: any(page_key in allowed_pages for page_key in page_keys),
        "can_do": lambda action_key: user_has_action_access(user, action_key),
        "can_do_any": lambda action_keys: user_has_any_action_access(user, action_keys),
        "action_permission_options": ACTION_PERMISSION_OPTIONS,
        "all_action_permission_groups": grouped_action_permissions(page_permission_key_set()),
        "role_action_permissions_map": role_action_permissions_map(),
        "get_user_action_permissions": get_user_action_permissions,
        "selected_action_permissions_for_form": selected_action_permissions_for_form,
        "product_edit_action_keys": PRODUCT_EDIT_ACTION_KEYS,
        "page_permission_options": PAGE_PERMISSION_OPTIONS,
        "base_franchise_options": BASE_FRANCHISE_OPTIONS,
        "base_unit_options": BASE_UNIT_OPTIONS,
        "franchise_unit_options": FRANCHISE_UNIT_OPTIONS,
        "asset_unit_options": BASE_FRANCHISE_OPTIONS,
        "admin_organization_options": ADMIN_ORGANIZATION_OPTIONS,
        "asset_regional_options": ASSET_REGIONAL_OPTIONS,
        "asset_regional_for_base": asset_regional_for_base,
        "current_is_dev": current_is_dev,
        "current_can_manage_dev_roles": current_can_manage_dev_roles,
        "current_can_bootstrap_dev_role": current_can_bootstrap_dev_role,
        "current_can_edit_dev_password": current_can_edit_dev_password,
        "current_allowed_role_options": current_allowed_role_options,
        "current_can_edit_admin_role": current_can_edit_admin_role,
        "current_can_edit_user": lambda target: can_edit_user_target(user, target),
        "current_can_create_users": current_can_create_users,
        "current_can_edit_user_field": lambda target, field_key, is_new=False: can_edit_user_field(user, target, field_key, bool(is_new)),
        "current_editable_user_roles": lambda: sorted(get_user_editable_roles(user)),
        "current_editable_user_fields": lambda: sorted(get_user_editable_fields(user)),
        "current_can_manage_stock_tags": can_manage_stock_tags,
        "current_can_manage_product_blocks": lambda: can_manage_product_request_blocks(user),
        "can_manage_product_blocks": can_manage_product_request_blocks(user),
    }


def format_brl(cents: int | None) -> str:
    value = (cents or 0) / 100
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def status_label(status: str) -> str:
    labels = {
        "pending": "Pendente",
        "approved": "Aprovado",
        "rejected": "Recusado",
        "deleted": "Excluído",
    }
    return labels.get(status, status)


PAGE_PERMISSION_OPTIONS = [
    {
        "key": "home",
        "label": "Solicitar insumos",
        "description": "Permite acessar a tela inicial e enviar solicitações de insumos.",
        "admin_only": False,
    },
    {
        "key": "my_requests",
        "label": "Minhas solicitações",
        "description": "Permite visualizar as solicitações feitas pelo próprio usuário e baixar PDFs.",
        "admin_only": False,
    },
    {
        "key": "admin_dashboard",
        "label": "Painel admin",
        "description": "Permite acessar o resumo administrativo do portal.",
        "admin_only": True,
    },
    {
        "key": "admin_products",
        "label": "Produtos",
        "description": "Permite cadastrar, editar, importar e exportar produtos.",
        "admin_only": True,
    },
    {
        "key": "admin_stock",
        "label": "Gestão de estoque",
        "description": "Permite acompanhar estoque, gráficos e movimentações.",
        "admin_only": True,
    },
    {
        "key": "admin_users",
        "label": "Usuários e acessos",
        "description": "Permite aprovar, criar, editar permissões e excluir usuários.",
        "admin_only": True,
    },
    {
        "key": "admin_requests",
        "label": "Solicitações pendentes",
        "description": "Permite analisar, editar quantidades, aprovar ou recusar pedidos pendentes.",
        "admin_only": True,
    },
    {
        "key": "admin_requests_attended",
        "label": "Solicitações atendidas",
        "description": "Permite visualizar pedidos já aprovados ou recusados.",
        "admin_only": True,
    },
]


ACTION_PERMISSION_OPTIONS = [
    {"key": "products_create", "page_key": "admin_products", "label": "Criar produtos", "description": "Permite cadastrar novos produtos no catálogo."},
    {"key": "products_edit_basic", "page_key": "admin_products", "label": "Editar dados principais", "description": "Permite alterar nome, descrição e imagem dos produtos."},
    {"key": "products_edit_category", "page_key": "admin_products", "label": "Editar categorias", "description": "Permite alterar categoria, ícone e lista de categorias."},
    {"key": "products_edit_unit", "page_key": "admin_products", "label": "Editar unidade/kit", "description": "Permite alterar unidade de medida, kit e quantidade por kit."},
    {"key": "products_edit_price", "page_key": "admin_products", "label": "Editar preço", "description": "Permite alterar valor unitário dos produtos."},
    {"key": "products_edit_stock", "page_key": "admin_products", "label": "Editar estoque", "description": "Permite alterar estoque disponível diretamente no cadastro."},
    {"key": "products_edit_limits", "page_key": "admin_products", "label": "Editar limites", "description": "Permite alterar limite de pedido, estoque mínimo/máximo e quantidade mínima."},
    {"key": "products_edit_visibility", "page_key": "admin_products", "label": "Editar visibilidade/status", "description": "Permite ativar/inativar, marcar item interno e definir público base/franquia."},
    {"key": "products_import", "page_key": "admin_products", "label": "Importar produtos", "description": "Permite importar planilhas de produtos."},
    {"key": "products_export", "page_key": "admin_products", "label": "Exportar produtos", "description": "Permite exportar planilhas de produtos."},
    {"key": "products_delete", "page_key": "admin_products", "label": "Excluir produtos", "description": "Permite excluir produtos definitivamente."},
    {"key": "stock_material_entries", "page_key": "admin_stock", "label": "Entrada de materiais", "description": "Permite registrar ou importar entradas e alterar estoque por entrada."},
    {"key": "stock_assets_create", "page_key": "admin_stock", "label": "Criar ativos e baixar estoque", "description": "Permite cadastrar ativos e descontar itens do estoque."},
    {"key": "stock_reports", "page_key": "admin_stock", "label": "Gerar relatórios de estoque", "description": "Permite gerar relatórios de solicitações, entradas e ativos."},
    {"key": "requests_edit_items", "page_key": "admin_requests", "label": "Editar itens de solicitações", "description": "Permite alterar quantidades de pedidos pendentes."},
    {"key": "requests_approve_reject", "page_key": "admin_requests", "label": "Aprovar ou recusar solicitações", "description": "Permite aprovar, recusar e movimentar estoque por solicitação."},
    {"key": "requests_delete", "page_key": "admin_requests", "label": "Excluir solicitações", "description": "Permite excluir solicitações definitivamente do banco de dados."},
    {"key": "users_create_edit", "page_key": "admin_users", "label": "Criar/editar usuários", "description": "Permite cadastrar usuários e alterar dados de acesso."},
    {"key": "users_import_export", "page_key": "admin_users", "label": "Importar/exportar usuários", "description": "Permite baixar modelos, exportar tabela e importar usuários."},
    {"key": "users_status_delete", "page_key": "admin_users", "label": "Status e exclusão de usuários", "description": "Permite aprovar, recusar, reativar ou excluir usuários."},
]


PRODUCT_EDIT_ACTION_KEYS = (
    "products_edit_basic",
    "products_edit_category",
    "products_edit_unit",
    "products_edit_price",
    "products_edit_stock",
    "products_edit_limits",
    "products_edit_visibility",
)


USER_EDIT_FIELD_OPTIONS = [
    {"key": "responsible_name", "label": "Nome do responsavel", "description": "Permite alterar quem usa o acesso."},
    {"key": "username", "label": "Nome de usuario", "description": "Permite alterar o login."},
    {"key": "password", "label": "Senha individual", "description": "Permite definir ou trocar a senha do usuario."},
    {"key": "role", "label": "Tipo de acesso", "description": "Permite trocar o cargo do usuario entre os cargos autorizados."},
    {"key": "status", "label": "Status do cadastro", "description": "Permite aprovar, recusar ou deixar pendente pela tela de edicao."},
    {"key": "organization_name", "label": "Base ou setor", "description": "Permite alterar base, setor ou unidade vinculada."},
    {"key": "franchise_name", "label": "Nome da franquia", "description": "Permite alterar a franquia exibida no portal."},
    {"key": "franchise_number", "label": "Telefone da franquia", "description": "Permite alterar o telefone da franquia."},
    {"key": "cnpj", "label": "CNPJ", "description": "Permite alterar o CNPJ da franquia."},
    {"key": "page_permissions", "label": "Paginas e acoes liberadas", "description": "Permite ajustar paginas e acoes individuais do usuario."},
]


def action_permission_key_set() -> set[str]:
    return {item["key"] for item in ACTION_PERMISSION_OPTIONS}


def user_edit_field_key_set() -> set[str]:
    return {item["key"] for item in USER_EDIT_FIELD_OPTIONS}


def action_permissions_for_pages(page_keys: list[str] | set[str]) -> list[dict[str, Any]]:
    allowed_pages = set(page_keys)
    return [item for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in allowed_pages]



STATIC_ROLE_LABELS = {"dev": "Dev", "admin": "Administrador", "base": "Base", "franchise": "Franquia"}
STATIC_ROLE_KEYS = set(STATIC_ROLE_LABELS.keys())
STATIC_ROLE_ORDER = ["dev", "admin", "base", "franchise"]
STATIC_ROLE_DESCRIPTIONS = {
    "base": "Tipo padrÃ£o para bases. PermissÃµes administrativas nÃ£o liberadas.",
    "franchise": "Tipo padrÃ£o para franquias. PermissÃµes administrativas nÃ£o liberadas.",
    "admin": "Tipo padrÃ£o administrativo com acesso completo Ã s Ã¡reas administrativas.",
    "dev": "Tipo Dev com controle total do portal, tipos de acesso e permissÃµes.",
}


def page_permission_key_set() -> set[str]:
    return {item["key"] for item in PAGE_PERMISSION_OPTIONS}


def admin_page_key_set() -> set[str]:
    return {item["key"] for item in PAGE_PERMISSION_OPTIONS if item["admin_only"]}


def default_static_page_permissions(role: str) -> list[str]:
    key = canonical_role_key(role, "")
    if key in {"admin", "dev"}:
        return [item["key"] for item in PAGE_PERMISSION_OPTIONS]
    return [item["key"] for item in PAGE_PERMISSION_OPTIONS if not item["admin_only"]]


def default_static_action_permissions(role: str) -> list[str]:
    key = canonical_role_key(role, "")
    if key in {"admin", "dev"}:
        return [item["key"] for item in ACTION_PERMISSION_OPTIONS]
    return []


def default_static_user_edit_roles(role: str) -> list[str]:
    key = canonical_role_key(role, "")
    if key == "dev":
        return all_role_options()
    if key == "admin":
        return [role_key for role_key in all_role_options() if role_key != "dev"]
    return []


def default_static_user_edit_fields(role: str) -> list[str]:
    key = canonical_role_key(role, "")
    if key == "dev":
        return [item["key"] for item in USER_EDIT_FIELD_OPTIONS]
    if key == "admin":
        # Por padrão, Admin não altera identidade/login do usuário.
        # O Dev pode liberar esses campos em Tipos de acesso > Administrador.
        protected_by_default = {"responsible_name", "username"}
        return [item["key"] for item in USER_EDIT_FIELD_OPTIONS if item["key"] not in protected_by_default]
    return []


def default_static_access_role(role: str) -> AccessRoleType | None:
    key = str(role or "").strip().lower()
    if key not in STATIC_ROLE_KEYS:
        return None
    return AccessRoleType(
        role_key=key,
        name=STATIC_ROLE_LABELS.get(key, key),
        description=STATIC_ROLE_DESCRIPTIONS.get(key, "Tipo padrÃ£o do sistema."),
        permissions=default_static_page_permissions(key),
        action_permissions=default_static_action_permissions(key),
        editable_roles=default_static_user_edit_roles(key),
        editable_user_fields=default_static_user_edit_fields(key),
        is_static=True,
    )


def custom_access_role_key(value: Any) -> str:
    raw = normalize_header(str(value or "")).replace("_", " ").strip()
    slug = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
    if not slug:
        slug = "tipo"
    if slug in STATIC_ROLE_KEYS:
        slug = f"personalizado_{slug}"
    return f"custom_{slug}"[:80].rstrip("_")


def editable_custom_access_role_key(value: Any, fallback: Any = "") -> str:
    raw = normalize_header(str(value or fallback or "")).replace("_", " ").strip()
    slug = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
    if slug.startswith("custom_") and len(slug) > len("custom_"):
        slug = slug[:80].rstrip("_")
    if not slug:
        slug = "tipo"
    if slug in STATIC_ROLE_KEYS:
        slug = f"personalizado_{slug}"
    return slug[:80].rstrip("_")


def display_access_role_key(role_key: Any) -> str:
    key = str(role_key or "").strip()
    if key.startswith("custom_") and len(key) > len("custom_"):
        return key[len("custom_"):]
    return key


def parse_role_permissions(value: Any) -> list[str]:
    valid = page_permission_key_set()
    if isinstance(value, (list, tuple, set)):
        raw_items = value
    else:
        try:
            decoded = json.loads(str(value or "[]"))
            raw_items = decoded if isinstance(decoded, list) else []
        except Exception:
            raw_items = []
    seen: list[str] = []
    for item in raw_items:
        key = str(item or "").strip()
        if key in valid and key not in seen:
            seen.append(key)
    if not seen:
        seen = ["home", "my_requests"]
    return seen


def parse_action_permissions(value: Any, allowed_page_keys: list[str] | set[str] | None = None) -> list[str]:
    valid = action_permission_key_set()
    if allowed_page_keys is not None:
        valid = {item["key"] for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in set(allowed_page_keys)}
    if isinstance(value, (list, tuple, set)):
        raw_items = value
    else:
        try:
            decoded = json.loads(str(value or "[]"))
            raw_items = decoded if isinstance(decoded, list) else []
        except Exception:
            raw_items = []
    seen: list[str] = []
    for item in raw_items:
        key = str(item or "").strip()
        if key in valid and key not in seen:
            seen.append(key)
    return seen


def parse_user_edit_roles(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        raw_items = value
    else:
        try:
            decoded = json.loads(str(value or "[]"))
            raw_items = decoded if isinstance(decoded, list) else []
        except Exception:
            raw_items = []
    seen: list[str] = []
    for item in raw_items:
        key = str(item or "").strip().lower()
        if key and re.fullmatch(r"[a-z0-9_]+", key) and key not in seen:
            seen.append(key[:80])
    return seen


def parse_user_edit_fields(value: Any) -> list[str]:
    valid = user_edit_field_key_set()
    if isinstance(value, (list, tuple, set)):
        raw_items = value
    else:
        try:
            decoded = json.loads(str(value or "[]"))
            raw_items = decoded if isinstance(decoded, list) else []
        except Exception:
            raw_items = []
    seen: list[str] = []
    for item in raw_items:
        key = str(item or "").strip()
        if key in valid and key not in seen:
            seen.append(key)
    return seen


def row_to_access_role(row: Any | None) -> AccessRoleType | None:
    if row is None:
        return None
    role_key = str(row["role_key"] or "").strip().lower()
    return AccessRoleType(
        role_key=role_key,
        name=str(row["name"] or ""),
        description=str(row["description"] or ""),
        permissions=parse_role_permissions(row["permissions_json"] if "permissions_json" in row.keys() else "[]"),
        action_permissions=parse_action_permissions(row["action_permissions_json"] if "action_permissions_json" in row.keys() else "[]", parse_role_permissions(row["permissions_json"] if "permissions_json" in row.keys() else "[]")),
        editable_roles=parse_user_edit_roles(row["editable_roles_json"] if "editable_roles_json" in row.keys() else "[]"),
        editable_user_fields=parse_user_edit_fields(row["editable_user_fields_json"] if "editable_user_fields_json" in row.keys() else "[]"),
        created_at=parse_dt(row["created_at"]) if "created_at" in row.keys() else None,
        updated_at=parse_dt(row["updated_at"]) if "updated_at" in row.keys() else None,
        is_static=role_key in STATIC_ROLE_KEYS,
    )


@lru_cache(maxsize=128)
def get_access_role_override(role_key: str | None) -> AccessRoleType | None:
    key = str(role_key or "").strip().lower()
    if not key or key == "dev":
        return None
    try:
        with db_connect() as conn:
            row = conn.execute(
                "SELECT role_key, name, description, permissions_json, action_permissions_json, editable_roles_json, editable_user_fields_json, created_at, updated_at FROM access_role_types WHERE role_key = ?",
                (key,),
            ).fetchone()
        return row_to_access_role(row)
    except Exception:
        return None


@lru_cache(maxsize=128)
def get_custom_access_role(role_key: str | None) -> AccessRoleType | None:
    key = str(role_key or "").strip().lower()
    if not key or key in STATIC_ROLE_KEYS:
        return None
    role = get_access_role_override(key)
    return role if role is not None and not role.is_static else None


def get_access_role_definition(role_key: str | None) -> AccessRoleType | None:
    key = str(role_key or "").strip().lower()
    if not key:
        return None
    if key == "dev":
        return default_static_access_role("dev")
    override = get_access_role_override(key)
    if key == "admin":
        if override is not None:
            # Páginas, ações e cargos do Admin continuam completos.
            # Campos editáveis respeitam o que o Dev marcou no tipo de acesso.
            override.action_permissions = default_static_action_permissions("admin")
            override.editable_roles = default_static_user_edit_roles("admin")
            override.permissions = default_static_page_permissions("admin")
            if not override.editable_user_fields:
                override.editable_user_fields = default_static_user_edit_fields("admin")
            return override
        return default_static_access_role("admin")
    if override is not None:
        return override
    if key in STATIC_ROLE_KEYS:
        return default_static_access_role(key)
    return get_custom_access_role(key)


def list_custom_access_roles() -> list[AccessRoleType]:
    try:
        with db_connect() as conn:
            rows = conn.execute(
                "SELECT role_key, name, description, permissions_json, action_permissions_json, editable_roles_json, editable_user_fields_json, created_at, updated_at FROM access_role_types ORDER BY name COLLATE NOCASE ASC"
            ).fetchall()
        return [role for row in rows if (role := row_to_access_role(row)) is not None and not role.is_static]
    except Exception:
        return []


def all_role_options() -> list[str]:
    return ["base", "franchise", "admin", *[role.role_key for role in list_custom_access_roles()], "dev"]


def role_option_labels() -> dict[str, str]:
    labels = dict(STATIC_ROLE_LABELS)
    labels.update({role.role_key: role.name for role in list_custom_access_roles()})
    return labels


def role_permissions_map(options: list[str] | None = None) -> dict[str, list[str]]:
    role_options = options or all_role_options()
    return {role: sorted(default_page_keys_for_role(role)) for role in role_options}


def role_action_permissions_map(options: list[str] | None = None) -> dict[str, list[str]]:
    role_options = options or all_role_options()
    output: dict[str, list[str]] = {}
    for role in role_options:
        role_definition = get_access_role_definition(role)
        output[role] = sorted(role_definition.action_permissions if role_definition is not None else [])
    return output


def role_is_admin_like(role: str | None) -> bool:
    key = canonical_role_key(role, "")
    if key == "dev":
        return True
    role_definition = get_access_role_definition(key)
    return bool(role_definition and role_definition.is_admin_like)


def default_page_keys_for_role(role: str) -> set[str]:
    role = canonical_role_key(role, "base")
    if role == "dev":
        return {item["key"] for item in PAGE_PERMISSION_OPTIONS}
    role_definition = get_access_role_definition(role)
    if role_definition is not None:
        return set(role_definition.permissions)
    return {item["key"] for item in PAGE_PERMISSION_OPTIONS if not item["admin_only"]}


def permission_options_for_role(role: str) -> list[dict[str, Any]]:
    role = canonical_role_key(role, "base")
    if role == "dev":
        return PAGE_PERMISSION_OPTIONS
    role_definition = get_access_role_definition(role)
    if role_definition is not None:
        allowed = set(role_definition.permissions)
        return [item for item in PAGE_PERMISSION_OPTIONS if item["key"] in allowed]
    return [item for item in PAGE_PERMISSION_OPTIONS if not item["admin_only"]]


def get_user_page_permissions(user: User | None) -> set[str]:
    if user is None:
        return set()
    role_key = canonical_role_key(user.role, "base")
    cache_key = f"_page_permissions_{user.id}"
    if has_request_context() and hasattr(g, cache_key):
        return set(getattr(g, cache_key))
    if role_key == "dev":
        # Dev permanece com acesso total para não perder o controle do portal.
        allowed = page_permission_key_set()
    elif not user.page_permissions_configured:
        allowed = default_page_keys_for_role(role_key)
    else:
        with db_connect() as conn:
            rows = conn.execute(
                "SELECT page_key FROM user_page_permissions WHERE user_id = ?",
                (user.id,),
            ).fetchall()
        # Permissões individuais salvas pelo Dev são independentes do cargo.
        # Assim o Dev pode liberar/remover páginas por usuário sem depender do tipo de acesso.
        allowed = {str(row["page_key"]) for row in rows} & page_permission_key_set()
    if has_request_context():
        setattr(g, cache_key, set(allowed))
    return set(allowed)


def get_user_action_permissions(user: User | None) -> set[str]:
    if user is None:
        return set()
    role_key = canonical_role_key(user.role, "base")
    cache_key = f"_action_permissions_{user.id}"
    if has_request_context() and hasattr(g, cache_key):
        return set(getattr(g, cache_key))
    allowed_pages = get_user_page_permissions(user)
    if role_key == "dev":
        allowed = action_permission_key_set()
    elif user.action_permissions_configured:
        with db_connect() as conn:
            rows = conn.execute(
                "SELECT action_key FROM user_action_permissions WHERE user_id = ?",
                (user.id,),
            ).fetchall()
        selected = {str(row["action_key"] or "").strip() for row in rows}
        allowed = selected & {item["key"] for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in allowed_pages}
    else:
        role_definition = get_access_role_definition(role_key)
        allowed = set(role_definition.action_permissions) if role_definition is not None else set()
        allowed = {key for key in allowed if any(item["key"] == key and item["page_key"] in allowed_pages for item in ACTION_PERMISSION_OPTIONS)}
    if has_request_context():
        setattr(g, cache_key, set(allowed))
    return set(allowed)


def user_has_action_access(user: User | None, action_key: str) -> bool:
    return str(action_key or "").strip() in get_user_action_permissions(user)


def user_has_any_action_access(user: User | None, action_keys: list[str] | tuple[str, ...] | set[str]) -> bool:
    allowed = get_user_action_permissions(user)
    return any(str(action_key or "").strip() in allowed for action_key in action_keys)


def get_user_editable_roles(user: User | None) -> set[str]:
    if user is None:
        return set()
    role_key = canonical_role_key(user.role, "base")
    if role_key == "dev":
        return set(all_role_options())
    if role_key == "admin":
        roles = set(all_role_options())
        roles.discard("dev")
        return roles
    role_definition = get_access_role_definition(role_key)
    roles = set(role_definition.editable_roles if role_definition is not None else [])
    roles.discard("dev")
    return {role for role in roles if role in set(all_role_options())}


def get_user_editable_fields(user: User | None) -> set[str]:
    if user is None:
        return set()
    role_key = canonical_role_key(user.role, "base")
    if role_key == "dev":
        return user_edit_field_key_set()
    role_definition = get_access_role_definition(role_key)
    fields = set(role_definition.editable_user_fields if role_definition is not None else [])
    return fields & user_edit_field_key_set()


def can_edit_user_target(current: User | None, target: User | None) -> bool:
    if current is None or target is None:
        return False
    if current.is_dev and current.is_approved:
        return True
    if not user_has_action_access(current, "users_create_edit"):
        return False
    if target.role == "dev":
        return False
    return target.role in get_user_editable_roles(current)


def can_create_user_role(current: User | None, role: str | None) -> bool:
    role_key = str(role or "").strip().lower()
    if current is None:
        return False
    if current.is_dev and current.is_approved:
        return True
    if not user_has_action_access(current, "users_create_edit"):
        return False
    if role_key == "dev":
        return bool(current.is_admin and not dev_user_exists())
    return role_key in get_user_editable_roles(current)


def can_edit_user_field(current: User | None, target: User | None, field_key: str, is_new: bool = False) -> bool:
    key = str(field_key or "").strip()
    if key not in user_edit_field_key_set():
        return False
    if current is None:
        return False
    if current.is_dev and current.is_approved:
        return True
    if not user_has_action_access(current, "users_create_edit"):
        return False
    # Na criação do acesso, responsável/login/senha precisam ser preenchidos.
    # Depois de criado, responsável e login só mudam se o Dev liberar esses campos.
    if is_new and key in {"responsible_name", "username", "password"}:
        return True
    if not is_new and target is not None and not can_edit_user_target(current, target):
        return False
    return key in get_user_editable_fields(current)


def current_can_create_users() -> bool:
    user = current_user()
    return user_has_action_access(user, "users_create_edit") and bool(get_user_editable_roles(user) or (user and user.is_dev))


def require_action_permission(action_key: str, message: str | None = None, redirect_endpoint: str = "admin_products"):
    if user_has_action_access(current_user(), action_key):
        return None
    flash(message or "Seu tipo de acesso não tem permissão para executar esta ação.", "warning")
    return redirect(request.referrer or url_for(redirect_endpoint))


def user_has_page_access(user: User | None, page_key: str) -> bool:
    if user is None:
        return False
    return page_key in get_user_page_permissions(user)


def user_has_any_page_access(user: User | None, page_keys: list[str]) -> bool:
    if user is None:
        return False
    allowed = get_user_page_permissions(user)
    return any(key in allowed for key in page_keys)


def normalize_user_page_selection_for_editor(editor: User | None, role: str, selected_keys: list[str] | set[str]) -> set[str]:
    role_key = canonical_role_key(role, "base")
    valid_pages = page_permission_key_set()
    if role_key == "dev":
        return set(valid_pages)
    if editor is not None and editor.is_dev and editor.is_approved:
        return {str(key or "").strip() for key in selected_keys if str(key or "").strip() in valid_pages}
    allowed_for_role = default_page_keys_for_role(role_key)
    return {str(key or "").strip() for key in selected_keys if str(key or "").strip() in allowed_for_role}


def normalize_user_action_selection_for_editor(editor: User | None, role: str, selected_action_keys: list[str] | set[str], selected_page_keys: list[str] | set[str]) -> set[str]:
    role_key = canonical_role_key(role, "base")
    allowed_pages = set(selected_page_keys or []) & page_permission_key_set()
    valid_actions_for_pages = {item["key"] for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in allowed_pages}
    if role_key == "dev":
        return action_permission_key_set()
    if editor is not None and editor.is_dev and editor.is_approved:
        return {str(key or "").strip() for key in selected_action_keys if str(key or "").strip() in valid_actions_for_pages}
    role_definition = get_access_role_definition(role_key)
    allowed_for_role = set(role_definition.action_permissions) if role_definition is not None else set()
    return {
        str(key or "").strip()
        for key in selected_action_keys
        if str(key or "").strip() in allowed_for_role and str(key or "").strip() in valid_actions_for_pages
    }


def save_user_page_permissions(conn: Any, user_id: int, role: str, selected_keys: list[str] | set[str]) -> None:
    normalized = normalize_user_page_selection_for_editor(current_user() if has_request_context() else None, role, selected_keys)
    conn.execute("DELETE FROM user_page_permissions WHERE user_id = ?", (user_id,))
    created_at = now_iso()
    for key in sorted(normalized):
        conn.execute(
            "INSERT OR IGNORE INTO user_page_permissions (user_id, page_key, created_at) VALUES (?, ?, ?)",
            (user_id, key, created_at),
        )
    conn.execute(
        "UPDATE users SET page_permissions_configured = 1, updated_at = ? WHERE id = ?",
        (now_iso(), user_id),
    )


def save_user_action_permissions(conn: Any, user_id: int, role: str, selected_action_keys: list[str] | set[str], selected_page_keys: list[str] | set[str]) -> None:
    normalized = normalize_user_action_selection_for_editor(current_user() if has_request_context() else None, role, selected_action_keys, selected_page_keys)
    conn.execute("DELETE FROM user_action_permissions WHERE user_id = ?", (user_id,))
    created_at = now_iso()
    for key in sorted(normalized):
        conn.execute(
            "INSERT OR IGNORE INTO user_action_permissions (user_id, action_key, created_at) VALUES (?, ?, ?)",
            (user_id, key, created_at),
        )
    conn.execute(
        "UPDATE users SET action_permissions_configured = 1, updated_at = ? WHERE id = ?",
        (now_iso(), user_id),
    )


def selected_permissions_for_form(user: User | None, role: str | None = None) -> set[str]:
    if user is not None:
        return get_user_page_permissions(user)
    return default_page_keys_for_role(role or "base")


def selected_action_permissions_for_form(user: User | None, role: str | None = None, selected_pages: list[str] | set[str] | None = None) -> set[str]:
    pages = set(selected_pages or (get_user_page_permissions(user) if user is not None else default_page_keys_for_role(role or "base")))
    if user is not None:
        return get_user_action_permissions(user) & {item["key"] for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in pages}
    role_key = canonical_role_key(role, "base")
    role_definition = get_access_role_definition(role_key)
    allowed = set(role_definition.action_permissions) if role_definition is not None else set()
    return allowed & {item["key"] for item in ACTION_PERMISSION_OPTIONS if item["page_key"] in pages}


def page_access_required(page_key: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if user is None:
                abort(401)
            if not user_has_page_access(user, page_key):
                flash("Seu usuário não possui acesso a esta página.", "warning")
                if user.is_admin and user_has_page_access(user, "admin_dashboard"):
                    return redirect(url_for("admin_dashboard"))
                if user_has_page_access(user, "home"):
                    return redirect(url_for("home"))
                if user_has_page_access(user, "my_requests"):
                    return redirect(url_for("my_requests"))
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def page_access_any_required(page_keys: list[str]):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if user is None:
                abort(401)
            if not user_has_any_page_access(user, page_keys):
                flash("Seu usuário não possui acesso a esta página.", "warning")
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def stock_status_class(product: Product) -> str:
    minimum = product.min_stock
    maximum = product.max_stock
    stock = product.stock_quantity
    if minimum is None and maximum is None:
        return "normal"
    if minimum is not None:
        alert_limit = minimum
        if maximum is not None and maximum > minimum:
            alert_limit = minimum + max(1, int((maximum - minimum) * 0.15))
        if stock <= alert_limit:
            return "critical"
    if maximum is not None:
        return "good" if stock >= maximum else "normal"
    if minimum is not None and stock > minimum:
        return "good"
    return "normal"


def stock_status_label(product: Product) -> str:
    status = stock_status_class(product)
    if status == "critical":
        return "Atenção"
    if status == "good":
        return "Saudável"
    return "Normal"


def movement_type_label(movement_type: str) -> str:
    labels = {
        "request_approved": "Saída por solicitação",
        "manual_adjustment": "Ajuste manual",
        "product_created": "Cadastro de produto",
        "asset_allocation": "Saída para ativo",
        "material_entry": "Entrada de materiais",
        "material_import": "Importação de entrada",
        "import_adjustment": "Ajuste por importação",
    }
    return labels.get(movement_type, movement_type)


def record_stock_movement(
    conn: Any,
    product_id: int,
    quantity_delta: int,
    stock_before: int,
    stock_after: int,
    movement_type: str,
    note: str = "",
    request_id: int | None = None,
    created_by_id: int | None = None,
) -> None:
    created_at = now_iso()
    conn.execute(
        """
        INSERT INTO stock_movements (product_id, request_id, created_by_id, movement_type, quantity_delta, stock_before, stock_after, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (product_id, request_id, created_by_id, movement_type, quantity_delta, stock_before, stock_after, note, created_at),
    )
    notify_feishu_stock_movement(
        conn,
        product_id=product_id,
        quantity_delta=quantity_delta,
        stock_before=stock_before,
        stock_after=stock_after,
        movement_type=movement_type,
        note=note,
        request_id=request_id,
        created_by_id=created_by_id,
        created_at=created_at,
    )


FEISHU_ITEM_LIMIT = 18


def compact_text(value: Any, fallback: str = "-", limit: int = 240) -> str:
    text = " ".join(str(value or "").split())
    if not text:
        text = fallback
    if limit > 3 and len(text) > limit:
        return text[: limit - 3].rstrip() + "..."
    return text[:limit]


def feishu_md(value: Any, fallback: str = "-", limit: int = 240) -> str:
    text = compact_text(value, fallback=fallback, limit=limit)
    for char in ("\\", "*", "_", "`", "[", "]", "(", ")", "#"):
        text = text.replace(char, "\\" + char)
    return text


def feishu_line(label: str, value: Any) -> str:
    return f"**{label}:** {feishu_md(value)}"


def user_role_label(role: str | None) -> str:
    key = str(role or "").strip().lower()
    if key in STATIC_ROLE_LABELS:
        return "Admin" if key == "admin" else STATIC_ROLE_LABELS[key]
    custom = get_custom_access_role(key)
    return custom.name if custom else (role or "-")


def format_feishu_datetime(value: datetime | None = None) -> str:
    return format_sao_paulo_datetime(value, suffix=" (Brasilia)")


def public_url_for(endpoint: str, **values: Any) -> str:
    clean_values = {key: value for key, value in values.items() if value is not None}
    if not has_request_context():
        return PUBLIC_BASE_URL
    try:
        path = url_for(endpoint, **clean_values)
        if PUBLIC_BASE_URL:
            return f"{PUBLIC_BASE_URL}{path}"
        return url_for(endpoint, _external=True, **clean_values)
    except Exception:
        app.logger.exception("Falha ao montar link publico para o Feishu")
        return PUBLIC_BASE_URL


def dispatch_feishu_webhook(payload: dict[str, Any]) -> None:
    webhook_url = FEISHU_STOCK_WEBHOOK_URL
    if not webhook_url:
        return

    def send_payload() -> None:
        try:
            response = requests.post(webhook_url, json=payload, timeout=8)
            response.raise_for_status()
            try:
                response_payload = response.json()
            except ValueError:
                response_payload = {}
            if isinstance(response_payload, dict) and response_payload.get("code") not in (0, None):
                app.logger.warning("Feishu webhook retornou erro: %s", response_payload)
        except Exception:
            app.logger.exception("Falha ao enviar notificacao para o Feishu")

    threading.Thread(target=send_payload, daemon=True).start()


def send_feishu_card(title: str, lines: list[str], link_text: str, link_url: str, template: str = "red") -> None:
    content = "\n".join(line for line in lines if line is not None).strip()
    if len(content) > 3500:
        content = content[:3497].rstrip() + "..."

    elements: list[dict[str, Any]] = []
    if content:
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": content}})
    if link_url:
        elements.append(
            {
                "tag": "action",
                "actions": [
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": compact_text(link_text, limit=50)},
                        "type": "primary",
                        "url": link_url,
                    }
                ],
            }
        )

    if not elements:
        return

    dispatch_feishu_webhook(
        {
            "msg_type": "interactive",
            "card": {
                "config": {"wide_screen_mode": True},
                "header": {
                    "template": template or "red",
                    "title": {"tag": "plain_text", "content": compact_text(title, limit=80)},
                },
                "elements": elements,
            },
        }
    )


def request_item_feishu_lines(items: list[RequestItem]) -> list[str]:
    lines: list[str] = []
    for index, item in enumerate(items[:FEISHU_ITEM_LIMIT], start=1):
        product = item.product
        product_name = product.name if product else item.product_name_snapshot
        unit = product.unit_measure if product and product.unit_measure else "un"
        lines.extend(
            [
                f"**{index}. {feishu_md(product_name)}**",
                f"> Quantidade: **{feishu_md(item.quantity, limit=40)} {feishu_md(unit, limit=40)}**",
                "",
            ]
        )
    remaining = len(items) - FEISHU_ITEM_LIMIT
    if remaining > 0:
        lines.append(f"+{remaining} item(ns) adicionais no portal")
    return lines or ["Sem itens"]


def notify_feishu_supply_request_created(supply_request: SupplyRequest, link_url: str) -> None:
    requester = supply_request.user
    requester_name = requester.responsible_name if requester else f"Usuario #{supply_request.user_id}"
    requester_org = requester.organization_name if requester else "-"
    requester_role = user_role_label(requester.role if requester else "")

    lines = [
        feishu_line("Solicitacao", f"#{supply_request.id}"),
        feishu_line("Pedido por", requester_name),
        feishu_line("Setor", requester_org),
        feishu_line("Tipo", requester_role),
        feishu_line("Status", "Pendente"),
        feishu_line("Data", format_feishu_datetime()),
    ]
    if supply_request.user_note:
        lines.append(feishu_line("Observacao do pedido", supply_request.user_note))
    lines.extend(["", "---", "**Itens solicitados**", "", *request_item_feishu_lines(supply_request.items)])

    send_feishu_card("Nova solicitação de insumos", lines, "Abrir solicitação", link_url)


def notify_feishu_supply_request_action(supply_request: SupplyRequest, action: str, actor: User, link_url: str, admin_note: str = "") -> None:
    action = (action or "").strip().lower()
    requester = supply_request.user
    requester_name = requester.responsible_name if requester else f"Usuario #{supply_request.user_id}"
    requester_org = requester.organization_name if requester else "-"
    requester_role = user_role_label(requester.role if requester else "")
    status_text = {"approved": "Aprovada", "rejected": "Recusada", "deleted": "Excluída"}.get(action, request_action_label(action))
    actor_label = {"approved": "Aprovada por", "rejected": "Recusada por", "deleted": "Excluída por"}.get(action, "Executada por")
    title = {
        "approved": "Solicitação de insumos aprovada",
        "rejected": "Solicitação de insumos recusada",
        "deleted": "Solicitação de insumos excluída",
    }.get(action, "Solicitação de insumos atualizada")
    template = "green" if action == "approved" else "red"
    lines = [
        feishu_line("Solicitação", f"#{supply_request.id}"),
        feishu_line("Pedido por", requester_name),
        feishu_line("Setor", requester_org),
        feishu_line("Tipo", requester_role),
        feishu_line("Status", status_text),
        feishu_line(actor_label, request_action_actor_name(actor)),
        feishu_line("Data da ação", format_feishu_datetime()),
    ]
    note_text = admin_note or supply_request.admin_note
    if note_text:
        lines.append(feishu_line("Observação admin", note_text))
    if supply_request.items:
        lines.extend(["", "---", "**Itens solicitados**", "", *request_item_feishu_lines(supply_request.items)])
    send_feishu_card(title, lines, "Abrir solicitação", link_url, template=template)


def notify_feishu_user_registration_requested(user: User, link_url: str) -> None:
    org_label = user.franchise_name or user.organization_name or "-"
    lines = [
        feishu_line("Responsável", user.responsible_name),
        feishu_line("Usuário", user.username),
        feishu_line("Tipo", user_role_label(user.role)),
        feishu_line("Setor", org_label),
        feishu_line("Telefone", user.formatted_phone or "-"),
        feishu_line("CNPJ", user.formatted_cnpj or "-"),
        feishu_line("Status", "Pendente"),
        feishu_line("Data", format_feishu_datetime(user.created_at)),
    ]
    send_feishu_card("Novo cadastro solicitado", lines, "Abrir cadastros", link_url)


def notify_feishu_asset_created(
    asset_id: int,
    name: str,
    base: str,
    regional: str,
    sector: str,
    manager: str,
    created_by: User,
    item_rows: list[dict[str, Any]],
    product_map: dict[int, Product],
    link_url: str,
) -> None:
    item_lines: list[str] = []
    for index, item in enumerate(item_rows[:FEISHU_ITEM_LIMIT], start=1):
        product = product_map.get(int(item["product_id"]))
        product_name = product.name if product else item.get("item_name", "Item")
        unit = product.unit_measure if product and product.unit_measure else "un"
        serial_number = compact_text(item.get("serial_number"), fallback="", limit=120)
        item_lines.extend(
            [
                f"**{index}. {feishu_md(product_name)}**",
                f"> Quantidade: **{feishu_md(item.get('quantity'), limit=40)} {feishu_md(unit, limit=40)}**",
            ]
        )
        if serial_number:
            item_lines.append(f"> Patrimônio/Série: {feishu_md(serial_number, fallback='', limit=120)}")
        item_lines.append("")
    remaining = len(item_rows) - FEISHU_ITEM_LIMIT
    if remaining > 0:
        item_lines.append(f"+{remaining} item(ns) adicionais no portal")

    lines = [
        feishu_line("Ativo", f"#{asset_id} - {name}"),
        feishu_line("Regional", regional),
        feishu_line("Base", base),
        feishu_line("Setor", sector),
        feishu_line("Gestor", manager),
        feishu_line("Cadastrado por", created_by.responsible_name),
        feishu_line("Data", format_feishu_datetime()),
        "",
        "---",
        "**Itens vinculados**",
        "",
        *(item_lines or ["Sem itens"]),
    ]

    send_feishu_card("Novo ativo cadastrado", lines, "Abrir ativo", link_url)




def notify_feishu_stock_movement(
    conn: Any,
    *,
    product_id: int,
    quantity_delta: int,
    stock_before: int,
    stock_after: int,
    movement_type: str,
    note: str = "",
    request_id: int | None = None,
    created_by_id: int | None = None,
    created_at: str | None = None,
) -> None:
    if not has_request_context() or not quantity_delta:
        return
    try:
        product_row = conn.execute("SELECT name, unit_measure FROM products WHERE id = ?", (product_id,)).fetchone()
        product_name = product_row["name"] if product_row else f"Produto #{product_id}"
        unit_measure = (product_row["unit_measure"] if product_row and "unit_measure" in product_row.keys() else "") or "un"
        responsible = "Sistema"
        if created_by_id:
            user_row = conn.execute("SELECT responsible_name, username FROM users WHERE id = ?", (created_by_id,)).fetchone()
            if user_row:
                responsible = user_row["responsible_name"] or user_row["username"] or f"Usuário #{created_by_id}"
        direction = "Entrada" if quantity_delta > 0 else "Saída"
        qty_prefix = "+" if quantity_delta > 0 else ""
        qty_text = f"{qty_prefix}{quantity_delta} {unit_measure}".strip()
        movement_dt = parse_dt(created_at) if created_at else datetime.utcnow()
        request_unit = ""
        request_requester = ""
        if request_id:
            request_row = conn.execute(
                """
                SELECT u.responsible_name, u.username, u.organization_name
                  FROM supply_requests sr
                  LEFT JOIN users u ON u.id = sr.user_id
                 WHERE sr.id = ?
                """,
                (request_id,),
            ).fetchone()
            if request_row is not None:
                request_unit = str(request_row["organization_name"] or "").strip()
                request_requester = str(request_row["responsible_name"] or request_row["username"] or "").strip()
        responsible_label = "Aprovado por" if movement_type == "request_approved" else "Responsável"
        lines = [
            feishu_line(responsible_label, responsible),
            feishu_line("Produto", product_name),
            feishu_line("Movimentação", direction),
            feishu_line("Quantidade", qty_text),
            feishu_line("Estoque", f"{stock_before} → {stock_after}"),
            feishu_line("Tipo", movement_type_label(movement_type)),
            feishu_line("Data", format_feishu_datetime(movement_dt)),
        ]
        if request_id:
            lines.insert(1, feishu_line("Solicitação", f"#{request_id}"))
            if request_unit:
                lines.insert(2, feishu_line("Unidade/Base", request_unit))
            if request_requester:
                lines.insert(3, feishu_line("Pedido por", request_requester))
            if movement_type == "request_approved":
                reason = f"Saída por aprovação da solicitação #{request_id}"
                if request_unit:
                    reason += f" para {request_unit}"
                lines.insert(4, feishu_line("Motivo da saída", reason))
        if note:
            lines.append(feishu_line("Observação", note))
        link_url = public_url_for("admin_request_detail", request_id=request_id) if request_id else public_url_for("admin_material_entries")
        link_text = "Abrir solicitação" if request_id else "Abrir entrada de materiais"
        title = "Estoque atualizado após aprovação" if movement_type == "request_approved" else "Movimentação de estoque"
        send_feishu_card(title, lines, link_text, link_url, template="green")
    except Exception:
        app.logger.exception("Falha ao preparar notificacao Feishu da movimentacao de estoque")


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    for char in [".", ":", ";", "-", "_", "/", "\\", "(", ")"]:
        text = text.replace(char, " ")
    return " ".join(text.split())


def parse_money_to_cents(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        try:
            return int(round(float(value) * 100))
        except (TypeError, ValueError, OverflowError):
            return 0
    text_value = str(value).strip()
    for token in ["R$", "BRL", "￥", "¥"]:
        text_value = text_value.replace(token, "")
    text_value = text_value.replace(" ", "")
    text_value = re.sub(r"[^0-9,.-]", "", text_value)
    if not text_value:
        return 0
    if "," in text_value and "." in text_value:
        if text_value.rfind(",") > text_value.rfind("."):
            text_value = text_value.replace(".", "").replace(",", ".")
        else:
            text_value = text_value.replace(",", "")
    elif "," in text_value:
        if re.fullmatch(r"\d{1,3}(,\d{3})+", text_value):
            text_value = text_value.replace(",", "")
        else:
            text_value = text_value.replace(".", "").replace(",", ".")
    elif "." in text_value:
        if re.fullmatch(r"\d{1,3}(\.\d{3})+", text_value):
            text_value = text_value.replace(".", "")
    try:
        return max(0, int(round(float(text_value or "0") * 100)))
    except (TypeError, ValueError, OverflowError):
        return 0


def parse_bool_value(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    text_raw = str(value).strip().lower()
    text = normalize_header(value)
    true_values = {"1", "sim", "s", "yes", "y", "true", "ativo", "ativado", "active", "habilitado", "enabled", "enable", "是", "启用", "啟用", "正常"}
    false_values = {"0", "nao", "não", "n", "no", "false", "inativo", "desativado", "inactive", "disabled", "disable", "否", "停用", "禁用"}
    if text in true_values or text_raw in true_values:
        return True
    if text in false_values or text_raw in false_values:
        return False
    return default


def safe_filename(text_value: str) -> str:
    normalized = normalize_header(text_value).replace(" ", "_")
    allowed = "".join(ch for ch in normalized if ch.isalnum() or ch in {"_", "-"})
    return allowed[:70] or "arquivo"


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None:
            flash("Faça login para continuar.", "warning")
            return redirect(url_for("login"))
        if not user.is_admin and not user.is_approved:
            session.clear()
            flash("Seu cadastro ainda não foi aprovado por um administrador.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)

    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if user is None or not user.is_admin or not user.is_approved:
            abort(403)
        return fn(*args, **kwargs)

    return wrapper


def generate_code(length: int = 6) -> str:
    return "".join(random.choice(string.digits) for _ in range(length))


def email_is_configured() -> bool:
    return bool(MAIL_SERVER and MAIL_USERNAME and MAIL_PASSWORD)


def send_admin_login_code(user: User, code: str) -> bool:
    subject = "Código de confirmação de login - Portal de Insumos J&T"
    body = (
        f"Olá, {user.responsible_name}.\n\n"
        f"Seu código de confirmação de login é: {code}\n"
        f"Validade: {ADMIN_CODE_MINUTES} minutos.\n\n"
        "Caso não tenha solicitado, ignore este e-mail."
    )

    if email_is_configured() and is_real_email(user.email):
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = MAIL_DEFAULT_SENDER
        msg["To"] = user.email
        msg.set_content(body)
        try:
            if MAIL_USE_SSL:
                smtp_context = smtplib.SMTP_SSL(MAIL_SERVER, MAIL_PORT, timeout=20)
            else:
                smtp_context = smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=20)
            with smtp_context as smtp:
                if MAIL_USE_TLS and not MAIL_USE_SSL:
                    smtp.starttls()
                smtp.login(MAIL_USERNAME, MAIL_PASSWORD)
                smtp.send_message(msg)
            return True
        except Exception as exc:
            print("=" * 80)
            print("ERRO AO ENVIAR CÓDIGO DE CONFIRMAÇÃO DE LOGIN POR E-MAIL")
            print(f"Destino: {user.email}")
            print(f"SMTP: {MAIL_SERVER}:{MAIL_PORT} | TLS={MAIL_USE_TLS} | SSL={MAIL_USE_SSL}")
            print(f"Erro: {type(exc).__name__} - {exc}")
            print("=" * 80)
            flash("Não consegui enviar o código de confirmação por e-mail. Confira as variáveis SMTP no Render e tente novamente.", "danger")
            return False

    print("=" * 80)
    print(f"CÓDIGO DE CONFIRMAÇÃO DE LOGIN DEV para {user.username}: {code}")
    print("=" * 80)
    flash(f"Insira o código para confirmar o seu login: {code}", "info")
    return True


def product_limit_for(product: Product, user: User) -> int | None:
    if user.role == "franchise":
        return product.limit_franchise
    if user.role == "base":
        return product.limit_base
    return None


def product_limit_unit_label(product: Product) -> str:
    return "kits" if product.is_kit and product.kit_quantity > 1 else (product.unit_measure or "un")


def product_limit_block_days(product: Product) -> int:
    try:
        days = int(product.limit_block_days or 60)
    except (TypeError, ValueError):
        days = 60
    return max(1, days)


def product_block_start_for_period(product: Product) -> datetime:
    return datetime.utcnow() - timedelta(days=product_limit_block_days(product))


def stored_quantity_to_limit_quantity(product: Product, stored_quantity: Any) -> int:
    try:
        qty = max(0, int(stored_quantity or 0))
    except (TypeError, ValueError):
        qty = 0
    multiplier = product.kit_quantity if product.is_kit and product.kit_quantity > 1 else 1
    if multiplier > 1:
        return (qty + multiplier - 1) // multiplier
    return qty


def requested_quantity_to_limit_quantity(product: Product, requested_quantity: Any) -> int:
    try:
        return max(0, int(requested_quantity or 0))
    except (TypeError, ValueError):
        return 0


def can_manage_product_request_blocks(user: User | None = None) -> bool:
    user = user if user is not None else current_user()
    return bool(user and canonical_role_key(user.role, "") in {"admin", "dev"})


def local_date_input_to_blocked_until_utc(value: Any) -> datetime | None:
    text_value = str(value or "").strip()
    if not text_value:
        return None
    try:
        local_date = datetime.strptime(text_value, "%Y-%m-%d")
    except ValueError:
        return None
    local_end = local_date.replace(hour=23, minute=59, second=59, microsecond=0)
    return local_end - SAO_PAULO_OFFSET


def get_active_product_request_block(user_id: int, product_id: int, conn: Any | None = None) -> ProductRequestBlock | None:
    sql = """
        SELECT prb.*, COALESCE(p.name, '') AS product_name, COALESCE(sr.status, '') AS source_request_status
          FROM product_request_blocks prb
          LEFT JOIN products p ON p.id = prb.product_id
          LEFT JOIN supply_requests sr ON sr.id = prb.created_by_request_id
         WHERE prb.user_id = ?
           AND prb.product_id = ?
           AND prb.revoked_at IS NULL
           AND prb.blocked_until > ?
           AND (prb.created_by_request_id IS NULL OR sr.status IN ('pending', 'approved'))
         LIMIT 1
    """
    params = (int(user_id), int(product_id), now_iso())
    if conn is not None:
        row = conn.execute(sql, params).fetchone()
        return row_to_product_request_block(row)
    with db_connect() as local_conn:
        row = local_conn.execute(sql, params).fetchone()
    return row_to_product_request_block(row)


def list_product_request_blocks_for_user(user_id: int, conn: Any | None = None) -> list[ProductRequestBlock]:
    sql = """
        SELECT prb.*, COALESCE(p.name, '') AS product_name, COALESCE(sr.status, '') AS source_request_status
          FROM product_request_blocks prb
          LEFT JOIN products p ON p.id = prb.product_id
          LEFT JOIN supply_requests sr ON sr.id = prb.created_by_request_id
         WHERE prb.user_id = ?
         ORDER BY (prb.revoked_at IS NULL AND prb.blocked_until > ? AND (prb.created_by_request_id IS NULL OR sr.status IN ('pending', 'approved'))) DESC, prb.blocked_until DESC, prb.id DESC
    """
    params = (int(user_id), now_iso())
    if conn is not None:
        rows = conn.execute(sql, params).fetchall()
    else:
        with db_connect() as local_conn:
            rows = local_conn.execute(sql, params).fetchall()
    return [block for row in rows if (block := row_to_product_request_block(row)) is not None]


def list_products_for_request_block_options(limit: int = 500) -> list[Product]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT *
              FROM products
             WHERE catalog_archived = 0
             ORDER BY name COLLATE NOCASE ASC
             LIMIT ?
            """,
            (bounded_int(limit, 500, 50, 1000),),
        ).fetchall()
    return [product for row in rows if (product := row_to_product(row)) is not None]


def product_requested_in_period(user_id: int, product: Product, conn: Any | None = None) -> int:
    since_dt = product_block_start_for_period(product)

    def query_total(active_conn: Any) -> int:
        nonlocal since_dt
        block_row = active_conn.execute(
            "SELECT revoked_at FROM product_request_blocks WHERE user_id = ? AND product_id = ?",
            (int(user_id), int(product.id)),
        ).fetchone()
        if block_row is not None and block_row["revoked_at"]:
            revoked_dt = parse_dt(block_row["revoked_at"])
            if revoked_dt is not None and revoked_dt > since_dt:
                since_dt = revoked_dt
        since_iso = since_dt.strftime("%Y-%m-%d %H:%M:%S")
        row = active_conn.execute(
            """
            SELECT COALESCE(SUM(ri.quantity), 0) AS total
              FROM request_items ri
              JOIN supply_requests sr ON sr.id = ri.request_id
             WHERE sr.user_id = ?
               AND ri.product_id = ?
               AND sr.status IN ('pending', 'approved')
               AND sr.created_at >= ?
            """,
            (int(user_id), int(product.id), since_iso),
        ).fetchone()
        return stored_quantity_to_limit_quantity(product, (row["total"] if row else 0) or 0)

    if conn is not None:
        return query_total(conn)
    with db_connect() as local_conn:
        return query_total(local_conn)


def upsert_product_request_block(conn: Any, user_id: int, product: Product, request_id: int | None, blocked_until: datetime, reason: str, updated_by_id: int | None = None) -> None:
    conn.execute(
        """
        INSERT INTO product_request_blocks (user_id, product_id, blocked_until, reason, created_by_request_id, created_at, updated_at, revoked_at, updated_by_id)
        VALUES (?, ?, ?, ?, ?, ?, NULL, NULL, ?)
        ON CONFLICT(user_id, product_id) DO UPDATE SET
            blocked_until = excluded.blocked_until,
            reason = excluded.reason,
            created_by_request_id = excluded.created_by_request_id,
            updated_at = excluded.created_at,
            revoked_at = NULL,
            updated_by_id = excluded.updated_by_id
        """,
        (int(user_id), int(product.id), blocked_until.strftime("%Y-%m-%d %H:%M:%S"), reason, request_id, now_iso(), updated_by_id),
    )


def apply_automatic_blocks_after_request(conn: Any, user: User, request_id: int, requested_items: list[tuple[Product, int, int]]) -> None:
    user_id = int(user.id)
    for product, _effective_quantity, _requested_limit_quantity in requested_items:
        limit = product_limit_for(product, user)
        if limit is None or limit <= 0:
            continue
        used = product_requested_in_period(user_id, product, conn)
        if used >= limit:
            blocked_until = datetime.utcnow() + timedelta(days=product_limit_block_days(product))
            reason = f"Limite de {limit} {product_limit_unit_label(product)} atingido automaticamente na solicitação #{request_id}."
            upsert_product_request_block(conn, user_id, product, request_id, blocked_until, reason)


def deactivate_automatic_blocks_for_request(conn: Any, request_id: int, updated_by_id: int | None = None) -> None:
    """Desativa bloqueios criados automaticamente quando a solicitação deixa de contar para o limite."""
    conn.execute(
        """
        UPDATE product_request_blocks
           SET revoked_at = ?, updated_at = ?, updated_by_id = ?
         WHERE created_by_request_id = ?
           AND revoked_at IS NULL
        """,
        (now_iso(), now_iso(), updated_by_id, int(request_id)),
    )


def create_limit_block_from_history(user: User, product: Product, limit: int) -> None:
    try:
        with db_connect() as conn:
            blocked_until = datetime.utcnow() + timedelta(days=product_limit_block_days(product))
            reason = f"Limite de {limit} {product_limit_unit_label(product)} já atingido no período de bloqueio."
            upsert_product_request_block(conn, user.id, product, None, blocked_until, reason)
            conn.commit()
    except Exception:
        app.logger.exception("Falha ao criar bloqueio automatico retroativo")


def apply_user_request_block_form(conn: Any, target_user_id: int, editor: User, form_data: Any) -> None:
    if not can_manage_product_request_blocks(editor):
        return
    existing_blocks = list_product_request_blocks_for_user(target_user_id, conn)
    for block in existing_blocks:
        revoke = str(form_data.get(f"block_revoke_{block.id}") or "").lower() in {"on", "1", "true"}
        if revoke:
            conn.execute(
                "UPDATE product_request_blocks SET revoked_at = ?, updated_at = ?, updated_by_id = ? WHERE id = ? AND user_id = ?",
                (now_iso(), now_iso(), editor.id, block.id, int(target_user_id)),
            )
            continue
        new_until = local_date_input_to_blocked_until_utc(form_data.get(f"block_until_{block.id}"))
        if new_until is not None:
            conn.execute(
                "UPDATE product_request_blocks SET blocked_until = ?, updated_at = ?, revoked_at = NULL, updated_by_id = ? WHERE id = ? AND user_id = ?",
                (new_until.strftime("%Y-%m-%d %H:%M:%S"), now_iso(), editor.id, block.id, int(target_user_id)),
            )
    new_product_id = parse_required_positive_int(form_data.get("new_block_product_id"))
    new_until = local_date_input_to_blocked_until_utc(form_data.get("new_block_until"))
    if new_product_id and new_until is not None:
        product = get_product(new_product_id)
        if product is not None:
            upsert_product_request_block(conn, target_user_id, product, None, new_until, "Bloqueio definido manualmente na edição do usuário.", editor.id)


def effective_product_quantity(product: Product, requested_quantity: int) -> int:
    multiplier = product.kit_quantity if product.is_kit and product.kit_quantity > 1 else 1
    return max(1, int(requested_quantity or 0)) * multiplier


def parse_required_positive_int(value: Any) -> int | None:
    if value is None:
        return None
    text_value = str(value).strip()
    if not text_value:
        return None
    try:
        number = int(text_value)
    except (TypeError, ValueError):
        return None
    if number <= 0:
        return None
    return number


def parse_optional_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        try:
            number = int(round(float(value)))
            return number if number >= 0 else None
        except (TypeError, ValueError, OverflowError):
            return None
    text_value = str(value).strip()
    text_value = re.sub(r"[^0-9,.-]", "", text_value)
    if not text_value:
        return None
    if "," in text_value and "." in text_value:
        if text_value.rfind(",") > text_value.rfind("."):
            text_value = text_value.replace(".", "").replace(",", ".")
        else:
            text_value = text_value.replace(",", "")
    elif "," in text_value:
        if re.fullmatch(r"\d{1,3}(,\d{3})+", text_value):
            text_value = text_value.replace(",", "")
        else:
            text_value = text_value.replace(",", ".")
    elif "." in text_value:
        if text_value.count(".") > 1 or re.fullmatch(r"\d{1,3}(\.\d{3})+", text_value):
            text_value = text_value.replace(".", "")
    try:
        number = int(round(float(text_value)))
    except (TypeError, ValueError, OverflowError):
        return None
    return number if number >= 0 else None


def validate_items_for_user(items_payload: Any, user: User) -> tuple[list[tuple[Product, int, int]], str | None]:
    if not isinstance(items_payload, list):
        return [], "Lista de itens inválida."
    if not items_payload:
        return [], "Adicione pelo menos um insumo ao pedido."

    seen: dict[int, int] = {}
    for raw in items_payload:
        if not isinstance(raw, dict):
            return [], "Item inválido no pedido."
        product_id = parse_required_positive_int(raw.get("product_id"))
        quantity = parse_required_positive_int(raw.get("quantity"))
        if product_id is None or quantity is None:
            return [], "Item inválido no pedido."
        seen[product_id] = seen.get(product_id, 0) + quantity

    normalized: list[tuple[Product, int, int]] = []
    for product_id, quantity in seen.items():
        product = get_product(product_id)
        if product is None or not product.active:
            return [], "Um dos insumos selecionados não está disponível."
        if normalize_stock_tag_slug(product.stock_tag) != SUPPLY_STOCK_TAG:
            return [], f"{product.name} pertence a outro estoque e não está disponível para solicitação de insumos."
        if not user.is_admin and product.stock_quantity <= 0:
            return [], f"{product.name} está sem estoque no momento."
        if not user.is_admin and product.internal:
            return [], f"{product.name} é um produto interno e não está disponível para solicitação."
        if user.role == "base" and not product.visible_base:
            return [], f"{product.name} não está disponível para bases."
        if user.role == "franchise" and not product.visible_franchise:
            return [], f"{product.name} não está disponível para franquias."

        requested_quantity = requested_quantity_to_limit_quantity(product, quantity)

        if not user.is_admin:
            active_block = get_active_product_request_block(user.id, product.id)
            if active_block is not None:
                return [], f"{product.name} está bloqueado para nova solicitação até {active_block.blocked_until_label}."

            limit = product_limit_for(product, user)
            if limit is not None and limit > 0:
                limit_label = product_limit_unit_label(product)
                already_requested = product_requested_in_period(user.id, product)
                if requested_quantity > limit:
                    return [], f"Limite de insumos excedido para {product.name}. Limite permitido: {limit} {limit_label}."
                if already_requested >= limit:
                    create_limit_block_from_history(user, product, limit)
                    return [], f"Você já atingiu o limite de {limit} {limit_label} para {product.name}."
                if already_requested + requested_quantity > limit:
                    remaining = max(0, limit - already_requested)
                    return [], f"Limite de insumos excedido para {product.name}. Restam {remaining} {limit_label} no período de bloqueio."

        effective_quantity = effective_product_quantity(product, requested_quantity)

        minimum = product.min_order_quantity
        if minimum is not None and minimum > 0 and effective_quantity < minimum:
            return [], f"A quantidade mínima para solicitar {product.name} é {minimum}."

        normalized.append((product, effective_quantity, requested_quantity))
    return normalized, None


def fill_product_from_form(product: Product) -> Product:
    product.name = request.form.get("name", "").strip()
    product.category = request.form.get("category", "").strip()
    product.category_emoji = clean_category_emoji(request.form.get("category_emoji"), product.category)
    product.unit_measure = request.form.get("unit_measure", "").strip() or "un"
    product.is_kit = request.form.get("is_kit") == "on"
    product.kit_quantity = parse_required_positive_int(request.form.get("kit_quantity")) or 1
    if not product.is_kit:
        product.kit_quantity = 1
    product.description = request.form.get("description", "").strip()
    product.stock_quantity = parse_required_positive_int(request.form.get("stock_quantity")) or 0
    price_raw = (request.form.get("price") or "0").strip()
    if "," in price_raw:
        price_raw = price_raw.replace(".", "").replace(",", ".")
    try:
        product.price_cents = int(round(float(price_raw or "0") * 100))
    except ValueError:
        product.price_cents = 0
    product.limit_base = parse_optional_int(request.form.get("limit_base"))
    product.limit_franchise = parse_optional_int(request.form.get("limit_franchise"))
    product.limit_block_days = parse_required_positive_int(request.form.get("limit_block_days")) or 60
    product.min_order_quantity = parse_optional_int(request.form.get("min_order_quantity"))
    if product.min_order_quantity is not None and product.min_order_quantity < 1:
        product.min_order_quantity = None
    product.min_stock = parse_optional_int(request.form.get("min_stock"))
    product.max_stock = parse_optional_int(request.form.get("max_stock"))
    product.active = request.form.get("active") == "on"
    product.internal = request.form.get("internal") == "on"
    product.visible_base = request.form.get("visible_base") == "on"
    product.visible_franchise = request.form.get("visible_franchise") == "on"
    if product.internal:
        product.visible_base = False
        product.visible_franchise = False
    if can_manage_stock_tags():
        product.stock_tag = allowed_product_stock_tag(request.form.get("stock_tag"), product.stock_tag or DEFAULT_STOCK_TAG)
        product.stock_tag_name = stock_tag_label(product.stock_tag)
    return product


def product_update_missing_action_permissions(old_product: Product, new_product: Product, uploaded_image_changed: bool = False, remove_image: bool = False) -> list[str]:
    checks: list[tuple[bool, str, str]] = [
        ((old_product.name != new_product.name) or (old_product.description != new_product.description) or uploaded_image_changed or remove_image, "products_edit_basic", "dados principais/imagem"),
        ((old_product.category != new_product.category) or (old_product.category_emoji != new_product.category_emoji), "products_edit_category", "categoria"),
        ((old_product.unit_measure != new_product.unit_measure) or (old_product.is_kit != new_product.is_kit) or (old_product.kit_quantity != new_product.kit_quantity), "products_edit_unit", "unidade de medida/kit"),
        (old_product.stock_quantity != new_product.stock_quantity, "products_edit_stock", "estoque"),
        (old_product.price_cents != new_product.price_cents, "products_edit_price", "preço"),
        ((old_product.limit_base != new_product.limit_base) or (old_product.limit_franchise != new_product.limit_franchise) or (old_product.limit_block_days != new_product.limit_block_days) or (old_product.min_order_quantity != new_product.min_order_quantity) or (old_product.min_stock != new_product.min_stock) or (old_product.max_stock != new_product.max_stock), "products_edit_limits", "limites/regras"),
        ((old_product.active != new_product.active) or (old_product.visible_base != new_product.visible_base) or (old_product.visible_franchise != new_product.visible_franchise) or (old_product.internal != new_product.internal), "products_edit_visibility", "visibilidade/status"),
    ]
    missing: list[str] = []
    for changed, action_key, label in checks:
        if changed and not user_has_action_access(current_user(), action_key):
            missing.append(label)
    return missing


def list_product_categories(user: User | None = None, stock_tag: str = "") -> list[dict[str, str]]:
    clauses = [
        "category IS NOT NULL",
        "TRIM(category) <> ''",
        "catalog_archived = 0",
    ]
    params: list[Any] = []
    tag_filter = normalize_stock_tag_slug(stock_tag, "") if stock_tag else ""
    if user is not None:
        tag_filter = SUPPLY_STOCK_TAG
        if user.role == "base":
            clauses.extend(["visible_base = 1", "COALESCE(internal, 0) = 0", "active = 1", "stock_quantity > 0"])
        elif user.role == "franchise":
            clauses.extend(["visible_franchise = 1", "COALESCE(internal, 0) = 0", "active = 1", "stock_quantity > 0"])
    if tag_filter:
        clauses.append("stock_tag = ?")
        params.append(tag_filter)
    where_sql = " AND ".join(clauses)
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT TRIM(category) AS category,
                   MAX(NULLIF(TRIM(category_emoji), '')) AS category_emoji
             FROM products
             WHERE """ + where_sql + """
             GROUP BY LOWER(TRIM(category))
             ORDER BY category COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
    return [
        {
            "name": str(row["category"]).strip(),
            "emoji": clean_category_emoji(row["category_emoji"], str(row["category"])),
        }
        for row in rows
        if row["category"]
    ]


def product_to_api(product: Product, user: User) -> dict[str, Any]:
    show_stock = user.is_admin
    show_price = user.is_admin or user.role == "franchise"
    limit = product_limit_for(product, user)
    active_block = None if user.is_admin else get_active_product_request_block(user.id, product.id)
    return {
        "id": product.id,
        "name": product.name,
        "category": product.category or "Sem categoria",
        "category_emoji": product.category_emoji or default_category_emoji(product.category),
        "image_url": url_for("product_image", product_id=product.id) if product.image_key else "",
        "unit_measure": product.unit_measure or "un",
        "is_kit": product.is_kit,
        "kit_quantity": product.kit_quantity if product.is_kit else 1,
        "description": product.description or "",
        "price": format_brl(product.price_cents),
        "stock_quantity": product.stock_quantity if show_stock else None,
        "limit": limit,
        "limit_block_days": product.limit_block_days,
        "blocked": active_block is not None,
        "blocked_until": active_block.blocked_until_date_input if active_block else "",
        "blocked_until_label": active_block.blocked_until_label if active_block else "",
        "min_order_quantity": product.min_order_quantity,
        "show_stock": show_stock,
        "show_price": show_price,
        "visible_base": product.visible_base,
        "visible_franchise": product.visible_franchise,
        "internal": product.internal,
    }





def store_generated_file(key: str, buffer: BytesIO, content_type: str, metadata: dict[str, str] | None = None) -> str | None:
    """Stores a generated file in Cloudflare R2 when configured.

    The buffer position is preserved for send_file downloads.
    """
    try:
        position = buffer.tell()
        buffer.seek(0)
        data = buffer.read()
        buffer.seek(position)
        return upload_bytes_to_r2(key, data, content_type, metadata or {})
    except Exception as exc:
        print(f"[R2] Não foi possível armazenar {key}: {exc}")
        try:
            buffer.seek(0)
        except Exception:
            pass
        return None


def storage_key(*parts: Any) -> str:
    clean_parts = [safe_filename(str(part)).strip("_") for part in parts if str(part or "").strip()]
    return "/".join(clean_parts)


def local_product_image_path(key: str) -> Path:
    filename = safe_filename(Path(key or "").stem)
    extension = Path(key or "").suffix.lower()
    if extension not in PRODUCT_IMAGE_TYPES:
        extension = ".img"
    return PRODUCT_IMAGE_DIR / f"{filename}{extension}"


def save_product_image_upload(uploaded: Any) -> tuple[str, str, str]:
    original_name = str(getattr(uploaded, "filename", "") or "").strip()
    extension = Path(original_name).suffix.lower()
    if extension not in PRODUCT_IMAGE_TYPES:
        raise ValueError("Use uma imagem PNG, JPG, JPEG, WEBP ou GIF.")
    data = uploaded.read()
    if not data:
        raise ValueError("A imagem selecionada está vazia.")
    if len(data) > PRODUCT_IMAGE_MAX_BYTES:
        raise ValueError("A imagem deve ter no máximo 3 MB.")
    content_type = PRODUCT_IMAGE_TYPES[extension]
    token = "".join(random.choice(string.ascii_lowercase + string.digits) for _ in range(12))
    key = storage_key("product_images", sao_paulo_filename_timestamp(), token) + extension
    local_path = local_product_image_path(key)
    local_path.write_bytes(data)
    try:
        upload_bytes_to_r2(key, data, content_type, {"type": "product_image"})
    except Exception as exc:
        print(f"[R2] Não foi possível salvar imagem do produto: {exc}")
    return original_name[:180], key[:500], content_type


def remove_local_product_image(key: str) -> None:
    if not key:
        return
    path = local_product_image_path(key)
    try:
        if path.exists():
            path.unlink()
    except OSError:
        pass


def pdf_clean_text(value: Any, default: str = "-") -> str:
    """Texto seguro para Paragraph do ReportLab.

    Alguns nomes antigos de base/franquia chegaram ao PDF com ponto e vírgula
    no lugar de espaços. Aqui normalizamos só para exibição, mantendo o valor
    salvo no banco intacto.
    """
    text_value = str(value if value is not None else "").strip()
    if not text_value:
        text_value = default
    text_value = text_value.replace(";", " ")
    text_value = re.sub(r"\s+", " ", text_value).strip()
    return html_escape(text_value or default)


def pdf_clean_plain_text(value: Any, default: str = "-") -> str:
    """Texto limpo para canvas.drawString, sem escapar HTML."""
    text_value = str(value if value is not None else "").strip()
    if not text_value:
        text_value = default
    text_value = text_value.replace(";", " ")
    text_value = re.sub(r"\s+", " ", text_value).strip()
    return text_value or default


def pdf_mixed_text(value: Any, default: str = "-") -> str:
    """Mantém tipografia latina compacta e usa fonte CJK só nos ideogramas."""
    text_value = pdf_clean_plain_text(value, default)
    cjk_pattern = re.compile(r"([\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff\u3000-\u303f]+)")
    parts: list[str] = []
    for part in cjk_pattern.split(text_value):
        if not part:
            continue
        escaped = html_escape(part)
        if cjk_pattern.fullmatch(part):
            parts.append(f'<font name="{PDF_CJK_FONT}">{escaped}</font>')
        else:
            parts.append(escaped)
    return "".join(parts) or html_escape(default)


def people_count_label(value: int | None) -> str:
    if value is None or int(value or 0) <= 0:
        return "-"
    number = int(value)
    return f"{number} pessoa" if number == 1 else f"{number} pessoas"


def build_request_pdf(supply_request: SupplyRequest, viewer: User) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=17 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="JTEyebrow", fontName=REQUEST_PDF_FONT_BOLD, fontSize=7.2, leading=9, textColor=colors.HexColor("#e60012")))
    styles.add(ParagraphStyle(name="JTTitle", fontName=REQUEST_PDF_FONT_BOLD, fontSize=17, leading=20, textColor=colors.HexColor("#171717"), spaceAfter=2))
    styles.add(ParagraphStyle(name="JTSubtitle", fontName=REQUEST_PDF_FONT, fontSize=8.3, leading=11.2, textColor=colors.HexColor("#666666")))
    styles.add(ParagraphStyle(name="JTSection", fontName=REQUEST_PDF_FONT_BOLD, fontSize=11.2, leading=14, textColor=colors.HexColor("#171717"), spaceBefore=3, spaceAfter=6))
    styles.add(ParagraphStyle(name="JTMeta", fontName=REQUEST_PDF_FONT, fontSize=8.4, leading=10.6, textColor=colors.HexColor("#252525")))
    styles.add(ParagraphStyle(name="JTMetaLabel", fontName=REQUEST_PDF_FONT_BOLD, fontSize=7.0, leading=8.5, textColor=colors.HexColor("#e60012")))
    styles.add(ParagraphStyle(name="JTCell", fontName=REQUEST_PDF_FONT, fontSize=8.2, leading=10.6, textColor=colors.HexColor("#252525")))
    styles.add(ParagraphStyle(name="JTCellBold", fontName=REQUEST_PDF_FONT_BOLD, fontSize=8.2, leading=10.6, textColor=colors.HexColor("#151515")))
    styles.add(ParagraphStyle(name="JTSmall", fontName=REQUEST_PDF_FONT, fontSize=7.7, leading=10.2, textColor=colors.HexColor("#666666")))

    def p(value: Any, style_name: str = "JTCell") -> Paragraph:
        return Paragraph(pdf_mixed_text(value), styles[style_name])

    story: list[Any] = []

    logo_path = BASE_DIR / "static" / "img" / "logo-jt-red.svg"
    logo_cell: Any
    try:
        drawing = svg2rlg(str(logo_path))
        if drawing is not None and drawing.width:
            scale = (31 * mm) / drawing.width
            drawing.width *= scale
            drawing.height *= scale
            drawing.scale(scale, scale)
            logo_cell = drawing
        else:
            raise ValueError("Logo SVG inválida")
    except Exception:
        logo_cell = Paragraph("J&amp;T EXPRESS", ParagraphStyle(name="FallbackLogoPDF", fontName=REQUEST_PDF_FONT_BOLD, fontSize=14, textColor=colors.HexColor("#e60012")))

    requester = supply_request.user
    requester_name = requester.responsible_name if requester else "-"
    requester_org = requester.organization_name if requester else "-"
    requester_username = requester.username if requester else "-"
    requester_role = requester.role if requester else "base"
    show_prices = viewer.is_admin or viewer.role == "franchise"
    role_label = "Administrador" if requester_role == "admin" else ("Base" if requester_role == "base" else "Franquia")

    title_block = [
        Paragraph("PORTAL DE INSUMOS", styles["JTEyebrow"]),
        Paragraph(f"Solicitação #{supply_request.id}", styles["JTTitle"]),
        Paragraph("J&amp;T Express Brazil", styles["JTSubtitle"]),
    ]
    status_block = Paragraph(
        f"<font color='#e60012'><b>{pdf_mixed_text(status_label(supply_request.status))}</b></font>"
        f"<br/><font color='#777777'>Emitido em {format_sao_paulo_datetime(suffix=' GMT-3')}</font>",
        styles["JTMeta"],
    )
    header_table = Table(
        [[
            logo_cell,
            title_block,
            status_block,
        ]],
        colWidths=[39 * mm, 86 * mm, 55 * mm],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7f8")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#edcfd3")),
        ("LINEBELOW", (0, 0), (-1, -1), 2.2, colors.HexColor("#e60012")),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 5 * mm))

    primary_meta = Table(
        [
            [Paragraph("SOLICITANTE", styles["JTMetaLabel"]), Paragraph("BASE / FRANQUIA", styles["JTMetaLabel"]), Paragraph("TIPO", styles["JTMetaLabel"])],
            [p(requester_name, "JTMeta"), p(requester_org, "JTMeta"), p(role_label, "JTMeta")],
        ],
        colWidths=[60 * mm, 60 * mm, 60 * mm],
    )
    primary_meta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fcfcfc")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(primary_meta)
    story.append(Spacer(1, 2 * mm))

    secondary_fields = [
        ("USUÁRIO", requester_username),
        ("DATA DA SOLICITAÇÃO", format_sao_paulo_datetime(supply_request.created_at)),
    ]
    if requester_role == "base":
        secondary_fields.insert(1, ("PESSOAS NA BASE", people_count_label(supply_request.people_count)))
    secondary_width = 180 * mm / len(secondary_fields)
    secondary_meta = Table(
        [
            [Paragraph(label, styles["JTMetaLabel"]) for label, _value in secondary_fields],
            [p(value, "JTMeta") for _label, value in secondary_fields],
        ],
        colWidths=[secondary_width] * len(secondary_fields),
    )
    secondary_meta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fcfcfc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(secondary_meta)
    story.append(Spacer(1, 7 * mm))

    headers = [p("Produto solicitado", "JTCellBold"), p("Qtd.", "JTCellBold")]
    col_widths = [126 * mm, 28 * mm]
    if show_prices:
        headers.extend([p("Valor unit.", "JTCellBold"), p("Subtotal", "JTCellBold")])
        col_widths = [86 * mm, 23 * mm, 30 * mm, 35 * mm]

    item_rows: list[list[Any]] = [headers]
    for item in supply_request.items:
        unit_label = item.product.unit_measure if item.product and item.product.unit_measure else "un"
        row: list[Any] = [
            p(item.product_name_snapshot),
            p(f"{item.quantity} {unit_label}", "JTCellBold"),
        ]
        if show_prices:
            row.extend([
                p(format_brl(item.price_cents_snapshot)),
                p(format_brl(item.price_cents_snapshot * item.quantity), "JTCellBold"),
            ])
        item_rows.append(row)

    if show_prices:
        item_rows.append(["", "", p("TOTAL", "JTCellBold"), p(format_brl(supply_request.total_cents), "JTCellBold")])

    story.append(Paragraph("Itens solicitados", styles["JTSection"]))
    items_table = Table(item_rows, colWidths=col_widths, repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d90012")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dddddd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#e8e8e8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
    ]
    if show_prices:
        style_commands.extend([
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fff3f4")),
            ("SPAN", (0, -1), (1, -1)),
        ])
    style_commands.append(("ROWBACKGROUNDS", (0, 1), (-1, -2 if show_prices else -1), [colors.white, colors.HexColor("#fafafa")]))
    items_table.setStyle(TableStyle(style_commands))
    story.append(items_table)
    story.append(Spacer(1, 8 * mm))

    info_blocks: list[list[Any]] = []
    if supply_request.user_note:
        info_blocks.append([p("Observação do solicitante", "JTCellBold"), p(supply_request.user_note, "JTSmall")])
    if supply_request.admin_note:
        info_blocks.append([p("Observação administrativa", "JTCellBold"), p(supply_request.admin_note, "JTSmall")])
    if supply_request.reviewed_at:
        info_blocks.append([p("Revisão", "JTCellBold"), p(f"Revisado em {format_sao_paulo_datetime(supply_request.reviewed_at)}", "JTSmall")])
    if info_blocks:
        story.append(Paragraph("Informações complementares", styles["JTSection"]))
        info_table = Table(info_blocks, colWidths=[52 * mm, 122 * mm])
        info_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
            ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fff8f8")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(info_table)

    def footer(canvas, document):
        canvas.saveState()
        width, _height = A4
        canvas.setStrokeColor(colors.HexColor("#e60012"))
        canvas.setLineWidth(0.7)
        canvas.line(16 * mm, 14 * mm, width - 16 * mm, 14 * mm)
        canvas.setFont(REQUEST_PDF_FONT, 7.5)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.drawString(16 * mm, 9 * mm, "J&T Express Brazil • CNPJ: 42.584.754/0092-13")
        canvas.drawRightString(width - 16 * mm, 9 * mm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return buffer



def parse_report_date(value: str, field_label: str) -> datetime:
    """Converte data yyyy-mm-dd do formulário em datetime.

    Mantemos a filtragem por texto/UTC do banco, mas o usuário escolhe datas
    simples no formulário. O intervalo final é tratado como inclusivo na rota.
    """
    try:
        return datetime.strptime((value or "").strip(), "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"{field_label} inválida. Use uma data válida.") from exc


def list_supply_requests_between(start_dt: datetime, end_dt: datetime, organization_name: str = "") -> list[SupplyRequest]:
    start_sql = start_dt.strftime("%Y-%m-%d %H:%M:%S")
    end_sql = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    params: list[Any] = [start_sql, end_sql]
    unit_filter = (organization_name or "").strip()
    unit_clause = ""
    if unit_filter:
        unit_clause = " AND u.organization_name = ?"
        params.append(unit_filter)
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT sr.*
              FROM supply_requests sr
              LEFT JOIN users u ON u.id = sr.user_id
             WHERE sr.created_at >= ?
               AND sr.created_at <= ?
               {unit_clause}
             ORDER BY sr.created_at ASC, sr.id ASC
            """,
            params,
        ).fetchall()
    return [req for row in rows if (req := row_to_supply_request(row)) is not None]


def list_assets_between(start_dt: datetime, end_dt: datetime, unit_name: str = "") -> list[AssetRecord]:
    start_sql = start_dt.strftime("%Y-%m-%d %H:%M:%S")
    end_sql = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    unit_filter = (unit_name or "").strip()
    unit_clause = ""
    params: list[Any] = [start_sql, end_sql]
    if unit_filter:
        unit_clause = " AND base = ?"
        params.append(unit_filter)
    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT *
              FROM assets
             WHERE created_at >= ?
               AND created_at <= ?
               {unit_clause}
             ORDER BY created_at ASC, id ASC
            """,
            params,
        ).fetchall()
        return [asset for row in rows if (asset := row_to_asset(row, conn=conn)) is not None]


def build_supply_requests_period_report_pdf(
    requests_list: list[SupplyRequest],
    start_dt: datetime,
    end_dt: datetime,
    viewer: User,
    unit_name: str = "",
    unit_kind: str = "",
) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", fontName=PDF_TEXT_FONT_BOLD, fontSize=18.5, leading=22.5, textColor=colors.HexColor("#151515"), spaceAfter=3))
    styles.add(ParagraphStyle(name="ReportSubtitle", fontName=PDF_TEXT_FONT, fontSize=9.0, leading=12.2, textColor=colors.HexColor("#666666")))
    styles.add(ParagraphStyle(name="ReportSection", fontName=PDF_TEXT_FONT_BOLD, fontSize=12.0, leading=15, textColor=colors.HexColor("#151515"), spaceBefore=6, spaceAfter=7))
    styles.add(ParagraphStyle(name="ReportLabel", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.2, leading=9, textColor=colors.HexColor("#e60012")))
    styles.add(ParagraphStyle(name="ReportMeta", fontName=PDF_TEXT_FONT, fontSize=8.0, leading=10.2, textColor=colors.HexColor("#222222")))
    styles.add(ParagraphStyle(name="ReportCell", fontName=PDF_TEXT_FONT, fontSize=7.4, leading=9.5, textColor=colors.HexColor("#222222")))
    styles.add(ParagraphStyle(name="ReportCellBold", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.4, leading=9.5, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="ReportSmall", fontName=PDF_TEXT_FONT, fontSize=6.8, leading=8.6, textColor=colors.HexColor("#666666")))

    def p(value: Any, style_name: str = "ReportCell") -> Paragraph:
        return Paragraph(pdf_clean_text(value), styles[style_name])

    def status_name(status: str) -> str:
        return status_label(status)

    logo_path = BASE_DIR / "static" / "img" / "logo-jt-red.svg"
    try:
        drawing = svg2rlg(str(logo_path))
        if drawing is not None and drawing.width:
            scale = (34 * mm) / drawing.width
            drawing.width *= scale
            drawing.height *= scale
            drawing.scale(scale, scale)
            logo_cell: Any = drawing
        else:
            raise ValueError("Logo SVG inválida")
    except Exception:
        logo_cell = Paragraph("J&amp;T EXPRESS", ParagraphStyle(name="ReportFallbackLogo", fontName=PDF_TEXT_FONT_BOLD, fontSize=13, textColor=colors.HexColor("#e60012")))

    period_label = f"{start_dt.strftime('%d/%m/%Y')} a {end_dt.strftime('%d/%m/%Y')}"
    unit_label = f"{unit_kind_label(unit_kind)}: {unit_name}" if unit_name else "Todas as unidades"
    generated_label = format_sao_paulo_datetime(suffix=" GMT-3")
    story: list[Any] = []
    header = Table(
        [[
            logo_cell,
            Paragraph("Relatório de Solicitações de Insumos", styles["ReportTitle"]),
            Paragraph(f"<b>Período</b><br/>{pdf_clean_text(period_label)}<br/><b>Unidade</b><br/>{pdf_clean_text(unit_label)}<br/><font color='#777777'>Gerado em {generated_label}</font>", styles["ReportMeta"]),
        ]],
        colWidths=[36 * mm, 92 * mm, 50 * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("LINEBELOW", (0, 0), (-1, -1), 2.0, colors.HexColor("#e60012")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(f"Relatório consolidado das solicitações registradas no intervalo selecionado para {pdf_clean_text(unit_label)}, incluindo solicitante, status, pessoas das bases, itens, quantidades, valores e observações.", styles["ReportSubtitle"]))
    story.append(Spacer(1, 6 * mm))

    status_counts: dict[str, int] = {}
    total_items = 0
    total_units = 0
    total_people = 0
    total_value = 0
    requester_units: set[str] = set()
    product_totals: dict[str, dict[str, Any]] = {}
    for req in requests_list:
        status_counts[req.status] = status_counts.get(req.status, 0) + 1
        if req.user and req.user.role == "base":
            total_people += int(req.people_count or 0)
        if req.user and req.user.organization_name:
            requester_units.add(pdf_clean_plain_text(req.user.organization_name))
        for item in req.items:
            total_items += 1
            total_units += int(item.quantity or 0)
            total_value += int(item.price_cents_snapshot or 0) * int(item.quantity or 0)
            product_key = item.product_name_snapshot or f"Produto #{item.product_id}"
            bucket = product_totals.setdefault(product_key, {"name": product_key, "quantity": 0, "value": 0, "unit": "un"})
            bucket["quantity"] += int(item.quantity or 0)
            bucket["value"] += int(item.price_cents_snapshot or 0) * int(item.quantity or 0)
            if item.product and item.product.unit_measure:
                bucket["unit"] = item.product.unit_measure

    summary_data = [
        [Paragraph("SOLICITAÇÕES", styles["ReportLabel"]), Paragraph("BASES/FRANQUIAS", styles["ReportLabel"]), Paragraph("ITENS", styles["ReportLabel"]), Paragraph("VALOR TOTAL", styles["ReportLabel"])],
        [p(len(requests_list), "ReportMeta"), p(len(requester_units), "ReportMeta"), p(f"{total_items} linhas / {total_units} un.", "ReportMeta"), p(format_brl(total_value), "ReportMeta")],
        [Paragraph("PENDENTES", styles["ReportLabel"]), Paragraph("APROVADAS", styles["ReportLabel"]), Paragraph("REJEITADAS", styles["ReportLabel"]), Paragraph("PESSOAS NAS BASES", styles["ReportLabel"])],
        [p(status_counts.get("pending", 0), "ReportMeta"), p(status_counts.get("approved", 0), "ReportMeta"), p(status_counts.get("rejected", 0), "ReportMeta"), p(total_people if total_people else "-", "ReportMeta")],
    ]
    summary_table = Table(summary_data, colWidths=[44.5 * mm, 44.5 * mm, 44.5 * mm, 44.5 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#fff3f4")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6 * mm))

    if product_totals:
        top_products = sorted(product_totals.values(), key=lambda row: row["quantity"], reverse=True)[:12]
        story.append(Paragraph("Resumo por produto", styles["ReportSection"]))
        product_rows = [[p("Produto", "ReportCellBold"), p("Quantidade", "ReportCellBold"), p("Valor", "ReportCellBold")]]
        for product in top_products:
            product_rows.append([
                p(product["name"]),
                p(f"{product['quantity']} {product.get('unit') or 'un'}", "ReportCellBold"),
                p(format_brl(product["value"]), "ReportCellBold"),
            ])
        product_table = Table(product_rows, colWidths=[108 * mm, 34 * mm, 36 * mm], repeatRows=1)
        product_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e60012")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff8f8")]),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dddddd")),
            ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ]))
        story.append(product_table)
        story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Solicitações do período", styles["ReportSection"]))
    if not requests_list:
        empty = Table([[Paragraph("Nenhuma solicitação encontrada para o período selecionado.", styles["ReportMeta"])]], colWidths=[178 * mm])
        empty.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff8f8")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(empty)
    else:
        for req in requests_list:
            requester = req.user
            requester_name = requester.responsible_name if requester else "-"
            requester_org = requester.organization_name if requester else "-"
            requester_username = requester.username if requester else "-"
            requester_role = user_role_label(requester.role if requester else "")
            requester_role_value = requester_role
            requester_role_label = "Tipo"
            if requester and requester.role == "base":
                requester_role_label = "Tipo / Pessoas"
                requester_role_value = f"{requester_role} • {people_count_label(req.people_count)}"
            item_text_parts = []
            for item in req.items:
                unit = item.product.unit_measure if item.product and item.product.unit_measure else "un"
                subtotal = int(item.price_cents_snapshot or 0) * int(item.quantity or 0)
                item_text_parts.append(f"{item.quantity} {unit} - {item.product_name_snapshot} ({format_brl(subtotal)})")
            if not item_text_parts:
                item_text_parts.append("Sem itens")
            notes_html = "-"
            if req.user_note and req.admin_note:
                notes_html = f"Solicitante: {pdf_clean_text(req.user_note)}<br/>Admin: {pdf_clean_text(req.admin_note)}"
            elif req.user_note:
                notes_html = pdf_clean_text(req.user_note)
            elif req.admin_note:
                notes_html = pdf_clean_text(req.admin_note)

            detail_rows = [
                [Paragraph(f"Solicitação #{req.id}", styles["ReportCellBold"]), Paragraph(pdf_clean_text(status_name(req.status)), styles["ReportCellBold"]), Paragraph(format_sao_paulo_datetime(req.created_at), styles["ReportCellBold"])],
                [Paragraph("Solicitante", styles["ReportLabel"]), Paragraph("Setor", styles["ReportLabel"]), Paragraph(requester_role_label, styles["ReportLabel"])],
                [p(f"{requester_name} ({requester_username})"), p(requester_org), p(requester_role_value)],
                [Paragraph("Itens", styles["ReportLabel"]), Paragraph("Observações", styles["ReportLabel"]), Paragraph("Total", styles["ReportLabel"])],
                [Paragraph("<br/>".join(pdf_clean_text(part) for part in item_text_parts), styles["ReportSmall"]), Paragraph(notes_html, styles["ReportSmall"]), p(format_brl(req.total_cents), "ReportCellBold")],
            ]
            if req.reviewed_at:
                detail_rows.append([Paragraph("Revisão", styles["ReportLabel"]), Paragraph(format_sao_paulo_datetime(req.reviewed_at), styles["ReportSmall"]), Paragraph("", styles["ReportSmall"])])
            detail = Table(detail_rows, colWidths=[68 * mm, 66 * mm, 44 * mm], repeatRows=0)
            detail.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fffafa")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#fffafa")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#e6d6d8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#eeeeee")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(detail)
            story.append(Spacer(1, 3.5 * mm))

    def footer(canvas, document):
        canvas.saveState()
        width, _height = A4
        canvas.setStrokeColor(colors.HexColor("#e60012"))
        canvas.setLineWidth(0.7)
        canvas.line(16 * mm, 13 * mm, width - 16 * mm, 13 * mm)
        canvas.setFont(PDF_TEXT_FONT, 7.2)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.drawString(16 * mm, 8.5 * mm, "J&T Express Brazil • Relatório mensal de solicitações")
        canvas.drawRightString(width - 16 * mm, 8.5 * mm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return buffer

def build_assets_period_report_pdf(
    assets: list[AssetRecord],
    start_dt: datetime,
    end_dt: datetime,
    viewer: User,
    unit_name: str,
    unit_kind: str,
) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="AssetReportTitle", fontName=PDF_TEXT_FONT_BOLD, fontSize=18.5, leading=22, textColor=colors.HexColor("#141414"), spaceAfter=2))
    styles.add(ParagraphStyle(name="AssetReportSubtitle", fontName=PDF_TEXT_FONT, fontSize=8.8, leading=12, textColor=colors.HexColor("#666666")))
    styles.add(ParagraphStyle(name="AssetReportLabel", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.2, leading=9, textColor=colors.HexColor("#e60012")))
    styles.add(ParagraphStyle(name="AssetReportCell", fontName=PDF_TEXT_FONT, fontSize=7.5, leading=9.8, textColor=colors.HexColor("#222222")))
    styles.add(ParagraphStyle(name="AssetReportCellBold", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.6, leading=9.8, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="AssetReportSection", fontName=PDF_TEXT_FONT_BOLD, fontSize=12.0, leading=15, textColor=colors.HexColor("#111111"), spaceBefore=7, spaceAfter=7))
    styles.add(ParagraphStyle(name="AssetReportSmall", fontName=PDF_TEXT_FONT, fontSize=6.8, leading=8.8, textColor=colors.HexColor("#666666")))

    def p(value: Any, style_name: str = "AssetReportCell") -> Paragraph:
        return Paragraph(pdf_clean_text(value), styles[style_name])

    logo_path = BASE_DIR / "static" / "img" / "logo-jt-red.svg"
    try:
        drawing = svg2rlg(str(logo_path))
        if drawing is not None and drawing.width:
            scale = (34 * mm) / drawing.width
            drawing.width *= scale
            drawing.height *= scale
            drawing.scale(scale, scale)
            logo_cell: Any = drawing
        else:
            raise ValueError("Logo SVG inválida")
    except Exception:
        logo_cell = Paragraph("J&amp;T EXPRESS", ParagraphStyle(name="AssetReportFallbackLogo", fontName=PDF_TEXT_FONT_BOLD, fontSize=13, textColor=colors.HexColor("#e60012")))

    period_label = f"{start_dt.strftime('%d/%m/%Y')} a {end_dt.strftime('%d/%m/%Y')}"
    unit_label = f"{unit_kind_label(unit_kind)}: {unit_name}" if unit_name else "Todas as unidades"
    generated_label = format_sao_paulo_datetime(suffix=" GMT-3")
    total_assets = len(assets)
    total_item_lines = sum(len(asset.items) for asset in assets)
    total_quantity = sum(sum(max(0, int(item.quantity or 0)) for item in asset.items) for asset in assets)
    sectors = len({asset.sector for asset in assets if asset.sector})
    managers = len({asset.manager for asset in assets if asset.manager})

    story: list[Any] = []
    header = Table(
        [[
            logo_cell,
            Paragraph("Relatório de Ativos", styles["AssetReportTitle"]),
            Paragraph(f"<b>Período</b><br/>{pdf_clean_text(period_label)}<br/><b>Unidade</b><br/>{pdf_clean_text(unit_label)}<br/><font color='#777777'>Gerado em {generated_label}</font>", styles["AssetReportCell"]),
        ]],
        colWidths=[36 * mm, 90 * mm, 52 * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("LINEBELOW", (0, 0), (-1, -1), 2.0, colors.HexColor("#e60012")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("Relatório consolidado dos ativos cadastrados no período selecionado, com dados da unidade, setor, gestor e itens vinculados.", styles["AssetReportSubtitle"]))
    story.append(Spacer(1, 6 * mm))

    summary = Table([
        [p("ATIVOS", "AssetReportLabel"), p("LINHAS DE ITENS", "AssetReportLabel"), p("QUANTIDADE TOTAL", "AssetReportLabel"), p("SETORES", "AssetReportLabel"), p("GESTORES", "AssetReportLabel")],
        [p(total_assets, "AssetReportCellBold"), p(total_item_lines, "AssetReportCellBold"), p(total_quantity, "AssetReportCellBold"), p(sectors, "AssetReportCellBold"), p(managers, "AssetReportCellBold")],
    ], colWidths=[35.6 * mm] * 5)
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#eeeeee")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(summary)
    story.append(Spacer(1, 7 * mm))

    story.append(Paragraph("Ativos do período", styles["AssetReportSection"]))
    if not assets:
        empty = Table([[Paragraph("Nenhum ativo encontrado para o período e unidade selecionados.", styles["AssetReportCell"])]], colWidths=[178 * mm])
        empty.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#eeeeee")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff8f8")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(empty)
    else:
        for asset in assets:
            item_parts = []
            for item in asset.items:
                product = get_product(item.product_id) if item.product_id else None
                unit = product.unit_measure if product and product.unit_measure else "un"
                serial = item.serial_number or "Sem patrimônio/série"
                item_parts.append(f"{item.quantity} {unit} - {item.item_name} • {serial}")
            if not item_parts:
                item_parts.append("Sem itens")
            detail_rows = [
                [p(f"Ativo #{asset.id}", "AssetReportCellBold"), p(asset.name, "AssetReportCellBold"), p(format_sao_paulo_datetime(asset.created_at), "AssetReportCellBold")],
                [p("Setor", "AssetReportLabel"), p("Setor", "AssetReportLabel"), p("Gestor", "AssetReportLabel")],
                [p(asset.base), p(asset.sector), p(asset.manager)],
                [p("Itens vinculados", "AssetReportLabel"), p("Regional", "AssetReportLabel"), p("Total do ativo", "AssetReportLabel")],
                [Paragraph("<br/>".join(pdf_clean_text(part) for part in item_parts), styles["AssetReportSmall"]), p(asset.regional), p(str(sum(max(0, int(item.quantity or 0)) for item in asset.items)), "AssetReportCellBold")],
            ]
            detail = Table(detail_rows, colWidths=[82 * mm, 54 * mm, 42 * mm])
            detail.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff3f4")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fffafa")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#fffafa")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#e6d6d8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#eeeeee")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(detail)
            story.append(Spacer(1, 3.5 * mm))

    def footer(canvas, document):
        canvas.saveState()
        width, _height = A4
        canvas.setStrokeColor(colors.HexColor("#e60012"))
        canvas.setLineWidth(0.7)
        canvas.line(16 * mm, 13 * mm, width - 16 * mm, 13 * mm)
        canvas.setFont(PDF_TEXT_FONT, 7.2)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.drawString(16 * mm, 8.5 * mm, "J&T Express Brazil • Relatório mensal de ativos")
        canvas.drawRightString(width - 16 * mm, 8.5 * mm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return buffer


def build_asset_pdf(asset: AssetRecord, viewer: User) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="AssetTitle", fontName=PDF_TEXT_FONT_BOLD, fontSize=20, leading=24, textColor=colors.HexColor("#111111"), spaceAfter=2))
    styles.add(ParagraphStyle(name="AssetSubtitle", fontName=PDF_TEXT_FONT, fontSize=8.7, leading=11.5, textColor=colors.HexColor("#666666")))
    styles.add(ParagraphStyle(name="AssetLabel", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.2, leading=8.6, textColor=colors.HexColor("#e60012"), uppercase=True))
    styles.add(ParagraphStyle(name="AssetValue", fontName=PDF_TEXT_FONT_BOLD, fontSize=9.4, leading=12, textColor=colors.HexColor("#141414")))
    styles.add(ParagraphStyle(name="AssetCell", fontName=PDF_TEXT_FONT, fontSize=8.6, leading=11.2, textColor=colors.HexColor("#252525")))
    styles.add(ParagraphStyle(name="AssetCellBold", fontName=PDF_TEXT_FONT_BOLD, fontSize=8.6, leading=11.2, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="AssetSmall", fontName=PDF_TEXT_FONT, fontSize=7.5, leading=9.6, textColor=colors.HexColor("#777777")))
    styles.add(ParagraphStyle(name="AssetSection", fontName=PDF_TEXT_FONT_BOLD, fontSize=12, leading=14.5, textColor=colors.HexColor("#111111"), spaceBefore=2, spaceAfter=6))

    story: list[Any] = []

    logo_path = BASE_DIR / "static" / "img" / "logo-jt-red.svg"
    try:
        drawing = svg2rlg(str(logo_path))
        if drawing is None or not drawing.width:
            raise ValueError("Logo SVG invalida")
        scale = (38 * mm) / drawing.width
        drawing.width *= scale
        drawing.height *= scale
        drawing.scale(scale, scale)
        logo_cell: Any = drawing
    except Exception:
        logo_cell = Paragraph("J&amp;T EXPRESS", ParagraphStyle(name="AssetFallbackLogo", fontName=PDF_TEXT_FONT_BOLD, fontSize=15, textColor=colors.HexColor("#e60012")))

    total_quantity = sum(max(0, int(item.quantity or 0)) for item in asset.items)
    total_lines = len(asset.items)
    created_by = get_user(asset.created_by_id) if asset.created_by_id else None
    created_by_name = created_by.responsible_name if created_by else "Portal de Insumos"
    generated_at = format_sao_paulo_datetime(suffix=" GMT-3")

    header = Table(
        [[
            logo_cell,
            Paragraph(f"Ficha de Ativo #{asset.id}", styles["AssetTitle"]),
            Paragraph(f"<b>Gerado em</b><br/>{generated_at}<br/><font color='#777777'>Portal de Insumos J&amp;T Express</font>", styles["AssetCell"]),
        ]],
        colWidths=[43 * mm, 82 * mm, 55 * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.9, colors.HexColor("#eeeeee")),
        ("LINEBELOW", (0, 0), (-1, -1), 2.2, colors.HexColor("#e60012")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    story.append(header)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("Documento patrimonial gerado automaticamente para controle interno de ativos vinculados a bases e franquias.", styles["AssetSubtitle"]))
    story.append(Spacer(1, 7 * mm))

    def label_cell(text: str) -> Paragraph:
        return Paragraph(pdf_clean_text(text), styles["AssetLabel"])

    def value_cell(value: Any) -> Paragraph:
        return Paragraph(pdf_clean_text(value), styles["AssetValue"])

    info_data = [
        [label_cell("ATIVO"), label_cell("BASE / FRANQUIA"), label_cell("REGIONAL")],
        [value_cell(asset.name), value_cell(asset.base), value_cell(asset.regional)],
        [label_cell("SETOR"), label_cell("GESTOR RESPONSAVEL"), label_cell("CADASTRADO POR")],
        [value_cell(asset.sector), value_cell(asset.manager), value_cell(created_by_name)],
        [label_cell("DATA DO CADASTRO"), label_cell("TOTAL DE ITENS"), label_cell("LINHAS DE ITENS")],
        [value_cell(format_sao_paulo_datetime(asset.created_at)), value_cell(str(total_quantity)), value_cell(str(total_lines))],
    ]
    info_table = Table(info_data, colWidths=[60 * mm, 60 * mm, 60 * mm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fcfcfc")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff1f2")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#fff1f2")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#fff1f2")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#e6e6e6")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#e9e9e9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6.2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6.2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Itens vinculados ao ativo", styles["AssetSection"]))
    item_rows: list[list[Any]] = [[
        Paragraph("#", styles["AssetCellBold"]),
        Paragraph("Item", styles["AssetCellBold"]),
        Paragraph("Quantidade", styles["AssetCellBold"]),
        Paragraph("Patrimonio / Serie", styles["AssetCellBold"]),
    ]]
    for index, item in enumerate(asset.items, start=1):
        product = get_product(item.product_id) if item.product_id else None
        unit = product.unit_measure if product and product.unit_measure else "un"
        quantity_text = f"{item.quantity} {unit}"
        item_rows.append([
            Paragraph(str(index), styles["AssetCellBold"]),
            Paragraph(pdf_clean_text(item.item_name), styles["AssetCell"]),
            Paragraph(pdf_clean_text(quantity_text), styles["AssetCellBold"]),
            Paragraph(pdf_clean_text(item.serial_number or "Sem patrimonio/serie"), styles["AssetCell"]),
        ])
    if len(item_rows) == 1:
        item_rows.append(["", Paragraph("Nenhum item cadastrado.", styles["AssetCell"]), "", ""])

    items_table = Table(item_rows, colWidths=[13 * mm, 87 * mm, 29 * mm, 51 * mm], repeatRows=1)
    item_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e60012")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dddddd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#e8e8e8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
    ]
    for row_index in range(1, len(item_rows), 2):
        item_styles.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#fbfbfb")))
    items_table.setStyle(TableStyle(item_styles))
    story.append(items_table)
    story.append(Spacer(1, 10 * mm))

    signature_table = Table(
        [[
            Paragraph("Responsavel pela unidade", styles["AssetSmall"]),
            Paragraph("Conferencia administrativa", styles["AssetSmall"]),
        ], [
            Paragraph("__________________________________________", styles["AssetCell"]),
            Paragraph("__________________________________________", styles["AssetCell"]),
        ]],
        colWidths=[87 * mm, 87 * mm],
    )
    signature_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(signature_table)

    def footer(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e60012"))
        canvas.setLineWidth(0.6)
        canvas.line(15 * mm, 13 * mm, A4[0] - 15 * mm, 13 * mm)
        canvas.setFillColor(colors.HexColor("#777777"))
        canvas.setFont(PDF_TEXT_FONT, 7.5)
        canvas.drawString(15 * mm, 8.7 * mm, f"Ativo #{asset.id} - {pdf_clean_plain_text(asset.base)}")
        canvas.drawRightString(A4[0] - 15 * mm, 8.7 * mm, f"Pagina {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return buffer


PRODUCT_EXPORT_HEADERS_PT = [
    "ID",
    "Nome do produto",
    "Categoria",
    "Ícone da categoria",
    "Unidade de medida",
    "Kit",
    "Quantidade por kit",
    "Descrição",
    "Estoque disponível",
    "Valor unitário",
    "Limite para bases",
    "Limite para franquias",
    "Dias de bloqueio após limite",
    "Quantidade mínima por pedido",
    "Estoque mínimo",
    "Estoque máximo",
    "Tag de estoque",
    "Ativo",
]

PRODUCT_EXPORT_HEADERS_ZH = [
    "ID",
    "产品名称",
    "类别",
    "类别图标",
    "计量单位",
    "套装",
    "每套数量",
    "描述",
    "可用库存",
    "单价",
    "基地限制",
    "加盟店限制",
    "限制后封锁天数",
    "最低订购量",
    "最低库存",
    "最高库存",
    "启用",
]

EXCEL_ZH_TRANSLATIONS = {
    "Envelope de segurança M": "M 型安全信封",
    "Envelope de segurança P": "P 型安全信封",
    "Envelope médio para envios padrão.": "用于标准寄件的中型信封。",
    "Envelope pequeno para envios leves.": "用于轻量寄件的小型信封。",
    "Etiqueta térmica": "热敏标签",
    "Rolo de etiqueta para impressora térmica.": "热敏打印机用标签卷。",
    "Lacre plástico": "塑料封条",
    "Lacre numerado para controle interno.": "用于内部管控的编号封条。",
    "Embalagens": "包装用品",
    "Etiquetas": "标签",
    "Operacional": "运营用品",
    "un": "个",
    "unidade": "个",
    "unidades": "个",
    "rolo": "卷",
    "rolos": "卷",
    "caixa": "箱",
    "caixas": "箱",
    "pacote": "包",
    "pacotes": "包",
    "metro": "米",
    "metros": "米",
    "kg": "公斤",
    "Sim": "是",
    "Não": "否",
    "Ativo": "启用",
    "Inativo": "停用",
    "Sem limite": "无限制",
}


def translate_excel_value_to_zh(value: Any) -> Any:
    if value is None:
        return ""
    text_value = str(value).strip()
    if not text_value:
        return ""
    return EXCEL_ZH_TRANSLATIONS.get(text_value, text_value)


def product_row_for_excel_language(product: Product, language: str = "pt") -> list[Any]:
    if language == "zh":
        return [
            product.id,
            translate_excel_value_to_zh(product.name),
            translate_excel_value_to_zh(product.category),
            product.category_emoji,
            translate_excel_value_to_zh(product.unit_measure),
            translate_excel_value_to_zh("Sim" if product.is_kit else "Não"),
            product.kit_quantity if product.is_kit else "",
            translate_excel_value_to_zh(product.description),
            product.stock_quantity,
            product.price_brl,
            product.limit_base,
            product.limit_franchise,
            product.limit_block_days,
            product.min_order_quantity,
            product.min_stock,
            product.max_stock,
            translate_excel_value_to_zh("Sim" if product.active else "Não"),
        ]
    return product_row_for_excel(product)

def product_row_for_excel(product: Product) -> list[Any]:
    return [
        product.id,
        product.name,
        product.category,
        product.category_emoji,
        product.unit_measure,
        "Sim" if product.is_kit else "Não",
        product.kit_quantity if product.is_kit else "",
        product.description,
        product.stock_quantity,
        product.price_brl,
        product.limit_base,
        product.limit_franchise,
        product.limit_block_days,
        product.min_order_quantity,
        product.min_stock,
        product.max_stock,
        product.stock_tag_name,
        "Sim" if product.active else "Não",
    ]


def get_header_value(row_values: list[Any], header_map: dict[str, int], names: list[str]) -> Any:
    normalized_names = [normalize_header(name) for name in names]
    for key in normalized_names:
        if key in header_map:
            idx = header_map[key]
            return row_values[idx] if idx < len(row_values) else None
    for key in normalized_names:
        if not key:
            continue
        for header_key, idx in header_map.items():
            if not header_key:
                continue
            if key in header_key or header_key in key:
                return row_values[idx] if idx < len(row_values) else None
    return None


PRODUCT_IMPORT_HEADER_ALIASES = {
    "id": ["ID", "Código", "Codigo", "Cod", "编号", "編號"],
    "name": ["Nome do produto", "Nome", "Produto", "Insumo", "Item", "Descrição do item", "Descricao do item", "产品名称", "產品名稱", "商品名称", "产品", "品名"],
    "category": ["Categoria", "Categoria do produto", "Grupo", "Tipo", "类别", "類別", "分类", "分類"],
    "category_emoji": ["Ícone da categoria", "Icone da categoria", "Emoji", "Ícone", "Icone", "类别图标"],
    "unit_measure": ["Unidade de medida", "Unidade", "Unid.", "Unid", "UM", "U.M.", "Medida", "计量单位", "計量單位", "单位", "單位"],
    "is_kit": ["Kit", "É kit", "E kit", "Produto kit", "套装"],
    "kit_quantity": ["Quantidade por kit", "Qtd por kit", "Itens por kit", "Unidades por kit", "每套数量"],
    "description": ["Descrição", "Descricao", "Descrição do produto", "Descricao do produto", "Observação", "Observacao", "Detalhes", "描述", "说明", "說明", "备注", "備註"],
    "stock_quantity": ["Estoque disponível", "Estoque disponivel", "Estoque", "Quantidade", "Qtd", "Qtde", "Saldo", "可用库存", "可用庫存", "库存", "庫存", "数量", "數量"],
    "price_cents": ["Valor unitário", "Valor unitario", "Valor", "Preço", "Preco", "Preço unitário", "Preco unitario", "Custo", "单价", "單價", "价格", "價格"],
    "limit_base": ["Limite para bases", "Limite base", "Base", "基地限制", "网点限制"],
    "limit_franchise": ["Limite para franquias", "Limite franquia", "Franquia", "加盟店限制", "加盟限制"],
    "limit_block_days": ["Dias de bloqueio após limite", "Bloqueio após limite", "Tempo de bloqueio", "Dias de bloqueio", "限制后封锁天数"],
    "min_order_quantity": ["Quantidade mínima por pedido", "Quantidade minima por pedido", "Pedido mínimo", "Pedido minimo", "Qtd mínima", "Qtd minima", "最低订购量"],
    "min_stock": ["Estoque mínimo", "Estoque minimo", "Mínimo", "Minimo", "Min stock", "Min", "最低库存", "最低庫存"],
    "max_stock": ["Estoque máximo", "Estoque maximo", "Máximo", "Maximo", "Max stock", "Max", "最高库存", "最高庫存"],
    "active": ["Ativo", "Status", "Produto ativo", "启用", "啟用", "状态", "狀態"],
}


PRODUCT_IMPORT_HEADER_ALIASES["stock_tag"] = ["Tag de estoque", "Tag", "Tipo de estoque", "Separacao de estoque", "Separação de estoque"]


def normalize_product_lookup_key(value: Any) -> str:
    text_value = str(value or "").strip().casefold()
    text_value = "".join(ch for ch in unicodedata.normalize("NFD", text_value) if unicodedata.category(ch) != "Mn")
    return " ".join(text_value.split())


def clean_import_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text_value = str(value).strip()
    if text_value.lower() in {"none", "nan", "null"}:
        return default
    return text_value


def get_import_value(row_values: list[Any], header_map: dict[str, int], field_key: str) -> Any:
    return get_header_value(row_values, header_map, PRODUCT_IMPORT_HEADER_ALIASES.get(field_key, []))


def excel_row_is_empty(row_values: list[Any]) -> bool:
    return not any(value is not None and str(value).strip() for value in row_values)


@dataclass
class ProductImportRecord:
    source_row: int
    product_id: int | None
    name: str
    category: str
    category_emoji: str
    unit_measure: str
    is_kit: bool
    kit_quantity: int
    description: str
    stock_quantity: int
    price_cents: int
    limit_base: int | None
    limit_franchise: int | None
    limit_block_days: int
    min_order_quantity: int | None
    min_stock: int | None
    max_stock: int | None
    stock_tag: str
    active: bool


def parse_product_import_record(row_number: int, row_values: list[Any], header_map: dict[str, int]) -> ProductImportRecord | None:
    name = clean_import_text(get_import_value(row_values, header_map, "name"))
    if not name:
        return None
    return ProductImportRecord(
        source_row=row_number,
        product_id=parse_optional_int(get_import_value(row_values, header_map, "id")),
        name=name,
        category=clean_import_text(get_import_value(row_values, header_map, "category")),
        category_emoji=clean_category_emoji(
            get_import_value(row_values, header_map, "category_emoji"),
            clean_import_text(get_import_value(row_values, header_map, "category")),
        ),
        unit_measure=clean_import_text(get_import_value(row_values, header_map, "unit_measure"), "un") or "un",
        is_kit=parse_bool_value(get_import_value(row_values, header_map, "is_kit"), default=False),
        kit_quantity=parse_optional_int(get_import_value(row_values, header_map, "kit_quantity")) or 1,
        description=clean_import_text(get_import_value(row_values, header_map, "description")),
        stock_quantity=parse_optional_int(get_import_value(row_values, header_map, "stock_quantity")) or 0,
        price_cents=parse_money_to_cents(get_import_value(row_values, header_map, "price_cents")),
        limit_base=parse_optional_int(get_import_value(row_values, header_map, "limit_base")),
        limit_franchise=parse_optional_int(get_import_value(row_values, header_map, "limit_franchise")),
        limit_block_days=parse_required_positive_int(get_import_value(row_values, header_map, "limit_block_days")) or 60,
        min_order_quantity=parse_optional_int(get_import_value(row_values, header_map, "min_order_quantity")),
        min_stock=parse_optional_int(get_import_value(row_values, header_map, "min_stock")),
        max_stock=parse_optional_int(get_import_value(row_values, header_map, "max_stock")),
        stock_tag=normalize_stock_tag_slug(get_import_value(row_values, header_map, "stock_tag")),
        active=parse_bool_value(get_import_value(row_values, header_map, "active"), default=True),
    )



def product_import_signature(record: ProductImportRecord) -> tuple[Any, ...]:
    """Assinatura exata da linha. Não junta produtos só pelo nome para não sumir item da planilha."""
    return (
        record.product_id,
        normalize_product_lookup_key(record.name),
        normalize_product_lookup_key(record.category),
        record.category_emoji,
        normalize_product_lookup_key(record.unit_measure),
        bool(record.is_kit),
        int(record.kit_quantity or 1),
        normalize_product_lookup_key(record.description),
        int(record.stock_quantity or 0),
        int(record.price_cents or 0),
        record.limit_base,
        record.limit_franchise,
        record.limit_block_days,
        record.min_order_quantity,
        record.min_stock,
        record.max_stock,
        record.stock_tag,
        bool(record.active),
    )


def dedupe_import_records(records: list[ProductImportRecord]) -> tuple[list[ProductImportRecord], int]:
    """Mantém apenas a última linha de cada nome para impedir produtos duplicados."""
    positions: dict[tuple[str, str], int] = {}
    result: list[ProductImportRecord] = []
    duplicates = 0
    for record in records:
        key = (normalize_stock_tag_slug(record.stock_tag), normalize_product_lookup_key(record.name))
        if key in positions:
            duplicates += 1
            result[positions[key]] = record
            continue
        positions[key] = len(result)
        result.append(record)
    return result, duplicates


def flash_import_errors(row_errors: list[str]) -> None:
    if not row_errors:
        return
    preview = "; ".join(row_errors[:5])
    suffix = "" if len(row_errors) <= 5 else f"; +{len(row_errors) - 5} outro(s) erro(s) nos logs."
    flash(f"Algumas linhas não foram importadas: {preview}{suffix}", "warning")



def worksheet_values(row: Any) -> list[Any]:
    return list(row or [])


def header_map_contains_alias(header_map: dict[str, int], aliases: list[str]) -> bool:
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        if alias in header_map:
            return True
    for alias in normalized_aliases:
        if not alias:
            continue
        for header_key in header_map.keys():
            if not header_key:
                continue
            if alias in header_key or header_key in alias:
                return True
    return False


def header_map_score(header_map: dict[str, int]) -> int:
    required = 0
    if header_map_contains_alias(header_map, PRODUCT_IMPORT_HEADER_ALIASES["name"]):
        required += 5
    for key in ["category", "unit_measure", "description", "stock_quantity", "price_cents", "active"]:
        if header_map_contains_alias(header_map, PRODUCT_IMPORT_HEADER_ALIASES[key]):
            required += 1
    return required


def detect_product_header_row(worksheet: Any, max_scan_rows: int = 40) -> tuple[int, dict[str, int]]:
    """Encontra a linha de cabeçalho mesmo quando a planilha tem título antes da tabela."""
    best_row_number = 1
    best_map: dict[str, int] = {}
    best_score = -1
    max_row = min(int(getattr(worksheet, "max_row", 1) or 1), max_scan_rows)
    for row_number, row_values_tuple in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        row_values = worksheet_values(row_values_tuple)
        header_map = {normalize_header(value): index for index, value in enumerate(row_values) if value is not None and str(value).strip()}
        score = header_map_score(header_map)
        if score > best_score:
            best_row_number = row_number
            best_map = header_map
            best_score = score
        # Nome do produto achado e pelo menos mais um campo: já é cabeçalho confiável.
        if score >= 6:
            return row_number, header_map
    return best_row_number, best_map


def chunk_list(items: list[Any], size: int) -> list[list[Any]]:
    return [items[index:index + size] for index in range(0, len(items), size)]


def product_record_db_values(record: ProductImportRecord) -> tuple[Any, ...]:
    return (
        record.name,
        record.category,
        record.category_emoji,
        record.unit_measure,
        1 if record.is_kit else 0,
        int(record.kit_quantity or 1) if record.is_kit else 1,
        record.description,
        int(record.stock_quantity or 0),
        int(record.price_cents or 0),
        record.limit_base,
        record.limit_franchise,
        record.limit_block_days,
        record.min_order_quantity,
        record.min_stock,
        record.max_stock,
        record.stock_tag,
        1 if record.active else 0,
    )


def sql_literal(value: Any) -> str:
    """Literal SQL seguro para reduzir centenas de chamadas HTTP ao D1 sem limite de parâmetros."""
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:
            return "NULL"
        return str(value)
    text_value = str(value)
    # Remove caracteres de controle que podem quebrar SQL/planilha.
    text_value = "".join(ch for ch in text_value if ch == "\n" or ch == "\t" or ord(ch) >= 32)
    return "'" + text_value.replace("'", "''") + "'"


def product_upsert_sql_rows(rows: list[tuple[int | None, ProductImportRecord]], created_at: str, updated_at: str) -> str:
    values_sql: list[str] = []
    for target_id, record in rows:
        values = [
            target_id,
            record.name,
            record.category,
            record.category_emoji,
            record.unit_measure,
            1 if record.is_kit else 0,
            int(record.kit_quantity or 1) if record.is_kit else 1,
            record.description,
            int(record.stock_quantity or 0),
            int(record.price_cents or 0),
            record.limit_base,
            record.limit_franchise,
            record.limit_block_days,
            record.min_order_quantity,
            record.min_stock,
            record.max_stock,
            record.stock_tag,
            1 if record.active else 0,
            0,
            created_at,
            updated_at,
        ]
        values_sql.append("(" + ", ".join(sql_literal(value) for value in values) + ")")
    return ",\n".join(values_sql)


def execute_upsert_products_chunked(conn: Any, rows: list[tuple[int | None, ProductImportRecord]], row_errors: list[str]) -> tuple[int, int, int]:
    """Cria/atualiza produtos em poucos comandos, sem explodir a página em timeout/500.

    A versão anterior caía para centenas de requisições linha a linha no Cloudflare D1.
    Isso fazia o Render estourar tempo, abrir Internal Server Error e deixar a importação pela metade.
    """
    created = sum(1 for target_id, _record in rows if target_id is None)
    updated = len(rows) - created
    skipped = 0
    fields = "id, name, category, category_emoji, unit_measure, is_kit, kit_quantity, description, stock_quantity, price_cents, limit_base, limit_franchise, limit_block_days, min_order_quantity, min_stock, max_stock, stock_tag, active, catalog_archived, created_at, updated_at"
    update_set = """
        name = excluded.name,
        category = excluded.category,
        category_emoji = excluded.category_emoji,
        unit_measure = excluded.unit_measure,
        is_kit = excluded.is_kit,
        kit_quantity = excluded.kit_quantity,
        description = excluded.description,
        stock_quantity = excluded.stock_quantity,
        price_cents = excluded.price_cents,
        limit_base = excluded.limit_base,
        limit_franchise = excluded.limit_franchise,
        limit_block_days = excluded.limit_block_days,
        min_order_quantity = excluded.min_order_quantity,
        min_stock = excluded.min_stock,
        max_stock = excluded.max_stock,
        stock_tag = excluded.stock_tag,
        active = excluded.active,
        catalog_archived = 0,
        updated_at = excluded.updated_at
    """
    # 60 linhas por SQL mantém o payload pequeno e evita timeout; ainda reduz muito as chamadas ao D1.
    for chunk in chunk_list(rows, 60):
        if not chunk:
            continue
        created_at = now_iso()
        updated_at = now_iso()
        sql = f"""
        INSERT INTO products ({fields})
        VALUES {product_upsert_sql_rows(chunk, created_at, updated_at)}
        ON CONFLICT(id) DO UPDATE SET {update_set}
        """
        try:
            conn.execute(sql)
        except Exception as chunk_exc:
            print(f"[IMPORTAÇÃO PRODUTOS] Upsert em bloco falhou; tentando bloco menor: {chunk_exc}")
            # Divide novamente para impedir que uma célula problemática derrube 60 produtos.
            if len(chunk) > 1:
                for small in chunk_list(chunk, 10):
                    try:
                        small_sql = f"""
                        INSERT INTO products ({fields})
                        VALUES {product_upsert_sql_rows(small, now_iso(), now_iso())}
                        ON CONFLICT(id) DO UPDATE SET {update_set}
                        """
                        conn.execute(small_sql)
                    except Exception as small_exc:
                        print(f"[IMPORTAÇÃO PRODUTOS] Upsert em sub-bloco falhou; tentando linha a linha: {small_exc}")
                        for target_id, record in small:
                            try:
                                one_sql = f"""
                                INSERT INTO products ({fields})
                                VALUES {product_upsert_sql_rows([(target_id, record)], now_iso(), now_iso())}
                                ON CONFLICT(id) DO UPDATE SET {update_set}
                                """
                                conn.execute(one_sql)
                            except Exception as exc:
                                skipped += 1
                                if target_id is None:
                                    created -= 1
                                else:
                                    updated -= 1
                                msg = f"linha {record.source_row}: {type(exc).__name__} - {str(exc)[:180]}"
                                row_errors.append(msg)
                                print(f"[IMPORTAÇÃO PRODUTOS] Erro ao importar linha {record.source_row} ({record.name}): {exc}")
            else:
                target_id, record = chunk[0]
                skipped += 1
                if target_id is None:
                    created -= 1
                else:
                    updated -= 1
                msg = f"linha {record.source_row}: {type(chunk_exc).__name__} - {str(chunk_exc)[:180]}"
                row_errors.append(msg)
                print(f"[IMPORTAÇÃO PRODUTOS] Erro ao importar linha {record.source_row} ({record.name}): {chunk_exc}")
    return max(0, created), max(0, updated), skipped


def ensure_product_import_columns(conn: Any) -> None:
    """Garante colunas novas antes da importação, inclusive em banco D1 já criado."""
    try:
        supply_request_columns = {row["name"] for row in conn.execute("PRAGMA table_info(supply_requests)").fetchall()}
        if "people_count" not in supply_request_columns:
            conn.execute("ALTER TABLE supply_requests ADD COLUMN people_count INTEGER")

        product_columns = {row["name"] for row in conn.execute("PRAGMA table_info(products)").fetchall()}
    except Exception as exc:
        print(f"[IMPORTAÇÃO PRODUTOS] Não foi possível verificar colunas por PRAGMA: {exc}")
        return
    migrations = [
        ("category_emoji", "ALTER TABLE products ADD COLUMN category_emoji TEXT"),
        ("image_name", "ALTER TABLE products ADD COLUMN image_name TEXT"),
        ("image_key", "ALTER TABLE products ADD COLUMN image_key TEXT"),
        ("image_content_type", "ALTER TABLE products ADD COLUMN image_content_type TEXT"),
        ("unit_measure", "ALTER TABLE products ADD COLUMN unit_measure TEXT NOT NULL DEFAULT 'un'"),
        ("min_order_quantity", "ALTER TABLE products ADD COLUMN min_order_quantity INTEGER"),
        ("min_stock", "ALTER TABLE products ADD COLUMN min_stock INTEGER"),
        ("max_stock", "ALTER TABLE products ADD COLUMN max_stock INTEGER"),
        ("catalog_archived", "ALTER TABLE products ADD COLUMN catalog_archived INTEGER NOT NULL DEFAULT 0"),
        ("visible_base", "ALTER TABLE products ADD COLUMN visible_base INTEGER NOT NULL DEFAULT 1"),
        ("visible_franchise", "ALTER TABLE products ADD COLUMN visible_franchise INTEGER NOT NULL DEFAULT 1"),
        ("internal", "ALTER TABLE products ADD COLUMN internal INTEGER NOT NULL DEFAULT 0"),
        ("is_kit", "ALTER TABLE products ADD COLUMN is_kit INTEGER NOT NULL DEFAULT 0"),
        ("kit_quantity", "ALTER TABLE products ADD COLUMN kit_quantity INTEGER NOT NULL DEFAULT 1"),
        ("stock_tag", "ALTER TABLE products ADD COLUMN stock_tag TEXT NOT NULL DEFAULT 'insumos'"),
        ("limit_block_days", "ALTER TABLE products ADD COLUMN limit_block_days INTEGER NOT NULL DEFAULT 60"),
    ]
    for column_name, sql in migrations:
        if column_name not in product_columns:
            try:
                conn.execute(sql)
            except Exception as exc:
                print(f"[IMPORTAÇÃO PRODUTOS] Migração ignorada para {column_name}: {exc}")


def product_import_key_from_values(name: Any, stock_tag: Any = DEFAULT_STOCK_TAG) -> tuple[str, str]:
    return (normalize_stock_tag_slug(stock_tag), normalize_product_lookup_key(name))


def product_import_record_key(record: ProductImportRecord) -> tuple[str, str]:
    return product_import_key_from_values(record.name, record.stock_tag)


def prepare_import_stock_tags(conn: Any, records: list[ProductImportRecord]) -> None:
    known = {
        normalize_stock_tag_slug(row["slug"])
        for row in conn.execute("SELECT slug FROM stock_tags").fetchall()
    }
    can_manage = can_manage_stock_tags()
    for record in records:
        record.stock_tag = normalize_stock_tag_slug(record.stock_tag)
        if record.stock_tag in known:
            continue
        if not can_manage:
            record.stock_tag = DEFAULT_STOCK_TAG
            continue
        label = default_stock_tag_name(record.stock_tag)
        conn.execute(
            """
            INSERT OR IGNORE INTO stock_tags (slug, name, description, active, system_key, created_at)
            VALUES (?, ?, ?, 1, 0, ?)
            """,
            (record.stock_tag, label, "Criada pela importacao de produtos.", now_iso()),
        )
        known.add(record.stock_tag)


def archive_products_outside_import(
    conn: Any,
    imported_keys: set[tuple[str, str]],
    preferred_ids: dict[tuple[str, str], int],
    visible_before: set[int],
) -> int:
    rows = conn.execute("SELECT id, name, stock_tag FROM products").fetchall()
    keep_ids: set[int] = set()
    grouped: dict[tuple[str, str], list[int]] = {}
    for row in rows:
        key = product_import_key_from_values(row["name"], row["stock_tag"] if "stock_tag" in row.keys() else DEFAULT_STOCK_TAG)
        grouped.setdefault(key, []).append(int(row["id"]))

    for key in imported_keys:
        ids = sorted(grouped.get(key, []))
        if not ids:
            continue
        preferred = preferred_ids.get(key)
        keep_ids.add(preferred if preferred in ids else ids[0])

    archive_ids = [
        int(row["id"])
        for row in rows
        if int(row["id"]) not in keep_ids
    ]
    for chunk in chunked_ids(archive_ids):
        placeholders = ",".join("?" for _ in chunk)
        conn.execute(
            f"UPDATE products SET catalog_archived = 1, active = 0, updated_at = ? WHERE id IN ({placeholders})",
            [now_iso(), *chunk],
        )
    if keep_ids:
        for chunk in chunked_ids(list(keep_ids)):
            placeholders = ",".join("?" for _ in chunk)
            conn.execute(
                f"UPDATE products SET catalog_archived = 0 WHERE id IN ({placeholders})",
                chunk,
            )
    return len(visible_before.intersection(archive_ids))


def import_products_from_workbook_bytes(
    uploaded_bytes: bytes,
    import_mode: str = "merge",
) -> tuple[int, int, int, int, list[str]]:
    import_mode = "replace" if import_mode == "replace" else "merge"
    workbook = load_workbook(BytesIO(uploaded_bytes), data_only=True, read_only=True)
    worksheet = workbook.active
    if not worksheet or int(getattr(worksheet, "max_row", 0) or 0) < 1:
        return 0, 0, 0, 0, ["planilha vazia"]

    header_row_number, header_map = detect_product_header_row(worksheet)
    if not header_map_contains_alias(header_map, PRODUCT_IMPORT_HEADER_ALIASES["name"]):
        try:
            workbook.close()
        except Exception:
            pass
        return 0, 0, 0, 0, ["não encontrei a coluna Nome do produto / 产品名称 na planilha"]

    parsed_records: list[ProductImportRecord] = []
    skipped = 0
    row_errors: list[str] = []

    for row_number, row_values_tuple in enumerate(worksheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
        row_values = worksheet_values(row_values_tuple)
        if excel_row_is_empty(row_values):
            continue
        try:
            record = parse_product_import_record(row_number, row_values, header_map)
            if record is None:
                skipped += 1
                continue
            parsed_records.append(record)
        except Exception as exc:
            skipped += 1
            row_errors.append(f"linha {row_number}: {type(exc).__name__} - {str(exc)[:180]}")
            print(f"[IMPORTAÇÃO PRODUTOS] Erro ao interpretar linha {row_number}: {exc}")

    parsed_records, duplicates_merged = dedupe_import_records(parsed_records)
    skipped += duplicates_merged
    if import_mode == "replace" and row_errors:
        try:
            workbook.close()
        except Exception:
            pass
        row_errors.append("a substituição foi cancelada para evitar um catálogo incompleto")
        return 0, 0, skipped, 0, row_errors
    if not parsed_records:
        try:
            workbook.close()
        except Exception:
            pass
        return 0, 0, skipped, 0, row_errors

    with db_connect() as conn:
        ensure_product_import_columns(conn)
        prepare_import_stock_tags(conn, parsed_records)
        existing_rows = conn.execute(
            "SELECT id, name, stock_tag, catalog_archived FROM products ORDER BY catalog_archived ASC, id ASC"
        ).fetchall()
        visible_before = {
            int(row["id"])
            for row in existing_rows
            if not bool(row["catalog_archived"])
        }
        existing_by_key: dict[tuple[str, str], Any] = {}
        for row in existing_rows:
            key = product_import_key_from_values(row["name"], row["stock_tag"] if "stock_tag" in row.keys() else DEFAULT_STOCK_TAG)
            if key[1] and key not in existing_by_key:
                existing_by_key[key] = row

        upsert_rows: list[tuple[int | None, ProductImportRecord]] = []
        preferred_ids: dict[tuple[str, str], int] = {}
        for record in parsed_records:
            key = product_import_record_key(record)
            existing_row = existing_by_key.get(key)
            target_id = int(existing_row["id"]) if existing_row is not None else None
            if target_id is not None:
                preferred_ids[key] = target_id
            upsert_rows.append((target_id, record))

        created, updated, skipped_upsert = execute_upsert_products_chunked(conn, upsert_rows, row_errors)
        skipped += skipped_upsert
        archived = 0
        if import_mode == "replace" and not row_errors:
            imported_keys = {product_import_record_key(record) for record in parsed_records}
            archived = archive_products_outside_import(conn, imported_keys, preferred_ids, visible_before)
        category_emojis = {
            normalize_product_lookup_key(record.category): (record.category, record.category_emoji)
            for record in parsed_records
            if record.category
        }
        for category_name, category_emoji in category_emojis.values():
            conn.execute(
                "UPDATE products SET category_emoji = ? WHERE LOWER(TRIM(category)) = LOWER(TRIM(?))",
                (category_emoji, category_name),
            )
        conn.commit()
    try:
        workbook.close()
    except Exception:
        pass
    return created, updated, skipped, archived, row_errors


USER_IMPORT_HEADER_ALIASES = {
    "responsible_name": ["Nome do responsável", "Nome do responsavel", "Responsável", "Responsavel"],
    "username": ["Nome de usuário", "Nome de usuario", "Usuário", "Usuario", "Login"],
    "password": ["Senha", "Senha inicial", "Nova senha", "Alterar senha", "Password"],
    "role": ["Tipo de acesso", "Perfil", "Tipo", "Acesso"],
    "status": ["Status do cadastro", "Status", "Situação", "Situacao"],
    "base_name": ["Nome da base", "Base", "Unidade", "Nome da base/franquia", "Setor", "Unidade / Franquia", "Unidade/Franquia", "Organização", "Organizacao", "Base ou franquia"],
    "franchise_name": ["Nome da franquia", "Franquia", "Setor", "Unidade / Franquia", "Unidade/Franquia", "Nome da base/franquia", "Base ou franquia"],
    "franchise_number": ["Telefone", "Número de telefone", "Numero de telefone", "Telefone da franquia", "Número da franquia", "Numero da franquia", "Código da franquia", "Codigo da franquia"],
    "cnpj": ["CNPJ", "CNPJ da franquia"],
}


@dataclass
class UserImportRecord:
    source_row: int
    responsible_name: str
    username: str
    password_hash: str
    role: str
    status: str
    organization_name: str
    franchise_name: str
    franchise_number: str
    cnpj: str


def get_user_import_value(row_values: list[Any], header_map: dict[str, int], field_key: str) -> Any:
    return get_header_value(row_values, header_map, USER_IMPORT_HEADER_ALIASES.get(field_key, []))


def user_header_map_score(header_map: dict[str, int]) -> int:
    score = 0
    for key in ["responsible_name", "username", "password", "role", "status"]:
        if header_map_contains_alias(header_map, USER_IMPORT_HEADER_ALIASES[key]):
            score += 3
    for key in ["base_name", "franchise_name", "franchise_number", "cnpj"]:
        if header_map_contains_alias(header_map, USER_IMPORT_HEADER_ALIASES[key]):
            score += 1
    return score


def detect_user_header_row(worksheet: Any, max_scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    best_row_number = 1
    best_map: dict[str, int] = {}
    best_score = -1
    max_row = min(int(getattr(worksheet, "max_row", 1) or 1), max_scan_rows)
    for row_number, row_values_tuple in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True),
        start=1,
    ):
        row_values = worksheet_values(row_values_tuple)
        header_map = {
            normalize_header(value): index
            for index, value in enumerate(row_values)
            if value is not None and str(value).strip()
        }
        score = user_header_map_score(header_map)
        if score > best_score:
            best_row_number = row_number
            best_map = header_map
            best_score = score
        if score >= 15:
            return row_number, header_map
    return best_row_number, best_map


@lru_cache(maxsize=4096)
def generate_user_import_password_hash(password: str) -> str:
    """Gera hash compatível com check_password_hash sem travar importação em massa.

    Em importações grandes, o método padrão do Werkzeug pode estourar tempo/memória
    do Render/Cloudflare. O cache acelera planilhas com senha padrão repetida e o
    PBKDF2 reduz o risco de Internal Server Error por timeout.
    """
    return generate_password_hash(str(password or ""), method="pbkdf2:sha256:20000", salt_length=12)


def parse_user_import_record(row_number: int, row_values: list[Any], header_map: dict[str, int]) -> UserImportRecord:
    responsible_name = clean_import_text(get_user_import_value(row_values, header_map, "responsible_name"))
    username = normalize_username(clean_import_text(get_user_import_value(row_values, header_map, "username")))
    password = clean_import_text(get_user_import_value(row_values, header_map, "password"))
    role = normalize_user_role(get_user_import_value(row_values, header_map, "role"), allow_admin=True)
    status = normalize_user_status(get_user_import_value(row_values, header_map, "status"), default="approved")

    if not responsible_name:
        raise ValueError("nome do responsável não informado")
    if not valid_username(username):
        raise ValueError("nome de usuário inválido")
    if role is None:
        raise ValueError("tipo de acesso inválido")
    if status is None:
        raise ValueError("status do cadastro inválido")

    base_value = get_user_import_value(row_values, header_map, "base_name")
    franchise_value = get_user_import_value(row_values, header_map, "franchise_name")
    # Planilhas antigas podem ter uma única coluna "Base/Franquia".
    # Para franquia, usa essa coluna como fallback do nome da franquia.
    if role == "franchise" and not clean_import_text(franchise_value) and clean_import_text(base_value):
        franchise_value = base_value

    organization_name, franchise_name, franchise_number, cnpj = validate_user_profile_fields(
        role,
        organization_name=base_value,
        franchise_name=franchise_value,
        franchise_number=get_user_import_value(row_values, header_map, "franchise_number"),
        cnpj=get_user_import_value(row_values, header_map, "cnpj"),
        strict_base=False,
    )
    return UserImportRecord(
        source_row=row_number,
        responsible_name=responsible_name[:160],
        username=username,
        password_hash=generate_user_import_password_hash(password) if password else "",
        role=role,
        status=status,
        organization_name=organization_name,
        franchise_name=franchise_name,
        franchise_number=franchise_number,
        cnpj=cnpj,
    )


def user_upsert_sql_rows(rows: list[tuple[int | None, UserImportRecord]], created_at: str, updated_at: str) -> str:
    values_sql: list[str] = []
    for _target_id, record in rows:
        values = [
            record.responsible_name,
            record.organization_name,
            record.franchise_name,
            record.franchise_number,
            record.cnpj,
            record.username,
            synthetic_email_for_username(record.username),
            record.password_hash,
            record.role,
            record.status,
            created_at,
            updated_at,
            0,
            0,
        ]
        values_sql.append("(" + ", ".join(sql_literal(value) for value in values) + ")")
    return ",\n".join(values_sql)


def execute_user_upsert_chunked(conn: Any, rows: list[tuple[int | None, UserImportRecord]], row_errors: list[str]) -> tuple[int, int, int]:
    """Cria/atualiza usuários em blocos pequenos e seguros.

    Usa o login (username) como chave de atualização. Isso evita INSERT com id NULL
    e reduz falhas no Cloudflare D1. Quando um bloco falha, diminui automaticamente
    até linha a linha sem derrubar a rota.
    """
    if not rows:
        return 0, 0, 0
    created = sum(1 for target_id, _record in rows if target_id is None)
    updated = len(rows) - created
    skipped = 0
    fields = (
        "responsible_name, organization_name, franchise_name, franchise_number, cnpj, "
        "username, email, password_hash, role, status, created_at, updated_at, page_permissions_configured, action_permissions_configured"
    )
    update_set = """
        responsible_name = excluded.responsible_name,
        organization_name = excluded.organization_name,
        franchise_name = excluded.franchise_name,
        franchise_number = excluded.franchise_number,
        cnpj = excluded.cnpj,
        email = excluded.email,
        password_hash = excluded.password_hash,
        role = excluded.role,
        status = excluded.status,
        updated_at = excluded.updated_at,
        page_permissions_configured = 0,
        action_permissions_configured = 0
    """

    def execute_chunk(chunk: list[tuple[int | None, UserImportRecord]]) -> bool:
        if not chunk:
            return True
        sql = f"""
        INSERT INTO users ({fields})
        VALUES {user_upsert_sql_rows(chunk, now_iso(), now_iso())}
        ON CONFLICT(username) DO UPDATE SET {update_set}
        """
        conn.execute(sql)
        return True

    for chunk in chunk_list(rows, 20):
        try:
            execute_chunk(chunk)
            continue
        except Exception as chunk_exc:
            print(f"[IMPORTAÇÃO USUÁRIOS] Upsert em bloco de {len(chunk)} falhou: {chunk_exc}")

        for small in chunk_list(chunk, 5):
            try:
                execute_chunk(small)
                continue
            except Exception as small_exc:
                print(f"[IMPORTAÇÃO USUÁRIOS] Upsert em sub-bloco de {len(small)} falhou: {small_exc}")

            for target_id, record in small:
                try:
                    execute_chunk([(target_id, record)])
                except Exception as exc:
                    skipped += 1
                    if target_id is None:
                        created -= 1
                    else:
                        updated -= 1
                    row_errors.append(f"linha {record.source_row}: {type(exc).__name__} - {str(exc)[:180]}")
                    print(f"[IMPORTAÇÃO USUÁRIOS] Erro na linha {record.source_row}: {exc}")
    return max(0, created), max(0, updated), skipped


def reject_users_outside_import(conn: Any, existing_rows: list[Any], imported_usernames: set[str], current_user_id: int | None) -> int:
    """Recusa usuários que não estão na planilha sem usar NOT IN gigante.

    O NOT IN com muitos parâmetros pode quebrar SQLite/D1 ou estourar payload.
    Aqui calculamos os IDs localmente e fazemos UPDATE por blocos pequenos.
    """
    ids_to_reject: list[int] = []
    for row in existing_rows:
        try:
            user_id = int(row["id"])
            username_key = normalize_username(row["username"] or "")
            status = (row["status"] or "").strip().lower() if "status" in row.keys() else ""
        except Exception:
            continue
        if current_user_id is not None and user_id == int(current_user_id):
            continue
        if username_key in imported_usernames:
            continue
        if status == "rejected":
            continue
        ids_to_reject.append(user_id)

    if not ids_to_reject:
        return 0
    updated_at = sql_literal(now_iso())
    total = 0
    for chunk in chunk_list(ids_to_reject, 400):
        ids_sql = ",".join(str(int(value)) for value in chunk)
        conn.execute(f"UPDATE users SET status = 'rejected', updated_at = {updated_at} WHERE id IN ({ids_sql})")
        total += len(chunk)
    return total


def import_users_from_workbook_bytes(uploaded_bytes: bytes, import_mode: str = "merge", current_user_id: int | None = None, current_user_role: str = "admin", update_passwords: bool = False) -> tuple[int, int, int, int, list[str]]:
    import_mode = (import_mode or "merge").strip().lower()
    if import_mode not in {"merge", "replace"}:
        import_mode = "merge"
    update_passwords = bool(update_passwords and current_user_role == "dev")

    try:
        workbook = load_workbook(BytesIO(uploaded_bytes), data_only=True, read_only=True)
    except Exception as exc:
        return 0, 0, 0, 0, [f"não foi possível abrir a planilha: {type(exc).__name__}"]

    try:
        worksheet = workbook.active
        if not worksheet or int(getattr(worksheet, "max_row", 0) or 0) < 1:
            return 0, 0, 0, 0, ["planilha vazia"]

        header_row_number, header_map = detect_user_header_row(worksheet)
        required_keys = ["responsible_name", "username", "role", "status"]
        missing_headers = [
            USER_IMPORT_HEADER_ALIASES[key][0]
            for key in required_keys
            if not header_map_contains_alias(header_map, USER_IMPORT_HEADER_ALIASES[key])
        ]
        if missing_headers:
            return 0, 0, 0, 0, ["colunas obrigatórias ausentes: " + ", ".join(missing_headers)]

        records: list[UserImportRecord] = []
        skipped = 0
        errors: list[str] = []
        seen_usernames: set[str] = set()
        max_import_rows = int(os.getenv("USER_IMPORT_MAX_ROWS", "1000"))
        processed_rows = 0

        for row_number, row_values_tuple in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            row_values = worksheet_values(row_values_tuple)
            if excel_row_is_empty(row_values):
                continue
            processed_rows += 1
            if processed_rows > max_import_rows:
                skipped += 1
                errors.append(f"limite de {max_import_rows} linhas atingido; divida a planilha para importar o restante")
                break
            try:
                record = parse_user_import_record(row_number, row_values, header_map)
                if record.role == "dev" and current_user_role != "dev":
                    raise ValueError("somente usuário Dev pode importar usuários com tipo Dev")
                if record.role not in STATIC_ROLE_KEYS and current_user_role != "dev":
                    raise ValueError("somente usuário Dev pode importar usuários com tipo personalizado")
                if record.username in seen_usernames:
                    raise ValueError("nome de usuário duplicado na planilha")
                seen_usernames.add(record.username)
                records.append(record)
            except Exception as exc:
                skipped += 1
                errors.append(f"linha {row_number}: {str(exc)[:160]}")

        if not records:
            return 0, 0, skipped, 0, errors

        with db_connect() as conn:
            existing_rows = conn.execute("SELECT id, username, status, role, password_hash FROM users").fetchall()
            existing_by_username = {
                normalize_username(row["username"]): {"id": int(row["id"]), "role": str(row["role"] or "base"), "password_hash": str(row["password_hash"] or "")}
                for row in existing_rows
                if row["username"]
            }
            upsert_rows: list[tuple[int | None, UserImportRecord]] = []
            imported_usernames = {record.username for record in records}

            for record in records:
                existing_info = existing_by_username.get(record.username)
                target_id = existing_info["id"] if existing_info else None
                existing_role = existing_info["role"] if existing_info else ""
                if target_id is None and not record.password_hash:
                    skipped += 1
                    errors.append(f"linha {record.source_row}: senha nao informada para novo usuario")
                    continue
                if target_id is not None and (not update_passwords or not record.password_hash):
                    record.password_hash = existing_info["password_hash"]
                if current_user_role != "dev" and role_is_admin_like(existing_role) and record.role != existing_role:
                    skipped += 1
                    errors.append(f"linha {record.source_row}: somente usuário Dev pode alterar o tipo de acesso de administradores")
                    continue
                upsert_rows.append((target_id, record))

            created, updated, upsert_skipped = execute_user_upsert_chunked(conn, upsert_rows, errors)
            skipped += upsert_skipped

            replaced = 0
            if import_mode == "replace":
                replaced = reject_users_outside_import(conn, existing_rows, imported_usernames, current_user_id)

            conn.commit()
        return created, updated, skipped, replaced, errors
    finally:
        try:
            workbook.close()
        except Exception:
            pass



# ---------- Entrada de Materiais ----------

MATERIAL_IMPORT_HEADER_ALIASES = {
    "item_name": ["Nome do item", "Item", "Produto", "Insumo", "Material", "Nome do produto", "产品名称", "物料名称"],
    "quantity": ["Quantidade", "Qtd", "Qtde", "数量", "數量"],
    "unit_price_cents": ["Valor unitário", "Valor unitario", "Preço unitário", "Preco unitario", "Unitário", "Unitario", "单价", "單價"],
    "unit_measure": ["Unidade de medida", "Unidade", "Un", "UM", "计量单位", "單位", "单位"],
    "invoice_number": ["Número da nota", "Numero da nota", "NF", "Nota fiscal", "Nº nota", "No nota", "发票号码"],
    "invoice_date": ["Data da nota", "Data NF", "Emissão", "Emissao", "发票日期"],
    "invoice_value_cents": ["Valor da nota", "Valor NF", "Total da nota", "发票金额"],
    "notes": ["Observações", "Observacoes", "Obs", "Notas", "备注"],
}


def parse_optional_date(value: Any) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    text_value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text_value, fmt)
        except ValueError:
            pass
    return None


def material_entry_header_map_score(header_map: dict[str, int]) -> int:
    score = 0
    for key in ["item_name", "quantity", "unit_price_cents", "unit_measure", "invoice_number", "invoice_date", "invoice_value_cents", "notes"]:
        if header_map_contains_alias(header_map, MATERIAL_IMPORT_HEADER_ALIASES[key]):
            score += 3 if key in {"item_name", "quantity"} else 1
    return score


def detect_material_entry_header_row(worksheet: Any, max_scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    best_row_number = 1
    best_map: dict[str, int] = {}
    best_score = -1
    max_row = min(int(getattr(worksheet, "max_row", 1) or 1), max_scan_rows)
    for row_number, row_values_tuple in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        row_values = worksheet_values(row_values_tuple)
        header_map = {normalize_header(value): index for index, value in enumerate(row_values) if value is not None and str(value).strip()}
        score = material_entry_header_map_score(header_map)
        if score > best_score:
            best_row_number = row_number
            best_map = header_map
            best_score = score
        if score >= 7:
            return row_number, header_map
    return best_row_number, best_map


def get_material_import_value(row_values: list[Any], header_map: dict[str, int], field_key: str) -> Any:
    return get_header_value(row_values, header_map, MATERIAL_IMPORT_HEADER_ALIASES.get(field_key, []))


def row_to_material_entry(row: Any | None, product: Product | None = None) -> MaterialEntry | None:
    if row is None:
        return None
    return MaterialEntry(
        id=int(row["id"]),
        product_id=row["product_id"] if "product_id" in row.keys() and row["product_id"] is not None else None,
        item_name=row["item_name"] or "",
        quantity=int(row["quantity"] or 0),
        unit_measure=row["unit_measure"] or "un",
        unit_price_cents=int(row["unit_price_cents"] or 0),
        invoice_file_name=row["invoice_file_name"] or "",
        invoice_file_key=row["invoice_file_key"] or "",
        invoice_number=row["invoice_number"] or "",
        invoice_date=parse_dt(row["invoice_date"]) if "invoice_date" in row.keys() and row["invoice_date"] else None,
        invoice_value_cents=int(row["invoice_value_cents"] or 0),
        notes=row["notes"] or "",
        created_by_id=row["created_by_id"] if "created_by_id" in row.keys() and row["created_by_id"] is not None else None,
        created_at=parse_dt(row["created_at"]),
        product=product,
        created_by_name=row["created_by_name"] if "created_by_name" in row.keys() and row["created_by_name"] else "",
    )


def find_product_by_name(conn: Any, name: str, stock_tag: str = DEFAULT_STOCK_TAG) -> Product | None:
    cleaned_name = (name or "").strip()
    if not cleaned_name:
        return None
    normalized_stock_tag = normalize_stock_tag_slug(stock_tag)
    # Primeiro tenta busca exata para evitar varrer a tabela inteira no D1 a cada entrada.
    row = conn.execute(
        "SELECT * FROM products WHERE catalog_archived = 0 AND stock_tag = ? AND name = ? LIMIT 1",
        (normalized_stock_tag, cleaned_name),
    ).fetchone()
    product = row_to_product(row) if row is not None else None
    if product is not None:
        return product
    # Fallback sem varrer a tabela inteira no Cloudflare D1.
    row = conn.execute(
        "SELECT * FROM products WHERE catalog_archived = 0 AND stock_tag = ? AND lower(name) = lower(?) LIMIT 1",
        (normalized_stock_tag, cleaned_name),
    ).fetchone()
    product = row_to_product(row) if row is not None else None
    if product is not None:
        return product

    # No D1, varrer todos os produtos a cada linha da planilha consome o limite de Rows read.
    # Mantém o fallback mais pesado apenas no SQLite/local.
    if low_row_read_mode():
        return None

    key = normalize_product_lookup_key(cleaned_name)
    rows = conn.execute("SELECT * FROM products WHERE catalog_archived = 0 AND stock_tag = ?", (normalized_stock_tag,)).fetchall()
    for row in rows:
        product = row_to_product(row)
        if product and normalize_product_lookup_key(product.name) == key:
            return product
    return None


def create_or_update_product_from_material_entry(
    conn: Any,
    item_name: str,
    quantity: int,
    unit_measure: str,
    unit_price_cents: int,
    created_by_id: int | None,
    movement_type: str = "material_entry",
    note: str = "Entrada de materiais.",
    product_id: int | None = None,
) -> int:
    item_name = (item_name or "").strip()[:180]
    unit_measure = (unit_measure or "un").strip()[:40] or "un"
    quantity = int(quantity or 0)
    unit_price_cents = int(unit_price_cents or 0)
    product = None
    if product_id:
        row = conn.execute(
            "SELECT * FROM products WHERE id = ? AND catalog_archived = 0 AND stock_tag = ? LIMIT 1",
            (int(product_id), SUPPLY_STOCK_TAG),
        ).fetchone()
        product = row_to_product(row) if row is not None else None
    if product is None:
        product = find_product_by_name(conn, item_name, SUPPLY_STOCK_TAG)
    now_value = now_iso()
    if product is None:
        cursor = conn.execute(
            """
            INSERT INTO products (name, category, unit_measure, description, stock_quantity, price_cents, limit_base, limit_franchise, min_stock, max_stock, active, stock_tag, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (item_name, "Entrada de Materiais", unit_measure, "Produto criado automaticamente pela entrada de materiais.", quantity, unit_price_cents, None, None, None, None, 1, SUPPLY_STOCK_TAG, now_value),
        )
        product_id = int(get_cursor_lastrowid(cursor) or 0)
        record_stock_movement(conn, product_id, quantity, 0, quantity, movement_type, note, created_by_id=created_by_id)
        return product_id
    stock_before = int(product.stock_quantity or 0)
    stock_after = stock_before + quantity
    conn.execute(
        """
        UPDATE products
           SET stock_quantity = ?, unit_measure = ?, price_cents = CASE WHEN ? > 0 THEN ? ELSE price_cents END, updated_at = ?
         WHERE id = ?
        """,
        (stock_after, unit_measure, unit_price_cents, unit_price_cents, now_value, product.id),
    )
    record_stock_movement(conn, product.id, quantity, stock_before, stock_after, movement_type, note, created_by_id=created_by_id)
    return product.id


def create_material_entry_record(
    conn: Any,
    *,
    item_name: str,
    quantity: int,
    unit_measure: str,
    unit_price_cents: int,
    invoice_file_name: str = "",
    invoice_file_key: str = "",
    invoice_number: str = "",
    invoice_date: datetime | None = None,
    invoice_value_cents: int = 0,
    notes: str = "",
    created_by_id: int | None = None,
    movement_type: str = "material_entry",
    selected_product_id: int | None = None,
) -> int:
    product_id = create_or_update_product_from_material_entry(
        conn,
        item_name=item_name,
        quantity=quantity,
        unit_measure=unit_measure,
        unit_price_cents=unit_price_cents,
        created_by_id=created_by_id,
        movement_type=movement_type,
        note=f"Entrada de materiais: {item_name}.",
        product_id=selected_product_id,
    )
    cursor = conn.execute(
        """
        INSERT INTO material_entries (
            product_id, item_name, quantity, unit_measure, unit_price_cents,
            invoice_file_name, invoice_file_key, invoice_number, invoice_date,
            invoice_value_cents, notes, created_by_id, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            product_id,
            item_name[:180],
            int(quantity or 0),
            (unit_measure or "un")[:40],
            int(unit_price_cents or 0),
            invoice_file_name[:180],
            invoice_file_key[:500],
            invoice_number[:80],
            invoice_date.isoformat() if invoice_date else None,
            int(invoice_value_cents or 0),
            notes[:1000],
            created_by_id,
            now_iso(),
        ),
    )
    return int(get_cursor_lastrowid(cursor) or 0)


def list_products_for_material_entry_options(limit: int = 1200) -> list[Product]:
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT *
              FROM products
             WHERE active = 1
               AND catalog_archived = 0
               AND stock_tag = ?
             ORDER BY name COLLATE NOCASE ASC
             LIMIT ?
            """,
            (SUPPLY_STOCK_TAG, bounded_int(limit, 1200, 50, 3000)),
        ).fetchall()
    return [product for row in rows if (product := row_to_product(row)) is not None]


def material_entry_product_options_payload(products: list[Product]) -> list[dict[str, Any]]:
    payload: list[dict[str, Any]] = []
    for product in products:
        payload.append({
            "id": product.id,
            "name": product.name,
            "unit_measure": product.unit_measure or "un",
            "unit_price_cents": int(product.price_cents or 0),
            "unit_price_input": (f"{(int(product.price_cents or 0) / 100):.2f}".replace(".", ",") if int(product.price_cents or 0) else ""),
            "stock_quantity": int(product.stock_quantity or 0),
            "category": product.category or "",
        })
    return payload


def list_material_entries(start_dt: datetime | None = None, end_dt: datetime | None = None, limit: int | None = None) -> list[MaterialEntry]:
    clauses: list[str] = []
    params: list[Any] = []
    if start_dt is not None:
        clauses.append("me.created_at >= ?")
        params.append(start_dt.strftime("%Y-%m-%d %H:%M:%S"))
    if end_dt is not None:
        clauses.append("me.created_at <= ?")
        params.append(end_dt.strftime("%Y-%m-%d %H:%M:%S"))
    sql = """
        SELECT me.*, u.responsible_name AS created_by_name
          FROM material_entries me
          LEFT JOIN users u ON u.id = me.created_by_id
    """
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY me.created_at DESC, me.id DESC"
    if limit is not None:
        sql += " LIMIT ?"
        params.append(limit)
    with db_connect() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [entry for row in rows if (entry := row_to_material_entry(row)) is not None]


def import_material_entries_from_workbook_bytes(uploaded_bytes: bytes, created_by_id: int | None) -> tuple[int, int, list[str]]:
    workbook = load_workbook(BytesIO(uploaded_bytes), data_only=True, read_only=True)
    worksheet = workbook.active
    if not worksheet or int(getattr(worksheet, "max_row", 0) or 0) < 1:
        return 0, 0, ["planilha vazia"]
    header_row_number, header_map = detect_material_entry_header_row(worksheet)
    if not header_map_contains_alias(header_map, MATERIAL_IMPORT_HEADER_ALIASES["item_name"]):
        return 0, 0, ["não encontrei a coluna Nome do item"]
    imported = 0
    skipped = 0
    errors: list[str] = []
    with db_connect() as conn:
        for row_number, row_values_tuple in enumerate(worksheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
            row_values = worksheet_values(row_values_tuple)
            if excel_row_is_empty(row_values):
                continue
            try:
                item_name = clean_import_text(get_material_import_value(row_values, header_map, "item_name"))
                quantity = parse_optional_int(get_material_import_value(row_values, header_map, "quantity")) or 0
                unit_measure = clean_import_text(get_material_import_value(row_values, header_map, "unit_measure"), "un") or "un"
                unit_price_cents = parse_money_to_cents(get_material_import_value(row_values, header_map, "unit_price_cents"))
                invoice_number = clean_import_text(get_material_import_value(row_values, header_map, "invoice_number"))
                invoice_date = parse_optional_date(get_material_import_value(row_values, header_map, "invoice_date"))
                invoice_value_cents = parse_money_to_cents(get_material_import_value(row_values, header_map, "invoice_value_cents"))
                notes = clean_import_text(get_material_import_value(row_values, header_map, "notes"))
                if not item_name or quantity <= 0:
                    skipped += 1
                    continue
                create_material_entry_record(
                    conn,
                    item_name=item_name,
                    quantity=quantity,
                    unit_measure=unit_measure,
                    unit_price_cents=unit_price_cents,
                    invoice_number=invoice_number,
                    invoice_date=invoice_date,
                    invoice_value_cents=invoice_value_cents,
                    notes=notes,
                    created_by_id=created_by_id,
                    movement_type="material_import",
                )
                imported += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"linha {row_number}: {type(exc).__name__} - {str(exc)[:120]}")
                print(f"[ENTRADA MATERIAIS] Erro ao importar linha {row_number}: {exc}")
        conn.commit()
    try:
        workbook.close()
    except Exception:
        pass
    return imported, skipped, errors


def build_material_entries_report_pdf(entries: list[MaterialEntry], start_dt: datetime, end_dt: datetime, viewer: User) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=15*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="MaterialTitle", fontName=PDF_TEXT_FONT_BOLD, fontSize=17, leading=21, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="MaterialSub", fontName=PDF_TEXT_FONT, fontSize=8.5, leading=11, textColor=colors.HexColor("#555555")))
    styles.add(ParagraphStyle(name="MaterialCell", fontName=PDF_TEXT_FONT, fontSize=7.4, leading=9.2, textColor=colors.HexColor("#222222")))
    styles.add(ParagraphStyle(name="MaterialBold", fontName=PDF_TEXT_FONT_BOLD, fontSize=7.6, leading=9.5, textColor=colors.HexColor("#111111")))
    styles.add(ParagraphStyle(name="MaterialSmall", fontName=PDF_TEXT_FONT, fontSize=6.8, leading=8.2, textColor=colors.HexColor("#666666")))
    def p(value: Any, style: str = "MaterialCell") -> Paragraph:
        return Paragraph(pdf_clean_text(value), styles[style])
    total_qty = sum(int(entry.quantity or 0) for entry in entries)
    total_value = sum(entry.total_cents for entry in entries)
    invoices = len([entry for entry in entries if entry.invoice_number or entry.invoice_file_name])
    story: list[Any] = []
    story.append(Paragraph("Relatório de Entrada de Materiais", styles["MaterialTitle"]))
    story.append(Paragraph(f"Período: {start_dt.strftime('%d/%m/%Y')} até {end_dt.strftime('%d/%m/%Y')} • Gerado por {pdf_clean_text(viewer.responsible_name)}", styles["MaterialSub"]))
    story.append(Spacer(1, 6*mm))
    summary = Table([
        [p("Entradas", "MaterialBold"), p("Quantidade total", "MaterialBold"), p("Valor total", "MaterialBold"), p("Notas fiscais", "MaterialBold")],
        [p(str(len(entries))), p(str(total_qty)), p(format_brl(total_value)), p(str(invoices))],
    ], colWidths=[42*mm, 42*mm, 42*mm, 42*mm])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E60012")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#F8F8F8")),
        ("BOX", (0,0), (-1,-1), 0.4, colors.HexColor("#DDDDDD")),
        ("INNERGRID", (0,0), (-1,-1), 0.3, colors.HexColor("#E5E5E5")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
    ]))
    story.append(summary)
    story.append(Spacer(1, 7*mm))
    rows = [[p("Data", "MaterialBold"), p("Item", "MaterialBold"), p("Qtd.", "MaterialBold"), p("Un.", "MaterialBold"), p("Valor unit.", "MaterialBold"), p("Nota fiscal", "MaterialBold"), p("Observações", "MaterialBold")]]
    if entries:
        for entry in entries:
            invoice_parts = []
            if entry.invoice_number:
                invoice_parts.append(f"Nº {entry.invoice_number}")
            if entry.invoice_date:
                invoice_parts.append(entry.invoice_date.strftime("%d/%m/%Y"))
            if entry.invoice_value_cents:
                invoice_parts.append(format_brl(entry.invoice_value_cents))
            if entry.invoice_file_name:
                invoice_parts.append(entry.invoice_file_name)
            rows.append([
                p(format_sao_paulo_datetime(entry.created_at)),
                p(entry.item_name, "MaterialBold"),
                p(str(entry.quantity)),
                p(entry.unit_measure),
                p(format_brl(entry.unit_price_cents)),
                p("<br/>".join(invoice_parts) if invoice_parts else "-"),
                p(entry.notes or "-"),
            ])
    else:
        rows.append([p("Nenhuma entrada encontrada no período.", "MaterialCell"), "", "", "", "", "", ""])
    table = Table(rows, colWidths=[23*mm, 44*mm, 15*mm, 16*mm, 25*mm, 37*mm, 30*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("BOX", (0,0), (-1,-1), 0.4, colors.HexColor("#D9D9D9")),
        ("INNERGRID", (0,0), (-1,-1), 0.25, colors.HexColor("#E8E8E8")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("SPAN", (0,1), (-1,1)) if not entries else ("LINEBELOW", (0,0), (-1,0), 0.5, colors.HexColor("#E60012")),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer

# ---------- Autenticação ----------

@app.route("/")
@login_required
@page_access_required("home")
def home():
    return render_template("index.html", product_categories=list_product_categories(require_current_user()))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = normalize_username(request.form.get("username", ""))
        password = request.form.get("password", "")
        user = get_user_by_username(username)

        if user is None or not check_password_hash(user.password_hash, password):
            flash("Nome de usuário ou senha inválidos.", "danger")
            return redirect(url_for("login"))

        if not user.is_approved:
            flash("Seu cadastro ainda não foi aprovado por um administrador.", "warning")
            return redirect(url_for("login"))

        session.pop("pending_admin_user_id", None)
        session["user_id"] = user.id
        flash("Login realizado com sucesso.", "success")
        if user.is_admin:
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("home"))

    return render_template("login.html")


@app.route("/verify-admin", methods=["GET", "POST"])
def verify_admin():
    pending_id = session.get("pending_admin_user_id")
    user = get_user(int(pending_id)) if pending_id is not None else None
    if user is None:
        flash("Faça login novamente.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        code = request.form.get("code", "").strip()
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT * FROM admin_login_codes
                WHERE user_id = ? AND used_at IS NULL
                ORDER BY created_at DESC
                LIMIT 5
                """,
                (user.id,),
            ).fetchall()
            for row in rows:
                expires = parse_dt(row["expires_at"]) or datetime.utcnow()
                if expires < datetime.utcnow():
                    continue
                if check_password_hash(row["code_hash"], code):
                    conn.execute("UPDATE admin_login_codes SET used_at = ? WHERE id = ?", (now_iso(), row["id"]))
                    conn.commit()
                    session.pop("pending_admin_user_id", None)
                    session["user_id"] = user.id
                    flash("Login confirmado com sucesso.", "success")
                    return redirect(url_for("admin_dashboard"))

        flash("Código inválido ou expirado.", "danger")
        return redirect(url_for("verify_admin"))

    return render_template("verify_admin.html", username=user.username, minutes=ADMIN_CODE_MINUTES)


@app.route("/register", methods=["GET", "POST"])
def register():
    flash("A solicitação pública de acesso foi desativada. Os logins devem ser criados dentro da plataforma por um cargo com permissão.", "warning")
    return redirect(url_for("login"))


@app.route("/logout")
def logout():
    session.clear()
    flash("Você saiu do portal.", "success")
    return redirect(url_for("login"))


# ---------- API Usuário ----------

@app.get("/api/products")
@login_required
@page_access_required("home")
def api_products():
    user = require_current_user()
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    sort = request.args.get("sort", "name").strip().lower()
    limit = api_page_limit(default=120, maximum=250)
    sql = """
        SELECT id, name, category, category_emoji, image_name, image_key, image_content_type,
               unit_measure, is_kit, kit_quantity, description, stock_quantity, price_cents,
               limit_base, limit_franchise, limit_block_days, min_order_quantity, min_stock, max_stock,
               active, visible_base, visible_franchise, internal, catalog_archived, created_at, updated_at
         FROM products
         WHERE active = 1 AND catalog_archived = 0
           AND stock_tag = ?
    """
    params: list[Any] = [SUPPLY_STOCK_TAG]
    if not user.is_admin:
        sql += " AND stock_quantity > 0 AND COALESCE(internal, 0) = 0"
    if user.role == "base":
        sql += " AND visible_base = 1"
    elif user.role == "franchise":
        sql += " AND visible_franchise = 1"
    if q:
        like = like_term(q)
        sql += " AND (name LIKE ? OR category LIKE ? OR description LIKE ? OR unit_measure LIKE ?)"
        params.extend([like, like, like, like])
    if category:
        sql += " AND LOWER(TRIM(COALESCE(category, ''))) = LOWER(TRIM(?))"
        params.append(category)
    sort_map = {
        "name": "name COLLATE NOCASE ASC",
        "stock_desc": "stock_quantity DESC, name COLLATE NOCASE ASC",
        "price_asc": "price_cents ASC, name COLLATE NOCASE ASC",
        "price_desc": "price_cents DESC, name COLLATE NOCASE ASC",
    }
    sql += " ORDER BY " + sort_map.get(sort, sort_map["name"])
    sql += " LIMIT ?"
    params.append(limit)
    with db_connect() as conn:
        rows = conn.execute(sql, params).fetchall()
    products = [product for row in rows if (product := row_to_product(row)) is not None]
    return jsonify([product_to_api(product, user) for product in products])


@app.get("/products/<int:product_id>/image")
@login_required
def product_image(product_id: int):
    product = get_product(product_id)
    if product is None or not product.image_key:
        abort(404)
    user = require_current_user()
    if not user.is_admin and product.internal:
        abort(404)
    if user.role == "base" and not product.visible_base:
        abort(404)
    if user.role == "franchise" and not product.visible_franchise:
        abort(404)
    local_path = local_product_image_path(product.image_key)
    if local_path.exists():
        return send_file(
            local_path,
            mimetype=product.image_content_type or PRODUCT_IMAGE_TYPES.get(local_path.suffix.lower(), "application/octet-stream"),
            max_age=3600,
        )
    try:
        downloaded = download_bytes_from_r2(product.image_key)
    except Exception:
        downloaded = None
    if not downloaded:
        abort(404)
    data, content_type = downloaded
    return send_file(
        BytesIO(data),
        mimetype=product.image_content_type or content_type,
        max_age=3600,
    )


@app.post("/api/requests")
@login_required
@page_access_required("home")
def api_create_request():
    user = require_current_user()
    payload = request.get_json(silent=True) or {}
    normalized, error = validate_items_for_user(payload.get("items"), user)
    if error:
        return jsonify({"ok": False, "message": error}), 400

    user_note = str(payload.get("user_note") or "").strip()
    people_count: int | None = None
    if user.role == "base":
        people_count_raw = str(payload.get("people_count") or "").strip()
        try:
            people_count = int(people_count_raw)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "message": "Informe o número de pessoas na base."}), 400
        if people_count <= 0 or people_count > 99999:
            return jsonify({"ok": False, "message": "Informe um número de pessoas válido."}), 400

    with db_connect() as conn:
        cursor = conn.execute(
            """
            INSERT INTO supply_requests (user_id, status, user_note, admin_note, people_count, created_at)
            VALUES (?, 'pending', ?, '', ?, ?)
            """,
            (user.id, user_note, people_count, now_iso()),
        )
        request_id = get_cursor_lastrowid(cursor)
        if request_id is None:
            conn.rollback()
            return jsonify({"ok": False, "message": "Não foi possível registrar a solicitação."}), 500

        for product, quantity, _requested_limit_quantity in normalized:
            conn.execute(
                """
                INSERT INTO request_items (request_id, product_id, product_name_snapshot, quantity, price_cents_snapshot)
                VALUES (?, ?, ?, ?, ?)
                """,
                (request_id, product.id, product.name, quantity, product.price_cents),
            )
        record_request_action(conn, int(request_id), "created", user.id, "Solicitação criada pelo solicitante.")
        apply_automatic_blocks_after_request(conn, user, int(request_id), normalized)
        conn.commit()

    try:
        created_request = get_supply_request(int(request_id))
        if created_request is not None:
            request_link = public_url_for("admin_request_detail", request_id=int(request_id))
            notify_feishu_supply_request_created(created_request, request_link)
    except Exception:
        app.logger.exception("Falha ao preparar notificacao Feishu da solicitacao")

    return jsonify({"ok": True, "message": "Solicitação enviada para aprovação.", "request_id": request_id})


@app.get("/minhas-solicitacoes")
@login_required
@page_access_required("my_requests")
def my_requests():
    user = require_current_user()
    request_filters = normalize_request_filters_from_args()
    extra_args = request_filters_to_query_args(request_filters)
    requests_list, request_pagination = list_supply_requests_page(
        user_id=user.id,
        endpoint="my_requests",
        filters=request_filters,
        extra_args=extra_args,
    )
    return render_template(
        "my_requests.html",
        requests_list=requests_list,
        request_pagination=request_pagination,
        request_filters=request_filters,
        request_product_options=list_products_for_request_filters(),
        request_sort_options=REQUEST_SORT_OPTIONS,
        request_type_options=REQUEST_TYPE_OPTIONS,
        request_regional_options=REQUEST_REGIONAL_OPTIONS,
        request_filter_action="my_requests",
        request_filter_show_type=True,
        request_filter_show_regional=True,
    )




@app.get("/solicitacoes/<int:request_id>/pdf")
@login_required
def request_pdf(request_id: int):
    viewer = require_current_user()
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    if not viewer.is_admin and supply_request.user_id != viewer.id:
        abort(403)
    if viewer.is_admin and supply_request.user_id != viewer.id and not can_view_supply_request_by_assignment(supply_request, viewer):
        abort(403)

    requester = supply_request.user
    org_name = (requester.organization_name if requester else "solicitacao").replace(";", " ")
    filename = f"solicitacao_{supply_request.id}_{safe_filename(org_name)}.pdf"
    buffer = build_request_pdf(supply_request, viewer)
    store_generated_file(
        storage_key("pdfs", str(supply_request.id), filename),
        buffer,
        "application/pdf",
        {"request_id": str(supply_request.id), "organization": org_name},
    )
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.get("/admin/assets/<int:asset_id>/pdf")
@admin_required
@page_access_required("admin_stock")
def asset_pdf(asset_id: int):
    denied = require_action_permission("stock_reports", "Seu tipo de acesso não pode gerar PDFs de ativos.", "admin_assets")
    if denied:
        return denied
    viewer = require_current_user()
    asset = get_asset(asset_id)
    if asset is None:
        abort(404)

    unit_name = asset.base.replace(";", " ")
    filename = f"ativo_{asset.id}_{safe_filename(unit_name)}.pdf"
    buffer = build_asset_pdf(asset, viewer)
    store_generated_file(
        storage_key("pdfs", "ativos", str(asset.id), filename),
        buffer,
        "application/pdf",
        {"asset_id": str(asset.id), "unit": unit_name, "regional": asset.regional},
    )
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


# ---------- Admin ----------

@app.route("/admin")
@admin_required
@page_access_required("admin_dashboard")
def admin_dashboard():
    with db_connect() as conn:
        if exact_counts_enabled():
            counts = {
                "users_pending": conn.execute("SELECT COUNT(*) AS total FROM users WHERE status = 'pending'").fetchone()["total"],
                "requests_pending": count_supply_requests(status="pending", viewer=current_user(), apply_assignment_visibility=True),
                "products": conn.execute("SELECT COUNT(*) AS total FROM products WHERE catalog_archived = 0").fetchone()["total"],
                "stock_total": conn.execute("SELECT COALESCE(SUM(stock_quantity), 0) AS total FROM products WHERE catalog_archived = 0").fetchone()["total"],
            }
        else:
            user_pending_rows = conn.execute("SELECT id FROM users WHERE status = 'pending' ORDER BY id DESC LIMIT 50").fetchall()
            request_pending_rows = []
            product_rows = conn.execute("SELECT id, stock_quantity FROM products WHERE catalog_archived = 0 ORDER BY id DESC LIMIT 200").fetchall()
            counts = {
                "users_pending": len(user_pending_rows),
                "requests_pending": count_supply_requests(status="pending", viewer=current_user(), apply_assignment_visibility=True),
                "products": len(product_rows),
                "stock_total": sum(int(row["stock_quantity"] or 0) for row in product_rows),
            }
        low_rows = conn.execute(
            """
            SELECT id, name, category, category_emoji, image_name, image_key, image_content_type,
                   unit_measure, is_kit, kit_quantity, description, stock_quantity, price_cents,
                   limit_base, limit_franchise, limit_block_days, min_order_quantity, min_stock, max_stock,
                   active, visible_base, visible_franchise, internal, catalog_archived, created_at, updated_at
              FROM products
             WHERE catalog_archived = 0 AND stock_quantity <= 20
             ORDER BY stock_quantity ASC
             LIMIT 8
            """
        ).fetchall()
    low_stock = [product for row in low_rows if (product := row_to_product(row)) is not None]
    dashboard_pages = get_user_page_permissions(current_user())
    if "admin_requests" in dashboard_pages:
        latest_requests = list_supply_requests(limit=8, viewer=current_user(), apply_assignment_visibility=True)
    elif "admin_requests_attended" in dashboard_pages:
        latest_requests = list_supply_requests(status="approved", limit=8, viewer=current_user(), apply_assignment_visibility=True)
    else:
        latest_requests = []
    return render_template("admin/dashboard.html", counts=counts, low_stock=low_stock, latest_requests=latest_requests)




@app.route("/admin/users")
@admin_required
@page_access_required("admin_users")
def admin_users():
    selected_status = (request.args.get("status", "") or "").strip().lower()
    if selected_status not in {"", "pending", "approved", "rejected"}:
        selected_status = ""

    selected_role = (request.args.get("role", "") or "").strip().lower()
    valid_filter_roles = {"", *all_role_options()}
    if selected_role not in valid_filter_roles:
        selected_role = ""

    search_query = (request.args.get("q", "") or "").strip()
    selected_sort = (request.args.get("sort", "responsible_asc") or "responsible_asc").strip().lower()
    sort_map = {
        "newest": "created_at DESC, id DESC",
        "oldest": "created_at ASC, id ASC",
        "responsible_asc": "responsible_name COLLATE NOCASE ASC, id DESC",
        "responsible_desc": "responsible_name COLLATE NOCASE DESC, id DESC",
        "username_asc": "username COLLATE NOCASE ASC, id DESC",
        "username_desc": "username COLLATE NOCASE DESC, id DESC",
        "role_asc": "role COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
        "unit_asc": "organization_name COLLATE NOCASE ASC, franchise_name COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
        "status_asc": "status COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
    }
    order_clause = sort_map.get(selected_sort, sort_map["responsible_asc"])
    if selected_sort not in sort_map:
        selected_sort = "responsible_asc"

    per_page = list_page_limit(default=DEFAULT_TABLE_PAGE_SIZE, maximum=500)
    page = bounded_int(request.args.get("page"), 1, 1, 100000)

    clauses: list[str] = []
    params: list[Any] = []
    if selected_status:
        clauses.append("status = ?")
        params.append(selected_status)
    if selected_role:
        clauses.append("role = ?")
        params.append(selected_role)
    if search_query:
        like = like_term(search_query)
        clauses.append(
            """(
                responsible_name LIKE ?
                OR username LIKE ?
                OR organization_name LIKE ?
                OR franchise_name LIKE ?
                OR franchise_number LIKE ?
                OR cnpj LIKE ?
                OR role LIKE ?
                OR status LIKE ?
            )"""
        )
        params.extend([like, like, like, like, like, like, like, like])

    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"""
        SELECT id, responsible_name, organization_name, franchise_name, franchise_number,
               cnpj, username, email, password_hash, role, status, created_at,
               updated_at, page_permissions_configured
          FROM users
          {where_sql}
         ORDER BY {order_clause}
         LIMIT ? OFFSET ?
    """

    with db_connect() as conn:
        total_row = conn.execute(f"SELECT COUNT(*) AS total FROM users{where_sql}", params).fetchone()
        total_users = int((total_row["total"] if total_row else 0) or 0)
        total_pages = max(1, (total_users + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page
        rows = conn.execute(sql, [*params, per_page, offset]).fetchall()
        if exact_counts_enabled():
            status_rows = conn.execute(f"SELECT status, COUNT(*) AS total FROM users{where_sql} GROUP BY status", params).fetchall()
            role_rows = conn.execute(f"SELECT role, COUNT(*) AS total FROM users{where_sql} GROUP BY role", params).fetchall()
        else:
            status_rows = []
            role_rows = []

    users = [user for row in rows if (user := row_to_user(row)) is not None]
    if exact_counts_enabled():
        status_counts = {row["status"]: int(row["total"] or 0) for row in status_rows}
        role_counts = {row["role"]: int(row["total"] or 0) for row in role_rows}
    else:
        status_counts: dict[str, int] = {}
        role_counts: dict[str, int] = {}
        for user in users:
            status_counts[user.status] = status_counts.get(user.status, 0) + 1
            role_counts[user.role] = role_counts.get(user.role, 0) + 1

    page_size_options = list(TABLE_PAGE_SIZE_OPTIONS)
    if per_page not in page_size_options:
        page_size_options.append(per_page)
        page_size_options.sort()

    def page_url(target_page: int, target_limit: int | None = None) -> str:
        args: dict[str, Any] = {"page": max(1, target_page), "limit": target_limit or per_page}
        if search_query:
            args["q"] = search_query
        if selected_status:
            args["status"] = selected_status
        if selected_role:
            args["role"] = selected_role
        if selected_sort:
            args["sort"] = selected_sort
        return url_for("admin_users", **args)

    visible_page_numbers: set[int] = {1, total_pages}
    for number in range(page - 2, page + 3):
        if 1 <= number <= total_pages:
            visible_page_numbers.add(number)
    page_links: list[dict[str, Any]] = []
    previous_number = 0
    for number in sorted(visible_page_numbers):
        if previous_number and number - previous_number > 1:
            page_links.append({"ellipsis": True})
        page_links.append({"number": number, "active": number == page, "url": page_url(number)})
        previous_number = number

    start_item = offset + 1 if total_users else 0
    end_item = min(offset + len(users), total_users)
    user_pagination = {
        "page": page,
        "limit": per_page,
        "total": total_users,
        "total_pages": total_pages,
        "start": start_item,
        "end": end_item,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "first_url": page_url(1),
        "prev_url": page_url(max(1, page - 1)),
        "next_url": page_url(min(total_pages, page + 1)),
        "last_url": page_url(total_pages),
        "page_links": page_links,
        "page_size_options": page_size_options,
    }

    user_counts = {
        "shown": len(users),
        "total": int(total_users or 0),
        "pending": status_counts.get("pending", 0),
        "approved": status_counts.get("approved", 0),
        "rejected": status_counts.get("rejected", 0),
        "dev": role_counts.get("dev", 0),
        "admin": role_counts.get("admin", 0),
        "base": role_counts.get("base", 0),
        "franchise": role_counts.get("franchise", 0),
        "page": page,
        "limit": per_page,
        "low_read": low_row_read_mode(),
    }
    user_filters = {
        "q": search_query,
        "status": selected_status,
        "role": selected_role,
        "sort": selected_sort,
        "limit": per_page,
        "page": page,
    }
    return render_template(
        "admin/users.html",
        users=users,
        selected_status=selected_status,
        user_filters=user_filters,
        user_counts=user_counts,
        user_pagination=user_pagination,
        role_labels=role_option_labels(),
        role_filter_options=all_role_options(),
    )


def filter_and_sort_users_for_export(users: list[User], q: str, status: str, role: str, sort_mode: str) -> list[User]:
    normalized_query = normalize_header(q).replace("_", " ").strip()
    terms = [term for term in normalized_query.split() if term]

    def searchable(user: User) -> str:
        return normalize_header(" ".join([
            user.responsible_name,
            user.username,
            user.organization_name,
            user.franchise_name,
            user.franchise_number,
            user.formatted_phone,
            user.cnpj,
            user.formatted_cnpj,
            user_role_label(user.role),
            status_label(user.status),
        ])).replace("_", " ")

    filtered: list[User] = []
    for user in users:
        if status and user.status != status:
            continue
        if role and user.role != role:
            continue
        text = searchable(user)
        if terms and not all(term in text for term in terms):
            continue
        filtered.append(user)

    def unit_name(user: User) -> str:
        return (user.franchise_name or user.organization_name or "").casefold()

    sort_mode = sort_mode if sort_mode in {
        "newest", "oldest", "responsible_asc", "responsible_desc", "username_asc", "username_desc", "role_asc", "unit_asc", "status_asc"
    } else "responsible_asc"
    if sort_mode == "oldest":
        filtered.sort(key=lambda user: (user.created_at or datetime.min, user.id))
    elif sort_mode == "responsible_asc":
        filtered.sort(key=lambda user: (user.responsible_name.casefold(), -user.id))
    elif sort_mode == "responsible_desc":
        filtered.sort(key=lambda user: (user.responsible_name.casefold(), user.id), reverse=True)
    elif sort_mode == "username_asc":
        filtered.sort(key=lambda user: (user.username.casefold(), -user.id))
    elif sort_mode == "username_desc":
        filtered.sort(key=lambda user: (user.username.casefold(), user.id), reverse=True)
    elif sort_mode == "role_asc":
        filtered.sort(key=lambda user: (user.role.casefold(), user.responsible_name.casefold(), user.id))
    elif sort_mode == "unit_asc":
        filtered.sort(key=lambda user: (unit_name(user), user.responsible_name.casefold(), user.id))
    elif sort_mode == "status_asc":
        filtered.sort(key=lambda user: (user.status.casefold(), user.responsible_name.casefold(), user.id))
    else:
        filtered.sort(key=lambda user: (user.responsible_name.casefold(), -user.id))
    return filtered


@app.get("/admin/users/export")
@admin_required
@page_access_required("admin_users")
def admin_users_export():
    denied = require_action_permission("users_import_export", "Seu tipo de acesso não pode exportar usuários.", "admin_users")
    if denied:
        return denied
    selected_status = (request.args.get("status", "") or "").strip().lower()
    if selected_status not in {"", "pending", "approved", "rejected"}:
        selected_status = ""
    selected_role = (request.args.get("role", "") or "").strip().lower()
    valid_filter_roles = {"", *all_role_options()}
    if selected_role not in valid_filter_roles:
        selected_role = ""
    search_query = (request.args.get("q", "") or "").strip()
    selected_sort = (request.args.get("sort", "responsible_asc") or "responsible_asc").strip().lower()
    sort_map = {
        "newest": "created_at DESC, id DESC",
        "oldest": "created_at ASC, id ASC",
        "responsible_asc": "responsible_name COLLATE NOCASE ASC, id DESC",
        "responsible_desc": "responsible_name COLLATE NOCASE DESC, id DESC",
        "username_asc": "username COLLATE NOCASE ASC, id DESC",
        "username_desc": "username COLLATE NOCASE DESC, id DESC",
        "role_asc": "role COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
        "unit_asc": "organization_name COLLATE NOCASE ASC, franchise_name COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
        "status_asc": "status COLLATE NOCASE ASC, responsible_name COLLATE NOCASE ASC",
    }
    order_clause = sort_map.get(selected_sort, sort_map["responsible_asc"])
    export_limit = bounded_int(request.args.get("limit"), int(os.getenv("D1_EXPORT_LIMIT", "500")), 25, 2000)
    export_page = bounded_int(request.args.get("page"), 1, 1, 100000)
    export_offset = (export_page - 1) * export_limit

    clauses: list[str] = []
    params: list[Any] = []
    if selected_status:
        clauses.append("status = ?")
        params.append(selected_status)
    if selected_role:
        clauses.append("role = ?")
        params.append(selected_role)
    if search_query:
        like = like_term(search_query)
        clauses.append(
            """(
                responsible_name LIKE ?
                OR username LIKE ?
                OR organization_name LIKE ?
                OR franchise_name LIKE ?
                OR franchise_number LIKE ?
                OR cnpj LIKE ?
                OR role LIKE ?
                OR status LIKE ?
            )"""
        )
        params.extend([like, like, like, like, like, like, like, like])
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""

    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT id, responsible_name, organization_name, franchise_name, franchise_number,
                   cnpj, username, email, password_hash, role, status, created_at,
                   updated_at, page_permissions_configured
              FROM users
              {where_sql}
             ORDER BY {order_clause}
             LIMIT ? OFFSET ?
            """,
            [*params, export_limit, export_offset],
        ).fetchall()
    users = [user for row in rows if (user := row_to_user(row)) is not None]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Usuários"
    headers = [
        "Responsável",
        "Usuário",
        "Tipo de acesso",
        "Setor",
        "Telefone",
        "CNPJ",
        "Status",
        "Criado em (GMT-3)",
        "Atualizado em (GMT-3)",
    ]
    worksheet.append(headers)
    for user in users:
        worksheet.append([
            user.responsible_name,
            user.username,
            user_role_label(user.role),
            user.franchise_name or user.organization_name,
            user.formatted_phone if user.role == "franchise" else "",
            user.formatted_cnpj or "",
            status_label(user.status),
            format_sao_paulo_datetime(user.created_at) if user.created_at else "",
            format_sao_paulo_datetime(user.updated_at) if user.updated_at else "",
        ])

    header_fill = PatternFill("solid", fgColor="E60012")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    widths = [34, 24, 18, 34, 22, 22, 16, 20, 20]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:I{max(1, worksheet.max_row)}"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"usuarios_tabela_atual_{sao_paulo_filename_timestamp()}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.get("/admin/users/model")
@admin_required
@page_access_required("admin_users")
def admin_users_template():
    denied = require_action_permission("users_import_export", "Seu tipo de acesso não pode baixar o modelo de usuários.", "admin_users")
    if denied:
        return denied
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Usuários"
    headers = [
        "Nome do responsável",
        "Nome de usuário",
        "Senha",
        "Tipo de acesso",
        "Status do cadastro",
        "Nome da base",
        "Nome da franquia",
        "Telefone",
        "CNPJ",
    ]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="E60012")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = [32, 24, 24, 20, 22, 24, 34, 26, 22]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:I1"

    lists = workbook.create_sheet("Listas")
    lists.append(["Tipos de acesso", "Status", "Bases"])
    import_role_labels = [role_option_labels().get(role, role) for role in all_role_options()]
    for row_index, role_label in enumerate(import_role_labels, start=2):
        lists.cell(row=row_index, column=1, value=role_label)
    for row_index, status_label in enumerate(["Aprovado", "Pendente", "Recusado"], start=2):
        lists.cell(row=row_index, column=2, value=status_label)
    for row_index, base_name in enumerate(BASE_UNIT_OPTIONS, start=2):
        lists.cell(row=row_index, column=3, value=base_name)
    lists.sheet_state = "hidden"

    role_validation = DataValidation(type="list", formula1=f"'Listas'!$A$2:$A${max(2, len(import_role_labels) + 1)}", allow_blank=False)
    status_validation = DataValidation(type="list", formula1="'Listas'!$B$2:$B$4", allow_blank=False)
    base_validation = DataValidation(
        type="list",
        formula1=f"'Listas'!$C$2:$C${max(2, len(BASE_UNIT_OPTIONS) + 1)}",
        allow_blank=True,
    )
    for validation in [role_validation, status_validation, base_validation]:
        worksheet.add_data_validation(validation)
    role_validation.add("D2:D1000")
    status_validation.add("E2:E1000")
    base_validation.add("F2:F1000")
    for row in range(2, 1001):
        worksheet.cell(row=row, column=8).number_format = "@"
        worksheet.cell(row=row, column=9).number_format = "@"

    instructions = workbook.create_sheet("Instruções")
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 95
    instructions.append(["Campo", "Como preencher"])
    instruction_rows = [
        ("Base", "Preencha Nome da base. Deixe Nome da franquia, Telefone e CNPJ vazios."),
        ("Franquia", "Preencha Nome da franquia e Telefone com DDD. CNPJ é opcional. Deixe Nome da base vazio."),
        ("Administrador", "Deixe os campos de base, franquia e CNPJ vazios. O administrador recebe acesso a todas as páginas."),
        ("Dev", "Use apenas com autorização. Somente um usuário Dev pode importar/promover usuários como Dev."),
        ("Senha", "Obrigatória para novos usuários. Para usuários existentes, o Dev pode marcar a opção de atualizar senhas para substituir pela senha da planilha; sem essa opção, a senha atual é preservada."),
        ("Status", "Use Aprovado, Pendente ou Recusado."),
    ]
    for row in instruction_rows:
        instructions.append(row)
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row in instructions.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="modelo_cadastro_usuarios.xlsx",
    )


@app.post("/admin/users/import")
@admin_required
@page_access_required("admin_users")
def admin_users_import():
    denied = require_action_permission("users_import_export", "Seu tipo de acesso não pode importar usuários.", "admin_users")
    if denied:
        return denied
    uploaded = request.files.get("spreadsheet")
    if uploaded is None or not uploaded.filename:
        flash("Selecione uma planilha .xlsx de usuários.", "warning")
        return redirect_to_return("admin_users")
    if not uploaded.filename.lower().endswith(".xlsx"):
        flash("Importe apenas arquivos .xlsx.", "warning")
        return redirect_to_return("admin_users")

    try:
        uploaded_bytes = uploaded.read()
        if not uploaded_bytes:
            flash("A planilha enviada está vazia.", "warning")
            return redirect_to_return("admin_users")
        import_mode = (request.form.get("import_mode") or "merge").strip().lower()
        if import_mode not in {"merge", "replace"}:
            import_mode = "merge"
        # Salvar cópia no R2 é opcional. Para planilhas grandes, não bloqueia a importação.
        if len(uploaded_bytes) <= int(os.getenv("IMPORT_BACKUP_MAX_BYTES", "5242880")):
            try:
                upload_bytes_to_r2(
                    storage_key(
                        "imports",
                        "usuarios_" + sao_paulo_filename_timestamp() + "_" + safe_filename(uploaded.filename),
                    ),
                    uploaded_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    {"type": "users_import", "mode": import_mode},
                )
            except Exception as exc:
                print(f"[R2] Não foi possível salvar a planilha de usuários: {exc}")

        current_user = require_current_user()
        update_existing_passwords = bool(current_user.is_dev and request.form.get("update_existing_passwords") == "on")
        created, updated, skipped, replaced, errors = import_users_from_workbook_bytes(
            uploaded_bytes,
            import_mode=import_mode,
            current_user_id=current_user.id,
            current_user_role=current_user.role,
            update_passwords=update_existing_passwords,
        )
        if errors:
            preview = "; ".join(errors[:5])
            suffix = "" if len(errors) <= 5 else f"; +{len(errors) - 5} outro(s) erro(s)."
            flash(f"Algumas linhas foram ignoradas: {preview}{suffix}", "warning")
        mode_message = " Cadastros fora da planilha foram recusados, exceto o usuário logado." if import_mode == "replace" else " Usuários repetidos foram atualizados pelo login."
        flash(
            f"Importação concluída: {created} criado(s), {updated} atualizado(s), {skipped} linha(s) ignorada(s), {replaced} substituído(s).{mode_message}",
            "success" if created or updated or replaced else "warning",
        )
    except Exception as exc:
        app.logger.exception("Falha ao importar usuários")
        flash(f"Não consegui importar a planilha. Erro tratado: {type(exc).__name__}. Veja os logs se continuar acontecendo.", "danger")
    return redirect_to_return("admin_users")


@app.route("/admin/users/new", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_users")
def admin_user_new():
    denied = require_action_permission("users_create_edit", "Seu tipo de acesso não pode criar usuários.", "admin_users")
    if denied:
        return denied
    if request.method == "POST":

        responsible_name = request.form.get("responsible_name", "").strip()
        username = normalize_username(request.form.get("username", ""))
        email = synthetic_email_for_username(username)
        password = request.form.get("password", "")
        current = require_current_user()
        requested_role = normalize_user_role(request.form.get("role", "base"), allow_admin=True) or "base"
        if not can_create_user_role(current, requested_role):
            flash("Seu tipo de acesso nao pode criar usuarios com esse cargo.", "warning")
            return redirect_to_return("admin_users")
        required_create_fields = {"role", "status"}
        if not current.is_dev and not required_create_fields.issubset(get_user_editable_fields(current)):
            flash("Seu tipo de acesso nao tem todos os campos necessarios para criar usuarios.", "warning")
            return redirect_to_return("admin_users")
        role_warning = None
        with db_connect() as role_conn:
            role, role_warning = safe_role_for_update(current, None, requested_role, request.form.get("dev_password", ""), role_conn)
            password_warning = maybe_update_dev_password_from_form(role_conn, current, request.form, None)
            if password_warning:
                role_conn.rollback()
                flash(password_warning, "warning")
                return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=default_page_keys_for_role("base"), allowed_role_options=allowed_role_options_for_editor(current), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current), can_edit_dev_password=can_edit_dev_password(current))
            role_conn.commit()
        if role_warning:
            flash(role_warning, "warning")
            if requested_role == "dev" and role != "dev":
                return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=default_page_keys_for_role("base"), allowed_role_options=allowed_role_options_for_editor(current), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current), can_edit_dev_password=can_edit_dev_password(current))
        status = normalize_user_status(request.form.get("status", "approved"), default="approved") or "approved"
        selected_pages = request.form.getlist("page_permissions") or list(default_page_keys_for_role(role))
        selected_actions = request.form.getlist("action_permissions")

        try:
            organization_name, franchise_name, franchise_number, cnpj = validate_user_profile_fields(
                role,
                organization_name=request.form.get("organization_name", ""),
                franchise_name=request.form.get("franchise_name", ""),
                franchise_number=request.form.get("franchise_number", ""),
                cnpj=request.form.get("cnpj", ""),
                strict_base=not current.is_dev,
            )
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))
        selected_pages = list(normalize_user_page_selection_for_editor(current, role, selected_pages))
        if not selected_pages:
            selected_pages = list(default_page_keys_for_role(role))
        selected_actions = list(normalize_user_action_selection_for_editor(current, role, selected_actions, selected_pages))
        if not responsible_name or not username or not password:
            flash("Preencha responsável, nome de usuário e senha.", "danger")
            return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))
        if not valid_username(username):
            flash("Use um nome de usuário com 3 a 40 caracteres: letras, números, ponto, hífen ou underline.", "danger")
            return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))
        if get_user_by_username(username) is not None:
            flash("Já existe usuário cadastrado com este nome de usuário.", "danger")
            return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))

        with db_connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO users (
                    responsible_name, organization_name, franchise_name, franchise_number, cnpj,
                    username, email, password_hash, role, status, created_at, page_permissions_configured, action_permissions_configured
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 1)
                """,
                (
                    responsible_name,
                    organization_name,
                    franchise_name,
                    franchise_number,
                    cnpj,
                    username,
                    email,
                    generate_password_hash(password),
                    role,
                    status,
                    now_iso(),
                ),
            )
            new_user_id = get_cursor_lastrowid(cursor)
            if new_user_id is None:
                conn.rollback()
                flash("Não foi possível adicionar o usuário.", "danger")
                return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))
            save_user_page_permissions(conn, new_user_id, role, selected_pages)
            save_user_action_permissions(conn, new_user_id, role, selected_actions, selected_pages)
            conn.commit()

        flash("Usuário adicionado com sucesso.", "success")
        return redirect_to_return("admin_users")

    return render_template("admin/user_form.html", user=None, is_new=True, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=default_page_keys_for_role("base"), allowed_role_options=allowed_role_options_for_editor(require_current_user()), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(require_current_user()), can_edit_dev_password=can_edit_dev_password(require_current_user()))


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_users")
def admin_user_edit(user_id: int):
    denied = require_action_permission("users_create_edit", "Seu tipo de acesso não pode editar usuários.", "admin_users")
    if denied:
        return denied
    target = get_user(user_id)
    if target is None:
        abort(404)

    current = require_current_user()
    if not can_edit_user_target(current, target):
        flash("Seu tipo de acesso nao pode editar esse usuario.", "warning")
        return redirect_to_return("admin_users")

    if request.method == "POST":
        responsible_name = request.form.get("responsible_name", "").strip()
        username = normalize_username(request.form.get("username", ""))
        email = synthetic_email_for_username(username)
        requested_role = normalize_user_role(request.form.get("role", target.role), allow_admin=True) or target.role
        role_warning = None
        with db_connect() as role_conn:
            role, role_warning = safe_role_for_update(current, target, requested_role, request.form.get("dev_password", ""), role_conn)
            password_warning = maybe_update_dev_password_from_form(role_conn, current, request.form, target)
            if password_warning:
                role_conn.rollback()
                flash(password_warning, "warning")
                return render_template("admin/user_form.html", user=target, is_new=False, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=selected_permissions_for_form(target), allowed_role_options=allowed_role_options_for_editor(current, target), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current, target), can_edit_dev_password=can_edit_dev_password(current))
            role_conn.commit()
        if role_warning:
            flash(role_warning, "warning")
        status = normalize_user_status(request.form.get("status", "approved"), default="approved") or "approved"
        password = request.form.get("password", "")
        selected_pages = request.form.getlist("page_permissions")
        selected_actions = request.form.getlist("action_permissions")
        editable_fields = get_user_editable_fields(current) if not current.is_dev else user_edit_field_key_set()

        if "responsible_name" not in editable_fields:
            responsible_name = target.responsible_name
        if "username" not in editable_fields:
            username = target.username
            email = target.email or synthetic_email_for_username(username)
        if "password" not in editable_fields:
            password = ""
        if "role" not in editable_fields:
            role = target.role
        if "status" not in editable_fields:
            status = target.status
        if "page_permissions" not in editable_fields:
            selected_pages = list(get_user_page_permissions(target))
            selected_actions = list(get_user_action_permissions(target))

        # Segurança: o usuário logado não pode remover o próprio acesso administrativo
        # nem bloquear a própria conta sem querer. Se ele estiver criando o primeiro Dev
        # com a senha correta, mantém o role calculado como dev.
        if target.id == current.id:
            if role != "dev":
                role = current.role
            status = "approved"
            selected_pages = list(default_page_keys_for_role(role))
            selected_actions = list(default_static_action_permissions(role))

        organization_value = request.form.get("organization_name", "")
        franchise_name_value = request.form.get("franchise_name", "")
        franchise_number_value = request.form.get("franchise_number", "")
        cnpj_value = request.form.get("cnpj", "")
        if role == "franchise" and not str(franchise_name_value or "").strip():
            franchise_name_value = target.franchise_name or target.organization_name
        if "organization_name" not in editable_fields:
            organization_value = target.organization_name
        if "franchise_name" not in editable_fields:
            franchise_name_value = target.franchise_name
        if "franchise_number" not in editable_fields:
            franchise_number_value = target.franchise_number
        if "cnpj" not in editable_fields:
            cnpj_value = target.cnpj

        try:
            organization_name, franchise_name, franchise_number, cnpj = validate_user_profile_fields(
                role,
                organization_name=organization_value,
                franchise_name=franchise_name_value,
                franchise_number=franchise_number_value,
                cnpj=cnpj_value,
                strict_base=not current.is_dev,
            )
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_template("admin/user_form.html", user=target, is_new=False, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(current, target), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current, target), can_edit_dev_password=can_edit_dev_password(current))

        selected_pages = list(normalize_user_page_selection_for_editor(current, role, selected_pages))
        if not selected_pages:
            selected_pages = list(default_page_keys_for_role(role))
        selected_actions = list(normalize_user_action_selection_for_editor(current, role, selected_actions, selected_pages))

        if not responsible_name or not username:
            flash("Preencha responsável e nome de usuário.", "danger")
            return render_template("admin/user_form.html", user=target, is_new=False, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(current, target), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current, target), can_edit_dev_password=can_edit_dev_password(current))
        if not valid_username(username):
            flash("Use um nome de usuário com 3 a 40 caracteres: letras, números, ponto, hífen ou underline.", "danger")
            return render_template("admin/user_form.html", user=target, is_new=False, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(current, target), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current, target), can_edit_dev_password=can_edit_dev_password(current))

        with db_connect() as conn:
            existing = conn.execute(
                "SELECT id FROM users WHERE lower(username) = lower(?) AND id <> ?",
                (username, user_id),
            ).fetchone()
            if existing is not None:
                flash("Já existe outro usuário cadastrado com este nome de usuário.", "danger")
                return render_template("admin/user_form.html", user=target, is_new=False, permission_options=PAGE_PERMISSION_OPTIONS, selected_permissions=set(selected_pages), allowed_role_options=allowed_role_options_for_editor(current, target), role_labels=role_option_labels(), role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)], role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None), can_edit_admin_role=can_change_admin_role(current, target), can_edit_dev_password=can_edit_dev_password(current))

            if password:
                conn.execute(
                    """
                    UPDATE users
                       SET responsible_name = ?, organization_name = ?, franchise_name = ?, franchise_number = ?, cnpj = ?,
                           username = ?, email = ?, password_hash = ?, role = ?, status = ?, updated_at = ?
                     WHERE id = ?
                    """,
                    (
                        responsible_name,
                        organization_name,
                        franchise_name,
                        franchise_number,
                        cnpj,
                        username,
                        email,
                        generate_password_hash(password),
                        role,
                        status,
                        now_iso(),
                        user_id,
                    ),
                )
            else:
                conn.execute(
                    """
                    UPDATE users
                       SET responsible_name = ?, organization_name = ?, franchise_name = ?, franchise_number = ?, cnpj = ?,
                           username = ?, email = ?, role = ?, status = ?, updated_at = ?
                     WHERE id = ?
                    """,
                    (
                        responsible_name,
                        organization_name,
                        franchise_name,
                        franchise_number,
                        cnpj,
                        username,
                        email,
                        role,
                        status,
                        now_iso(),
                        user_id,
                    ),
                )
            save_user_page_permissions(conn, user_id, role, selected_pages)
            save_user_action_permissions(conn, user_id, role, selected_actions, selected_pages)
            apply_user_request_block_form(conn, user_id, current, request.form)
            conn.commit()

        flash("Acesso do usuário atualizado.", "success")
        return redirect_to_return("admin_users")

    return render_template(
        "admin/user_form.html",
        user=target,
        is_new=False,
        permission_options=PAGE_PERMISSION_OPTIONS,
        selected_permissions=selected_permissions_for_form(target),
        allowed_role_options=allowed_role_options_for_editor(require_current_user(), target),
        role_labels=role_option_labels(),
        role_admin_keys=[key for key in all_role_options() if role_is_admin_like(key)],
        role_permissions_map=role_permissions_map(allowed_role_options_for_editor(current) if 'current' in locals() else None),
        can_edit_admin_role=can_change_admin_role(require_current_user(), target),
        can_edit_dev_password=can_edit_dev_password(require_current_user()),
        can_manage_product_blocks=can_manage_product_request_blocks(require_current_user()),
        user_request_blocks=list_product_request_blocks_for_user(target.id) if can_manage_product_request_blocks(require_current_user()) else [],
        block_product_options=list_products_for_request_block_options() if can_manage_product_request_blocks(require_current_user()) else [],
    )


@app.post("/admin/users/<int:user_id>/status")
@admin_required
@page_access_required("admin_users")
def admin_user_status(user_id: int):
    denied = require_action_permission("users_status_delete", "Seu tipo de acesso não pode alterar status de usuários.", "admin_users")
    if denied:
        return denied
    target = get_user(user_id)
    if target is None:
        abort(404)
    action = request.form.get("action")
    current = require_current_user()
    if not can_edit_user_target(current, target):
        flash("Seu tipo de acesso nao pode alterar esse usuario.", "warning")
        return redirect_to_return("admin_users")
    if target.role == "dev" and target.id != current.id and not current.is_dev:
        flash("Somente usuário Dev pode alterar status de administradores.", "warning")
        return redirect_to_return("admin_users")
    if target.is_admin and target.id == current.id and action != "approved":
        flash("Você não pode bloquear seu próprio usuário admin.", "warning")
        return redirect_to_return("admin_users")
    if action not in ["approved", "rejected", "pending"]:
        abort(400)
    with db_connect() as conn:
        conn.execute("UPDATE users SET status = ?, updated_at = ? WHERE id = ?", (action, now_iso(), user_id))
        conn.commit()
    flash("Status do usuário atualizado.", "success")
    return redirect_to_return("admin_users")


@app.post("/admin/users/<int:user_id>/delete")
@admin_required
@page_access_required("admin_users")
def admin_user_delete(user_id: int):
    denied = require_action_permission("users_status_delete", "Seu tipo de acesso não pode excluir usuários.", "admin_users")
    if denied:
        return denied
    target = get_user(user_id)
    if target is None:
        abort(404)
    current = require_current_user()
    if not can_edit_user_target(current, target):
        flash("Seu tipo de acesso nao pode excluir esse usuario.", "warning")
        return redirect_to_return("admin_users")
    if target.is_admin and target.id == current.id:
        flash("Você não pode excluir seu próprio usuário admin.", "warning")
        return redirect_to_return("admin_users")
    if target.role == "dev" and not current.is_dev:
        flash("Somente usuário Dev pode excluir administradores.", "warning")
        return redirect_to_return("admin_users")

    try:
        with db_connect() as conn:
            if target.is_admin:
                approved_admins = conn.execute("SELECT COUNT(*) AS total FROM users WHERE role IN ('admin', 'dev') AND status = 'approved'").fetchone()["total"]
                if approved_admins <= 1 and target.status == "approved":
                    flash("Não é possível excluir o último administrador aprovado.", "warning")
                    return redirect_to_return("admin_users")
            removed_requests, _ = permanently_delete_user(conn, user_id)
            conn.commit()
        if removed_requests:
            flash(f"Usuário excluído definitivamente. {removed_requests} solicitação(ões) vinculada(s) também foram removidas do banco.", "success")
        else:
            flash("Usuário excluído definitivamente do banco de dados.", "success")
    except Exception as exc:
        app.logger.exception("Falha ao excluir usuário definitivamente")
        flash(f"Não consegui excluir o usuário do banco. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_users")


def normalize_access_role_permissions_from_form(form: Any) -> list[str]:
    selected = [str(key or "").strip() for key in form.getlist("page_permissions")]
    valid = page_permission_key_set()
    permissions = [key for key in selected if key in valid]
    if not permissions:
        permissions = ["home", "my_requests"]
    return list(dict.fromkeys(permissions))


def normalize_access_role_action_permissions_from_form(form: Any, page_permissions: list[str] | set[str]) -> list[str]:
    selected = [str(key or "").strip() for key in form.getlist("action_permissions")]
    valid = {item["key"] for item in action_permissions_for_pages(page_permissions)}
    return list(dict.fromkeys([key for key in selected if key in valid]))


def normalize_access_role_editable_roles_from_form(form: Any) -> list[str]:
    selected = [str(key or "").strip().lower() for key in form.getlist("editable_roles")]
    valid = set(all_role_options())
    return list(dict.fromkeys([key for key in selected if key in valid and key != "dev"]))


def normalize_access_role_editable_fields_from_form(form: Any) -> list[str]:
    selected = [str(key or "").strip() for key in form.getlist("editable_user_fields")]
    valid = user_edit_field_key_set()
    return list(dict.fromkeys([key for key in selected if key in valid]))


def grouped_action_permissions(page_permissions: list[str] | set[str] | None = None) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    for page in PAGE_PERMISSION_OPTIONS:
        actions = [item for item in ACTION_PERMISSION_OPTIONS if item["page_key"] == page["key"]]
        if actions:
            groups.append({"page_key": page["key"], "page_label": page["label"], "actions": actions})
    return groups


def action_permission_label_map() -> dict[str, str]:
    return {item["key"]: item["label"] for item in ACTION_PERMISSION_OPTIONS}


def access_role_listing() -> list[dict[str, Any]]:
    page_labels = {item["key"]: item["label"] for item in PAGE_PERMISSION_OPTIONS}
    action_labels = action_permission_label_map()
    role_labels = role_option_labels()
    field_labels = {item["key"]: item["label"] for item in USER_EDIT_FIELD_OPTIONS}
    default_descriptions = {
        "base": "Tipo padrão para bases. Permissões administrativas não liberadas.",
        "franchise": "Tipo padrão para franquias. Permissões administrativas não liberadas.",
        "admin": "Tipo padrão administrativo com acesso completo às áreas administrativas.",
        "dev": "Tipo Dev com controle total do portal, tipos de acesso e permissões.",
    }
    roles: list[dict[str, Any]] = []
    for key in STATIC_ROLE_ORDER:
        role_definition = get_access_role_definition(key) or default_static_access_role(key)
        if role_definition is None:
            continue
        permissions = sorted(role_definition.permissions, key=lambda item: list(page_labels).index(item) if item in page_labels else 999)
        actions = sorted(role_definition.action_permissions, key=lambda item: list(action_labels).index(item) if item in action_labels else 999)
        roles.append({
            "role_key": key,
            "display_key": display_access_role_key(key),
            "name": role_definition.name,
            "description": default_descriptions.get(key, "Tipo padrão do sistema."),
            "permissions": permissions,
            "action_permissions": actions,
            "editable_roles": role_definition.editable_roles,
            "editable_role_labels": [role_labels.get(item, display_access_role_key(item)) for item in role_definition.editable_roles if item != "dev"],
            "editable_user_fields": role_definition.editable_user_fields,
            "editable_field_labels": [field_labels.get(item, item) for item in role_definition.editable_user_fields],
            "is_custom": False,
            "can_edit": key != "dev",
            "can_delete": False,
        })
    for role in list_custom_access_roles():
        roles.append({
            "role_key": role.role_key,
            "display_key": display_access_role_key(role.role_key),
            "name": role.name,
            "description": role.description or "-",
            "permissions": role.permissions,
            "action_permissions": role.action_permissions,
            "editable_roles": role.editable_roles,
            "editable_role_labels": [role_labels.get(item, display_access_role_key(item)) for item in role.editable_roles if item != "dev"],
            "editable_user_fields": role.editable_user_fields,
            "editable_field_labels": [field_labels.get(item, item) for item in role.editable_user_fields],
            "is_custom": True,
            "can_edit": True,
            "can_delete": True,
        })
    return roles


def access_role_form_context(
    role: AccessRoleType | None = None,
    permissions: list[str] | set[str] | None = None,
    action_permissions: list[str] | set[str] | None = None,
    editable_roles: list[str] | set[str] | None = None,
    editable_user_fields: list[str] | set[str] | None = None,
) -> dict[str, Any]:
    selected = set(permissions if permissions is not None else (role.permissions if role else ["home", "my_requests"]))
    selected_actions = set(action_permissions if action_permissions is not None else (role.action_permissions if role else []))
    selected_editable_roles = set(editable_roles if editable_roles is not None else (role.editable_roles if role else []))
    selected_editable_fields = set(editable_user_fields if editable_user_fields is not None else (role.editable_user_fields if role else []))
    return {
        "access_role": role,
        "permission_options": PAGE_PERMISSION_OPTIONS,
        "selected_permissions": selected,
        "action_permission_groups": grouped_action_permissions(selected),
        "selected_action_permissions": selected_actions,
        "user_edit_role_options": [{"key": key, "label": role_option_labels().get(key, display_access_role_key(key))} for key in all_role_options() if key != "dev"],
        "selected_editable_roles": selected_editable_roles,
        "user_edit_field_options": USER_EDIT_FIELD_OPTIONS,
        "selected_editable_user_fields": selected_editable_fields,
        "action_permission_label_map": action_permission_label_map(),
        "custom_roles": list_custom_access_roles(),
        "access_roles": access_role_listing(),
    }


@app.route("/admin/access-types", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_users")
def admin_access_types():
    current = require_current_user()
    if not current.is_dev:
        abort(403)
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()[:80]
        description = (request.form.get("description") or "").strip()[:240]
        permissions = normalize_access_role_permissions_from_form(request.form)
        action_permissions = normalize_access_role_action_permissions_from_form(request.form, permissions)
        editable_roles = normalize_access_role_editable_roles_from_form(request.form)
        editable_user_fields = normalize_access_role_editable_fields_from_form(request.form)
        if not name:
            flash("Informe o nome do tipo de acesso.", "warning")
            return render_template("admin/access_types.html", **access_role_form_context(None, permissions, action_permissions, editable_roles, editable_user_fields))
        role_key = custom_access_role_key(name)
        with db_connect() as conn:
            original_key = role_key
            suffix = 2
            while conn.execute("SELECT 1 FROM access_role_types WHERE role_key = ?", (role_key,)).fetchone() is not None:
                role_key = f"{original_key}_{suffix}"[:80].rstrip("_")
                suffix += 1
            conn.execute(
                """
                INSERT INTO access_role_types (role_key, name, description, permissions_json, action_permissions_json, editable_roles_json, editable_user_fields_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    role_key,
                    name,
                    description,
                    json.dumps(permissions, ensure_ascii=False),
                    json.dumps(action_permissions, ensure_ascii=False),
                    json.dumps(editable_roles, ensure_ascii=False),
                    json.dumps(editable_user_fields, ensure_ascii=False),
                    now_iso(),
                    now_iso(),
                ),
            )
            conn.commit()
        get_access_role_override.cache_clear()
        get_custom_access_role.cache_clear()
        flash("Tipo de acesso criado com sucesso.", "success")
        return redirect_to_return("admin_access_types")
    return render_template("admin/access_types.html", **access_role_form_context())


@app.route("/admin/access-types/<role_key>/edit", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_users")
def admin_access_type_edit(role_key: str):
    current = require_current_user()
    if not current.is_dev:
        abort(403)
    role = get_access_role_definition(role_key)
    if role is None or role.role_key == "dev":
        abort(404)
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()[:80]
        description = (request.form.get("description") or "").strip()[:240]
        updated_role_key = role.role_key
        if role.is_static:
            name = STATIC_ROLE_LABELS.get(role.role_key, role.name)
            description = description or STATIC_ROLE_DESCRIPTIONS.get(role.role_key, role.description)
        else:
            updated_role_key = editable_custom_access_role_key(request.form.get("role_key"), role.role_key)
        permissions = normalize_access_role_permissions_from_form(request.form)
        action_permissions = normalize_access_role_action_permissions_from_form(request.form, permissions)
        editable_roles = normalize_access_role_editable_roles_from_form(request.form)
        editable_user_fields = normalize_access_role_editable_fields_from_form(request.form)
        if not name:
            flash("Informe o nome do tipo de acesso.", "warning")
            return render_template("admin/access_type_form.html", **access_role_form_context(role, permissions, action_permissions, editable_roles, editable_user_fields))
        with db_connect() as conn:
            changed_at = now_iso()
            if updated_role_key != role.role_key:
                existing = conn.execute("SELECT 1 FROM access_role_types WHERE role_key = ?", (updated_role_key,)).fetchone()
                if existing is not None or updated_role_key in STATIC_ROLE_KEYS:
                    flash("Já existe outro tipo de acesso com esse identificador.", "warning")
                    return render_template("admin/access_type_form.html", **access_role_form_context(role, permissions, action_permissions, editable_roles, editable_user_fields))
                conn.execute(
                    """
                    UPDATE access_role_types
                       SET role_key = ?, name = ?, description = ?, permissions_json = ?,
                           action_permissions_json = ?, editable_roles_json = ?,
                           editable_user_fields_json = ?, updated_at = ?
                     WHERE role_key = ?
                    """,
                    (
                        updated_role_key,
                        name,
                        description,
                        json.dumps(permissions, ensure_ascii=False),
                        json.dumps(action_permissions, ensure_ascii=False),
                        json.dumps(editable_roles, ensure_ascii=False),
                        json.dumps(editable_user_fields, ensure_ascii=False),
                        changed_at,
                        role.role_key,
                    ),
                )
                conn.execute("UPDATE users SET role = ?, page_permissions_configured = 0, action_permissions_configured = 0, updated_at = ? WHERE role = ?", (updated_role_key, changed_at, role.role_key))
            else:
                conn.execute(
                    """
                    INSERT INTO access_role_types (role_key, name, description, permissions_json, action_permissions_json, editable_roles_json, editable_user_fields_json, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(role_key) DO UPDATE SET
                        name = excluded.name,
                        description = excluded.description,
                        permissions_json = excluded.permissions_json,
                        action_permissions_json = excluded.action_permissions_json,
                        editable_roles_json = excluded.editable_roles_json,
                        editable_user_fields_json = excluded.editable_user_fields_json,
                        updated_at = excluded.updated_at
                    """,
                    (
                        role.role_key,
                        name,
                        description,
                        json.dumps(permissions, ensure_ascii=False),
                        json.dumps(action_permissions, ensure_ascii=False),
                        json.dumps(editable_roles, ensure_ascii=False),
                        json.dumps(editable_user_fields, ensure_ascii=False),
                        changed_at,
                        changed_at,
                    ),
                )
            # Usuários desse tipo passam a herdar as permissões novas do tipo.
            conn.execute("UPDATE users SET page_permissions_configured = 0, action_permissions_configured = 0, updated_at = ? WHERE role = ?", (changed_at, updated_role_key))
            conn.commit()
        get_access_role_override.cache_clear()
        get_custom_access_role.cache_clear()
        flash("Tipo de acesso atualizado.", "success")
        return redirect_to_return("admin_access_types")
    return render_template("admin/access_type_form.html", **access_role_form_context(role))


@app.post("/admin/access-types/<role_key>/bulk-password")
@admin_required
@page_access_required("admin_users")
def admin_access_type_bulk_password(role_key: str):
    current = require_current_user()
    if not current.is_dev:
        abort(403)
    role = get_access_role_definition(role_key)
    if role is None or role.role_key == "dev":
        abort(404)
    new_password = (request.form.get("bulk_password") or "").strip()
    if not new_password:
        flash("Informe a nova senha para aplicar aos usuários desse tipo de acesso.", "warning")
        return redirect(url_for("admin_access_type_edit", role_key=role.role_key, return_to=request.form.get("return_to", "")))
    with db_connect() as conn:
        changed_at = now_iso()
        affected = int(conn.execute("SELECT COUNT(*) AS total FROM users WHERE role = ?", (role.role_key,)).fetchone()["total"] or 0)
        conn.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE role = ?",
            (generate_password_hash(new_password), changed_at, role.role_key),
        )
        conn.commit()
    flash(f"Senha aplicada para {affected} usuário(s) do tipo {role.name}.", "success")
    return redirect(url_for("admin_access_type_edit", role_key=role.role_key, return_to=request.form.get("return_to", "")))


@app.post("/admin/access-types/<role_key>/delete")
@admin_required
@page_access_required("admin_users")
def admin_access_type_delete(role_key: str):
    current = require_current_user()
    if not current.is_dev:
        abort(403)
    role = get_custom_access_role(role_key)
    if role is None:
        abort(404)
    with db_connect() as conn:
        in_use = conn.execute("SELECT COUNT(*) AS total FROM users WHERE role = ?", (role.role_key,)).fetchone()["total"]
        affected = int(in_use or 0)
        if affected > 0:
            # Ao excluir um tipo personalizado, nenhum usuário fica apontando para um cargo inexistente.
            # Eles voltam para Base e as permissões herdadas desse tipo são limpas.
            changed_at = now_iso()
            conn.execute(
                "DELETE FROM user_page_permissions WHERE user_id IN (SELECT id FROM users WHERE role = ?)",
                (role.role_key,),
            )
            conn.execute(
                "UPDATE users SET role = 'base', organization_name = COALESCE(NULLIF(organization_name, ''), ?), page_permissions_configured = 0, action_permissions_configured = 0, updated_at = ? WHERE role = ?",
                (BASE_UNIT_OPTIONS[0] if BASE_UNIT_OPTIONS else ADMIN_ORGANIZATION_NAME, changed_at, role.role_key),
            )
        conn.execute("DELETE FROM access_role_types WHERE role_key = ?", (role.role_key,))
        conn.commit()
    get_access_role_override.cache_clear()
    get_custom_access_role.cache_clear()
    flash("Tipo de acesso excluído do banco de dados." + (f" {affected} usuário(s) foram movidos para Base." if affected else ""), "success")
    return redirect_to_return("admin_access_types")


@app.route("/admin/products")
@admin_required
@page_access_required("admin_products")
def admin_products():
    search = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "all").strip().lower()
    sort_filter = (request.args.get("sort") or "default").strip().lower()
    category_filter = (request.args.get("category") or "").strip()
    stock_tag_filter_raw = (request.args.get("stock_tag") or "").strip()
    stock_tag_filter = normalize_stock_tag_slug(stock_tag_filter_raw, "") if stock_tag_filter_raw else ""
    if status_filter not in {"all", "active", "inactive"}:
        status_filter = "all"
    if sort_filter not in {"default", "category", "category_desc", "name", "name_desc", "value_asc", "value_desc", "stock_asc", "stock_desc"}:
        sort_filter = "default"

    per_page = list_page_limit(default=DEFAULT_TABLE_PAGE_SIZE, maximum=500)
    page = bounded_int(request.args.get("page"), 1, 1, 100000)

    clauses = ["catalog_archived = 0"]
    params: list[Any] = []
    if status_filter == "active":
        clauses.append("active = 1")
    elif status_filter == "inactive":
        clauses.append("active = 0")
    if category_filter:
        clauses.append("LOWER(TRIM(COALESCE(category, ''))) = LOWER(TRIM(?))")
        params.append(category_filter)
    if stock_tag_filter:
        clauses.append("stock_tag = ?")
        params.append(stock_tag_filter)
    if search:
        like = like_term(search)
        clauses.append("(name LIKE ? OR category LIKE ? OR description LIKE ? OR unit_measure LIKE ? OR stock_tag LIKE ?)")
        params.extend([like, like, like, like, like])

    sort_map = {
        "default": "active DESC, category COLLATE NOCASE ASC, name COLLATE NOCASE ASC",
        "name": "name COLLATE NOCASE ASC, id DESC",
        "name_desc": "name COLLATE NOCASE DESC, id DESC",
        "category": "category COLLATE NOCASE ASC, name COLLATE NOCASE ASC",
        "category_desc": "category COLLATE NOCASE DESC, name COLLATE NOCASE ASC",
        "value_asc": "price_cents ASC, name COLLATE NOCASE ASC",
        "value_desc": "price_cents DESC, name COLLATE NOCASE ASC",
        "stock_asc": "stock_quantity ASC, name COLLATE NOCASE ASC",
        "stock_desc": "stock_quantity DESC, name COLLATE NOCASE ASC",
    }
    where_sql = " WHERE " + " AND ".join(clauses)

    with db_connect() as conn:
        total_row = conn.execute(f"SELECT COUNT(*) AS total FROM products{where_sql}", params).fetchone()
        total_products = int((total_row["total"] if total_row else 0) or 0)
        total_pages = max(1, (total_products + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page
        rows = conn.execute(
            f"""
            SELECT id, name, category, category_emoji, image_name, image_key, image_content_type,
                   unit_measure, is_kit, kit_quantity, description, stock_quantity, price_cents,
                   limit_base, limit_franchise, limit_block_days, min_order_quantity, min_stock, max_stock,
                   active, visible_base, visible_franchise, internal, stock_tag,
                   (SELECT name FROM stock_tags WHERE slug = products.stock_tag) AS stock_tag_name,
                   catalog_archived, created_at, updated_at
              FROM products
              {where_sql}
             ORDER BY {sort_map.get(sort_filter, sort_map["default"])}
             LIMIT ? OFFSET ?
            """,
            [*params, per_page, offset],
        ).fetchall()

        if exact_counts_enabled():
            active_products = int(conn.execute(f"SELECT COUNT(*) AS total FROM products{where_sql} AND active = 1", params).fetchone()["total"] or 0)
            inactive_products = int(conn.execute(f"SELECT COUNT(*) AS total FROM products{where_sql} AND active = 0", params).fetchone()["total"] or 0)
        else:
            active_products = sum(1 for row in rows if int(row["active"] or 0) == 1)
            inactive_products = sum(1 for row in rows if int(row["active"] or 0) == 0)

    products = [product for row in rows if (product := row_to_product(row)) is not None]
    category_items = list_product_categories()
    categories = [item["name"] for item in category_items]

    page_size_options = list(TABLE_PAGE_SIZE_OPTIONS)
    if per_page not in page_size_options:
        page_size_options.append(per_page)
        page_size_options.sort()

    def page_url(target_page: int, target_limit: int | None = None) -> str:
        args: dict[str, Any] = {"page": max(1, target_page), "limit": target_limit or per_page}
        if search:
            args["q"] = search
        if status_filter and status_filter != "all":
            args["status"] = status_filter
        if sort_filter:
            args["sort"] = sort_filter
        if category_filter:
            args["category"] = category_filter
        if stock_tag_filter:
            args["stock_tag"] = stock_tag_filter
        return url_for("admin_products", **args)

    visible_page_numbers: set[int] = {1, total_pages}
    for number in range(page - 2, page + 3):
        if 1 <= number <= total_pages:
            visible_page_numbers.add(number)
    page_links: list[dict[str, Any]] = []
    previous_number = 0
    for number in sorted(visible_page_numbers):
        if previous_number and number - previous_number > 1:
            page_links.append({"ellipsis": True})
        page_links.append({"number": number, "active": number == page, "url": page_url(number)})
        previous_number = number

    start_item = offset + 1 if total_products else 0
    end_item = min(offset + len(products), total_products)
    product_pagination = {
        "page": page,
        "limit": per_page,
        "total": total_products,
        "total_pages": total_pages,
        "start": start_item,
        "end": end_item,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "first_url": page_url(1),
        "prev_url": page_url(max(1, page - 1)),
        "next_url": page_url(min(total_pages, page + 1)),
        "last_url": page_url(total_pages),
        "page_links": page_links,
        "page_size_options": page_size_options,
    }

    return render_template(
        "admin/products.html",
        products=products,
        product_categories_filter=categories,
        product_categories_manage=category_items,
        product_stock_tags=list_stock_tags(active_only=False),
        product_filters={"q": search, "status": status_filter, "sort": sort_filter, "category": category_filter, "stock_tag": stock_tag_filter, "limit": per_page, "page": page},
        product_counts={"total": total_products, "active": active_products, "inactive": inactive_products, "shown": len(products), "page": page, "limit": per_page, "low_read": low_row_read_mode()},
        product_pagination=product_pagination,
    )








@app.post("/admin/products/categories/update")
@admin_required
@page_access_required("admin_products")
def admin_product_categories_update():
    denied = require_action_permission("products_edit_category", "Seu tipo de acesso não pode editar categorias.", "admin_products")
    if denied:
        return denied
    old_names = request.form.getlist("category_old")
    new_names = request.form.getlist("category_name")
    emojis = request.form.getlist("category_emoji")
    updated = 0

    with db_connect() as conn:
        for old_name, new_name, emoji_value in zip(old_names, new_names, emojis):
            old_clean = str(old_name or "").strip()
            new_clean = str(new_name or "").strip()[:120]
            if not old_clean:
                continue
            if not new_clean:
                conn.rollback()
                flash("Nenhuma categoria pode ficar sem nome.", "warning")
                next_url = (request.form.get("next") or request.referrer or "").strip()
                return redirect(safe_local_redirect_target(request.form.get("return_to") or next_url, "admin_products"))
            emoji_clean = clean_category_emoji(emoji_value, new_clean)
            conn.execute(
                """
                UPDATE products
                   SET category = ?, category_emoji = ?, updated_at = ?
                 WHERE LOWER(TRIM(category)) = LOWER(TRIM(?))
                """,
                (new_clean, emoji_clean, now_iso(), old_clean),
            )
            updated += 1
        conn.commit()

    flash(f"Categorias atualizadas: {updated}.", "success")
    next_url = (request.form.get("next") or request.referrer or "").strip()
    return redirect(safe_local_redirect_target(request.form.get("return_to") or next_url, "admin_products"))


@app.post("/admin/products/stock-tags/update")
@admin_required
@page_access_required("admin_products")
def admin_product_stock_tags_update():
    if not can_manage_stock_tags():
        flash("Somente Admin e Dev podem criar ou editar tags de estoque.", "warning")
        return redirect_to_return("admin_products")

    slugs = request.form.getlist("tag_slug")
    names = request.form.getlist("tag_name")
    descriptions = request.form.getlist("tag_description")
    active_slugs = {normalize_stock_tag_slug(value, "") for value in request.form.getlist("tag_active")}
    updated = 0
    created = 0

    with db_connect() as conn:
        existing_rows = conn.execute("SELECT slug, system_key FROM stock_tags").fetchall()
        system_keys = {normalize_stock_tag_slug(row["slug"]) for row in existing_rows if bool(row["system_key"])}
        existing_slugs = {normalize_stock_tag_slug(row["slug"]) for row in existing_rows}

        for index, raw_slug in enumerate(slugs):
            slug = normalize_stock_tag_slug(raw_slug, "")
            if not slug:
                continue
            name = (names[index] if index < len(names) else "").strip()[:80]
            description = (descriptions[index] if index < len(descriptions) else "").strip()[:240]
            if not name:
                conn.rollback()
                flash("Nenhuma tag pode ficar sem nome.", "warning")
                return redirect_to_return("admin_products")
            active = 1 if slug in system_keys or slug in active_slugs else 0
            conn.execute(
                """
                UPDATE stock_tags
                   SET name = ?, description = ?, active = ?, updated_at = ?
                 WHERE slug = ?
                """,
                (name, description, active, now_iso(), slug),
            )
            updated += 1

        new_name = (request.form.get("new_tag_name") or "").strip()[:80]
        new_description = (request.form.get("new_tag_description") or "").strip()[:240]
        if new_name:
            new_slug = normalize_stock_tag_slug(new_name, "")
            if not new_slug:
                conn.rollback()
                flash("Informe um nome valido para a nova tag.", "warning")
                return redirect_to_return("admin_products")
            if new_slug in existing_slugs:
                conn.rollback()
                flash("Ja existe uma tag com esse nome.", "warning")
                return redirect_to_return("admin_products")
            conn.execute(
                """
                INSERT INTO stock_tags (slug, name, description, active, system_key, created_at)
                VALUES (?, ?, ?, 1, 0, ?)
                """,
                (new_slug, new_name, new_description, now_iso()),
            )
            created += 1

        conn.commit()

    flash(f"Tags de estoque salvas: {updated} atualizada(s), {created} criada(s).", "success")
    return redirect_to_return("admin_products")

@app.get("/admin/products/export")
@admin_required
@page_access_required("admin_products")
def admin_products_export():
    denied = require_action_permission("products_export", "Seu tipo de acesso não pode exportar produtos.", "admin_products")
    if denied:
        return denied
    export_language = (request.args.get("lang") or "pt").strip().lower()
    if export_language in {"zh", "zh-cn", "zh-hans", "zh-tw", "mandarin", "mandarim", "chinese", "simplified"}:
        export_language = "zh"
    else:
        export_language = "pt"

    search = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "all").strip().lower()
    sort_filter = (request.args.get("sort") or "default").strip().lower()
    category_filter = (request.args.get("category") or "").strip()
    stock_tag_filter_raw = (request.args.get("stock_tag") or "").strip()
    stock_tag_filter = normalize_stock_tag_slug(stock_tag_filter_raw, "") if stock_tag_filter_raw else ""
    if status_filter not in {"all", "active", "inactive"}:
        status_filter = "all"
    if sort_filter not in {"default", "category", "category_desc", "name", "name_desc", "value_asc", "value_desc", "stock_asc", "stock_desc"}:
        sort_filter = "default"

    clauses = ["catalog_archived = 0"]
    params: list[Any] = []
    if status_filter == "active":
        clauses.append("active = 1")
    elif status_filter == "inactive":
        clauses.append("active = 0")
    if category_filter:
        clauses.append("LOWER(TRIM(COALESCE(category, ''))) = LOWER(TRIM(?))")
        params.append(category_filter)
    if stock_tag_filter:
        clauses.append("stock_tag = ?")
        params.append(stock_tag_filter)
    if search:
        like = like_term(search)
        clauses.append("(name LIKE ? OR category LIKE ? OR description LIKE ? OR unit_measure LIKE ? OR stock_tag LIKE ?)")
        params.extend([like, like, like, like, like])

    sort_map = {
        "default": "active DESC, category COLLATE NOCASE ASC, name COLLATE NOCASE ASC",
        "name": "name COLLATE NOCASE ASC, id DESC",
        "name_desc": "name COLLATE NOCASE DESC, id DESC",
        "category": "category COLLATE NOCASE ASC, name COLLATE NOCASE ASC",
        "category_desc": "category COLLATE NOCASE DESC, name COLLATE NOCASE ASC",
        "value_asc": "price_cents ASC, name COLLATE NOCASE ASC",
        "value_desc": "price_cents DESC, name COLLATE NOCASE ASC",
        "stock_asc": "stock_quantity ASC, name COLLATE NOCASE ASC",
        "stock_desc": "stock_quantity DESC, name COLLATE NOCASE ASC",
    }
    where_sql = " WHERE " + " AND ".join(clauses)

    has_page_args = "limit" in request.args or "page" in request.args
    limit_sql = ""
    query_params: list[Any] = list(params)
    if has_page_args:
        export_limit = bounded_int(request.args.get("limit"), int(os.getenv("D1_EXPORT_LIMIT", "500")), 25, 2000)
        export_page = bounded_int(request.args.get("page"), 1, 1, 100000)
        export_offset = (export_page - 1) * export_limit
        limit_sql = " LIMIT ? OFFSET ?"
        query_params.extend([export_limit, export_offset])

    with db_connect() as conn:
        rows = conn.execute(
            f"""
            SELECT products.*,
                   (SELECT name FROM stock_tags WHERE slug = products.stock_tag) AS stock_tag_name
              FROM products
              {where_sql}
             ORDER BY {sort_map.get(sort_filter, sort_map['default'])}{limit_sql}
            """,
            query_params,
        ).fetchall()
    products = [product for row in rows if (product := row_to_product(row)) is not None]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "产品" if export_language == "zh" else "Produtos"
    headers = PRODUCT_EXPORT_HEADERS_ZH if export_language == "zh" else PRODUCT_EXPORT_HEADERS_PT
    worksheet.append(headers)
    for product in products:
        worksheet.append(product_row_for_excel_language(product, export_language))

    header_fill = PatternFill("solid", fgColor="E60012")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if len(row) >= 8:
            row[7].number_format = 'R$ #,##0.00'
    widths = [10, 38, 24, 16, 24, 12, 18, 48, 18, 18, 22, 24, 24, 24, 18, 18, 20, 14]
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width
    worksheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    language_label = "chines_simplificado" if export_language == "zh" else "portugues"
    filename = f"produtos_jt_insumos_{language_label}_{sao_paulo_filename_timestamp()}.xlsx"
    store_generated_file(
        storage_key("exports", filename),
        buffer,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        {"type": "products_export", "mode": "current_page" if has_page_args else "all"},
    )
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)


@app.post("/admin/products/import")
@admin_required
@page_access_required("admin_products")
def admin_products_import():
    denied = require_action_permission("products_import", "Seu tipo de acesso não pode importar produtos.", "admin_products")
    if denied:
        return denied
    try:
        import_mode = (request.form.get("import_mode") or "merge").strip().lower()
        if import_mode not in {"merge", "replace"}:
            import_mode = "merge"
        uploaded = request.files.get("spreadsheet")
        if uploaded is None or not uploaded.filename:
            flash("Selecione uma planilha .xlsx para importar.", "warning")
            return redirect_to_return("admin_products")
        if not uploaded.filename.lower().endswith(".xlsx"):
            flash("Importe apenas arquivos .xlsx.", "warning")
            return redirect_to_return("admin_products")

        try:
            uploaded_bytes = uploaded.read()
        except Exception as exc:
            print(f"[IMPORTAÇÃO PRODUTOS] Falha ao ler upload: {exc}")
            flash("Não foi possível ler o arquivo enviado.", "danger")
            return redirect_to_return("admin_products")

        if not uploaded_bytes:
            flash("A planilha enviada está vazia.", "warning")
            return redirect_to_return("admin_products")

        # Salvar cópia no R2 é opcional e nunca pode derrubar ou atrasar planilhas grandes.
        if len(uploaded_bytes) <= int(os.getenv("IMPORT_BACKUP_MAX_BYTES", "5242880")):
            try:
                upload_bytes_to_r2(
                    storage_key("imports", sao_paulo_filename_timestamp() + "_" + safe_filename(uploaded.filename)),
                    uploaded_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    {"type": "products_import", "mode": import_mode},
                )
            except Exception as exc:
                print(f"[R2] Não foi possível salvar cópia da planilha importada: {exc}")

        try:
            created, updated, skipped, archived, row_errors = import_products_from_workbook_bytes(
                uploaded_bytes,
                import_mode=import_mode,
            )
        except Exception as exc:
            try:
                import traceback
                traceback.print_exc()
            except Exception:
                pass
            print(f"[IMPORTAÇÃO PRODUTOS] Erro geral tratado: {type(exc).__name__} - {exc}")
            flash("Não consegui importar essa planilha. O erro foi registrado nos logs do Render; envie o trecho vermelho se continuar acontecendo.", "danger")
            return redirect_to_return("admin_products")

        flash_import_errors(row_errors)
        if created or updated:
            mode_message = (
                f" Catálogo substituído; {archived} produto(s) anterior(es) removido(s) das telas."
                if import_mode == "replace"
                else " Os produtos foram comparados pelo nome, sem criar duplicatas."
            )
            flash(
                f"Importação concluída: {created} criado(s), {updated} atualizado(s), {skipped} ignorado(s).{mode_message}",
                "success",
            )
        elif skipped or row_errors:
            flash(f"Nenhum produto foi criado ou atualizado. {skipped} linha(s) ignorada(s).", "warning")
        else:
            flash("Nenhum produto válido foi encontrado na planilha. Confira se a primeira linha contém os cabeçalhos corretos.", "warning")
        return redirect_to_return("admin_products")
    except Exception as exc:
        # Última barreira para impedir Internal Server Error branco na tela.
        try:
            import traceback
            traceback.print_exc()
        except Exception:
            pass
        print(f"[IMPORTAÇÃO PRODUTOS] Falha inesperada capturada: {type(exc).__name__} - {exc}")
        flash("A importação falhou, mas o site não quebrou. Veja os logs do Render para o detalhe do erro.", "danger")
        return redirect_to_return("admin_products")


@app.route("/admin/products/new", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_products")
def admin_product_new():
    if request.method == "POST":
        denied = require_action_permission("products_create", "Seu tipo de acesso não pode criar produtos.", "admin_products")
        if denied:
            return denied

        product = fill_product_from_form(Product())
        if not product.name:
            flash("Informe o nome do produto.", "warning")
            return redirect_to_return("admin_product_new")
        uploaded_image = request.files.get("product_image")
        if uploaded_image is not None and uploaded_image.filename:
            try:
                product.image_name, product.image_key, product.image_content_type = save_product_image_upload(uploaded_image)
            except ValueError as exc:
                flash(str(exc), "warning")
                return redirect_to_return("admin_product_new")
        with db_connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO products (
                    name, category, category_emoji, image_name, image_key, image_content_type,
                    unit_measure, is_kit, kit_quantity, description, stock_quantity,
                    price_cents, limit_base, limit_franchise, limit_block_days, min_order_quantity,
                    min_stock, max_stock, active, visible_base, visible_franchise, internal, stock_tag, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product.name,
                    product.category,
                    product.category_emoji,
                    product.image_name,
                    product.image_key,
                    product.image_content_type,
                    product.unit_measure,
                    1 if product.is_kit else 0,
                    product.kit_quantity if product.is_kit else 1,
                    product.description,
                    product.stock_quantity,
                    product.price_cents,
                    product.limit_base,
                    product.limit_franchise,
                    product.limit_block_days,
                    product.min_order_quantity,
                    product.min_stock,
                    product.max_stock,
                    1 if product.active else 0,
                    1 if product.visible_base else 0,
                    1 if product.visible_franchise else 0,
                    1 if product.internal else 0,
                    product.stock_tag,
                    now_iso(),
                ),
            )
            new_id = get_cursor_lastrowid(cursor)
            if product.category:
                conn.execute(
                    "UPDATE products SET category_emoji = ? WHERE LOWER(TRIM(category)) = LOWER(TRIM(?))",
                    (product.category_emoji, product.category),
                )
            if product.stock_quantity > 0 and new_id is not None:
                record_stock_movement(conn, int(new_id), product.stock_quantity, 0, product.stock_quantity, "product_created", "Produto cadastrado manualmente.", created_by_id=require_current_user().id)
            conn.commit()
        flash("Produto cadastrado.", "success")
        return redirect_to_return("admin_products")
    if not user_has_action_access(current_user(), "products_create"):
        flash("Seu tipo de acesso não pode criar produtos.", "warning")
        return redirect_to_return("admin_products")
    return render_template(
        "admin/product_form.html",
        product=None,
        product_categories=list_product_categories(),
        product_categories_manage=list_product_categories(),
        stock_tags=list_stock_tags(active_only=True, include_slug=DEFAULT_STOCK_TAG),
    )


@app.route("/admin/products/<int:product_id>/edit", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_products")
def admin_product_edit(product_id: int):
    product = get_product(product_id)
    if product is None:
        abort(404)
    if not user_has_any_action_access(current_user(), PRODUCT_EDIT_ACTION_KEYS):
        flash("Seu tipo de acesso nÃ£o pode editar produtos.", "warning")
        return redirect_to_return("admin_products")
    if request.method == "POST":
        old_product = Product(**product.__dict__)
        old_stock_quantity = product.stock_quantity
        old_image_key = product.image_key
        fill_product_from_form(product)
        if not product.name:
            flash("Informe o nome do produto.", "warning")
            return redirect_to_return("admin_product_edit", product_id=product_id)
        uploaded_image = request.files.get("product_image")
        remove_image = request.form.get("remove_image") == "on"
        if uploaded_image is not None and uploaded_image.filename:
            try:
                product.image_name, product.image_key, product.image_content_type = save_product_image_upload(uploaded_image)
            except ValueError as exc:
                flash(str(exc), "warning")
                return redirect_to_return("admin_product_edit", product_id=product_id)
        elif remove_image:
            product.image_name = ""
            product.image_key = ""
            product.image_content_type = ""
        missing_actions = product_update_missing_action_permissions(old_product, product, bool(uploaded_image is not None and uploaded_image.filename), remove_image)
        if missing_actions:
            flash("Seu tipo de acesso não permite alterar: " + ", ".join(missing_actions) + ".", "warning")
            return redirect_to_return("admin_product_edit", product_id=product_id)
        with db_connect() as conn:
            conn.execute(
                """
                UPDATE products
                SET name = ?, category = ?, category_emoji = ?, image_name = ?, image_key = ?,
                    image_content_type = ?, unit_measure = ?, is_kit = ?, kit_quantity = ?, description = ?,
                    stock_quantity = ?, price_cents = ?, limit_base = ?, limit_franchise = ?,
                    limit_block_days = ?, min_order_quantity = ?, min_stock = ?, max_stock = ?, active = ?,
                    visible_base = ?, visible_franchise = ?, internal = ?, stock_tag = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    product.name,
                    product.category,
                    product.category_emoji,
                    product.image_name,
                    product.image_key,
                    product.image_content_type,
                    product.unit_measure,
                    1 if product.is_kit else 0,
                    product.kit_quantity if product.is_kit else 1,
                    product.description,
                    product.stock_quantity,
                    product.price_cents,
                    product.limit_base,
                    product.limit_franchise,
                    product.limit_block_days,
                    product.min_order_quantity,
                    product.min_stock,
                    product.max_stock,
                    1 if product.active else 0,
                    1 if product.visible_base else 0,
                    1 if product.visible_franchise else 0,
                    1 if product.internal else 0,
                    product.stock_tag,
                    now_iso(),
                    product_id,
                ),
            )
            if product.category:
                conn.execute(
                    "UPDATE products SET category_emoji = ? WHERE LOWER(TRIM(category)) = LOWER(TRIM(?))",
                    (product.category_emoji, product.category),
                )
            if old_stock_quantity != product.stock_quantity:
                record_stock_movement(conn, product_id, product.stock_quantity - old_stock_quantity, old_stock_quantity, product.stock_quantity, "manual_adjustment", "Estoque alterado na edição do produto.", created_by_id=require_current_user().id)
            conn.commit()
        if old_image_key and old_image_key != product.image_key:
            remove_local_product_image(old_image_key)
        flash("Produto atualizado.", "success")
        return redirect_to_return("admin_products")
    return render_template(
        "admin/product_form.html",
        product=product,
        product_categories=list_product_categories(),
        product_categories_manage=list_product_categories(),
        stock_tags=list_stock_tags(active_only=True, include_slug=product.stock_tag),
    )


@app.post("/admin/products/<int:product_id>/toggle-active")
@admin_required
@page_access_required("admin_products")
def admin_product_toggle_active(product_id: int):
    denied = require_action_permission("products_edit_visibility", "Seu tipo de acesso não pode ativar ou inativar produtos.", "admin_products")
    if denied:
        return denied
    product = get_product(product_id)
    if product is None:
        abort(404)

    new_active = 0 if product.active else 1
    try:
        with db_connect() as conn:
            conn.execute(
                "UPDATE products SET active = ?, updated_at = ? WHERE id = ?",
                (new_active, now_iso(), product_id),
            )
            conn.commit()
        if new_active:
            flash(f"Produto '{product.name}' ativado para solicitação.", "success")
        else:
            flash(f"Produto '{product.name}' inativado para solicitação. Ele continua salvo no banco de dados.", "success")
    except Exception as exc:
        app.logger.exception("Falha ao alterar status do produto")
        flash(f"Não consegui alterar o status do produto. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_products")


@app.post("/admin/products/<int:product_id>/delete")
@admin_required
@page_access_required("admin_products")
def admin_product_delete(product_id: int):
    denied = require_action_permission("products_delete", "Seu tipo de acesso não pode excluir produtos.", "admin_products")
    if denied:
        return denied
    try:
        with db_connect() as conn:
            product_name, removed_items, removed_requests = permanently_delete_product(conn, product_id)
            conn.commit()
        details: list[str] = []
        if removed_items:
            details.append(f"{removed_items} vínculo(s) em solicitações")
        if removed_requests:
            details.append(f"{removed_requests} solicitação(ões) vazia(s)")
        suffix = f" Removidos também: {', '.join(details)}." if details else ""
        flash(f"Produto '{product_name}' excluído definitivamente do banco de dados.{suffix}", "success")
    except LookupError:
        abort(404)
    except Exception as exc:
        app.logger.exception("Falha ao excluir produto definitivamente")
        flash(f"Não consegui excluir o produto do banco. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_products")


@app.route("/admin/requests")
@admin_required
@page_access_required("admin_requests")
def admin_requests():
    current = require_current_user()
    selected_status = (request.args.get("status", "") or "").strip().lower()
    if selected_status not in {"", "pending", "approved", "rejected", "deleted"}:
        selected_status = ""
    request_filters = normalize_request_filters_from_args()
    extra_args: dict[str, Any] = request_filters_to_query_args(request_filters)
    if selected_status:
        extra_args["status"] = selected_status
    requests_list, request_pagination = list_supply_requests_page(
        status=selected_status,
        endpoint="admin_requests",
        extra_args=extra_args,
        filters=request_filters,
        viewer=current,
        apply_assignment_visibility=True,
    )
    status_tabs = []
    for value, label in [("", "Todas"), ("pending", "Pendentes"), ("approved", "Aprovadas"), ("rejected", "Recusadas"), ("deleted", "Excluídas")]:
        tab_args = request_filters_to_query_args(request_filters)
        tab_args["limit"] = request_pagination["limit"]
        if value:
            tab_args["status"] = value
        status_tabs.append({"value": value, "label": label, "active": selected_status == value, "url": url_for("admin_requests", **tab_args)})
    return render_template(
        "admin/requests.html",
        requests_list=requests_list,
        selected_status=selected_status,
        request_pagination=request_pagination,
        request_filters=request_filters,
        request_status_tabs=status_tabs,
        request_product_options=list_products_for_request_filters(),
        request_sort_options=REQUEST_SORT_OPTIONS,
        request_type_options=REQUEST_TYPE_OPTIONS,
        request_regional_options=REQUEST_REGIONAL_OPTIONS,
        request_filter_action="admin_requests",
        request_filter_show_type=True,
        request_filter_show_regional=True,
        regional_assignment_options=request_admin_assignment_admin_options() if current.is_dev else [],
        regional_assignments=list_request_regional_admin_assignment_details() if current.is_dev else {},
    )


@app.post("/admin/requests/regional-admins")
@admin_required
@page_access_required("admin_requests")
def admin_request_regional_admins_update():
    current = require_current_user()
    if not current.is_dev:
        abort(403)
    try:
        with db_connect() as conn:
            for option in REQUEST_REGIONAL_OPTIONS:
                regional = option["value"]
                raw_admin_id = (request.form.get(f"admin_{regional}") or "").strip()
                admin_id: int | None = None
                if raw_admin_id:
                    admin_id = int(raw_admin_id)
                set_request_regional_admin_assignment(conn, regional, admin_id, current.id)
            conn.commit()
        flash("Direcionamento de solicitações por regional atualizado.", "success")
    except ValueError as exc:
        flash(str(exc), "warning")
    except Exception as exc:
        app.logger.exception("Falha ao atualizar direcionamento de solicitações por regional")
        flash(f"Não consegui atualizar o direcionamento. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_requests")


@app.route("/admin/requests/attended")
@admin_required
@page_access_required("admin_requests_attended")
def admin_requests_attended():
    current = require_current_user()
    request_filters = normalize_request_filters_from_args()
    extra_args = request_filters_to_query_args(request_filters)
    requests_list, request_pagination = list_supply_requests_page(
        status="approved",
        endpoint="admin_requests_attended",
        filters=request_filters,
        extra_args=extra_args,
        viewer=current,
        apply_assignment_visibility=True,
    )
    return render_template(
        "admin/requests_attended.html",
        requests_list=requests_list,
        selected_status="approved",
        request_pagination=request_pagination,
        request_filters=request_filters,
        request_product_options=list_products_for_request_filters(),
        request_sort_options=REQUEST_SORT_OPTIONS,
        request_type_options=REQUEST_TYPE_OPTIONS,
        request_regional_options=REQUEST_REGIONAL_OPTIONS,
        request_filter_action="admin_requests_attended",
        request_filter_show_type=True,
        request_filter_show_regional=True,
    )



@app.route("/admin/material-entries", methods=["GET", "POST"])
@admin_required
@page_access_required("admin_stock")
def admin_material_entries():
    if not user_has_action_access(current_user(), "stock_material_entries"):
        flash("Seu tipo de acesso não pode acessar entrada de materiais.", "warning")
        return redirect(url_for("admin_stock"))
    if request.method == "POST":
        selected_product_id = parse_required_positive_int(request.form.get("product_id")) or None
        item_name = (request.form.get("item_name") or "").strip()
        quantity = parse_required_positive_int(request.form.get("quantity")) or 0
        unit_price_cents = parse_money_to_cents(request.form.get("unit_price"))
        unit_measure = (request.form.get("unit_measure") or "un").strip() or "un"
        notes = (request.form.get("notes") or "").strip()
        if selected_product_id:
            with db_connect() as lookup_conn:
                existing_row = lookup_conn.execute(
                    "SELECT * FROM products WHERE id = ? AND active = 1 AND catalog_archived = 0 AND stock_tag = ? LIMIT 1",
                    (int(selected_product_id), SUPPLY_STOCK_TAG),
                ).fetchone()
            existing_product = row_to_product(existing_row) if existing_row is not None else None
            if existing_product is not None:
                item_name = existing_product.name
                unit_measure = unit_measure or existing_product.unit_measure or "un"
                if unit_price_cents <= 0:
                    unit_price_cents = int(existing_product.price_cents or 0)
            else:
                selected_product_id = None
        invoice_file = request.files.get("invoice_file")
        has_invoice = bool(invoice_file and invoice_file.filename)
        invoice_number = (request.form.get("invoice_number") or "").strip() if has_invoice else ""
        invoice_date = parse_optional_date(request.form.get("invoice_date")) if has_invoice else None
        invoice_value_cents = parse_money_to_cents(request.form.get("invoice_value")) if has_invoice else 0
        if not item_name or quantity <= 0:
            flash("Informe nome do item e quantidade válida para adicionar a entrada.", "warning")
            return redirect_to_return("admin_material_entries")
        invoice_file_name = ""
        invoice_file_key = ""
        if has_invoice and invoice_file is not None:
            try:
                invoice_bytes = invoice_file.read()
                invoice_file_name = invoice_file.filename or "nota_fiscal"
                invoice_file_key = storage_key("notas_fiscais", sao_paulo_filename_timestamp() + "_" + safe_filename(invoice_file_name))
                try:
                    upload_bytes_to_r2(invoice_file_key, invoice_bytes, invoice_file.mimetype or "application/octet-stream", {"type": "material_invoice"})
                except Exception as exc:
                    print(f"[R2] Não foi possível salvar nota fiscal da entrada: {exc}")
                    invoice_file_key = ""
            except Exception as exc:
                print(f"[ENTRADA MATERIAIS] Falha ao ler nota fiscal: {exc}")
                invoice_file_name = ""
                invoice_file_key = ""
        try:
            with db_connect() as conn:
                create_material_entry_record(
                    conn,
                    item_name=item_name,
                    quantity=quantity,
                    unit_measure=unit_measure,
                    unit_price_cents=unit_price_cents,
                    invoice_file_name=invoice_file_name,
                    invoice_file_key=invoice_file_key,
                    invoice_number=invoice_number,
                    invoice_date=invoice_date,
                    invoice_value_cents=invoice_value_cents,
                    notes=notes,
                    created_by_id=require_current_user().id,
                    movement_type="material_entry",
                    selected_product_id=selected_product_id,
                )
                conn.commit()
            flash("Entrada de material registrada e estoque atualizado.", "success")
        except Exception as exc:
            app.logger.exception("Falha ao registrar entrada de materiais")
            flash(f"Não consegui registrar a entrada. Erro: {type(exc).__name__}.", "danger")
        return redirect_to_return("admin_material_entries")
    entries = list_material_entries(limit=DEFAULT_TABLE_PAGE_SIZE)
    material_product_options = list_products_for_material_entry_options()
    return render_template(
        "admin/material_entries.html",
        entries=entries,
        material_product_options=material_product_options,
        material_product_options_json=material_entry_product_options_payload(material_product_options),
    )


@app.get("/admin/material-entries/model")
@admin_required
@page_access_required("admin_stock")
def admin_material_entries_template():
    if not user_has_action_access(current_user(), "stock_material_entries"):
        flash("Seu tipo de acesso não pode baixar modelo de entrada de materiais.", "warning")
        return redirect(url_for("admin_stock"))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Entrada de Materiais"
    headers = ["Nome do item", "Quantidade", "Valor unitário", "Unidade de medida", "Número da nota", "Data da nota", "Valor da nota", "Observações"]
    worksheet.append(headers)
    worksheet.append(["Envelope de segurança M", 100, 0.65, "un", "NF-0001", "2026-06-23", 65.00, "Exemplo de preenchimento"])
    header_fill = PatternFill("solid", fgColor="E60012")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    widths = [38, 16, 18, 20, 22, 18, 18, 48]
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    worksheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="modelo_entrada_materiais.xlsx")


@app.post("/admin/material-entries/import")
@admin_required
@page_access_required("admin_stock")
def admin_material_entries_import():
    if not user_has_action_access(current_user(), "stock_material_entries"):
        flash("Seu tipo de acesso não pode importar entrada de materiais.", "warning")
        return redirect(url_for("admin_stock"))
    uploaded = request.files.get("spreadsheet")
    if uploaded is None or not uploaded.filename:
        flash("Selecione uma planilha .xlsx de entrada de materiais.", "warning")
        return redirect_to_return("admin_material_entries")
    if not uploaded.filename.lower().endswith(".xlsx"):
        flash("Importe apenas arquivos .xlsx.", "warning")
        return redirect_to_return("admin_material_entries")
    try:
        uploaded_bytes = uploaded.read()
        if not uploaded_bytes:
            flash("A planilha enviada está vazia.", "warning")
            return redirect_to_return("admin_material_entries")
        if len(uploaded_bytes) <= int(os.getenv("IMPORT_BACKUP_MAX_BYTES", "5242880")):
            try:
                upload_bytes_to_r2(storage_key("imports", "entrada_materiais_" + sao_paulo_filename_timestamp() + "_" + safe_filename(uploaded.filename)), uploaded_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", {"type": "material_entries_import"})
            except Exception as exc:
                print(f"[R2] Não foi possível salvar planilha de entrada: {exc}")
        imported, skipped, errors = import_material_entries_from_workbook_bytes(uploaded_bytes, require_current_user().id)
        if errors:
            flash("Algumas linhas foram ignoradas: " + "; ".join(errors[:4]), "warning")
        flash(f"Importação concluída: {imported} entrada(s) importada(s), {skipped} linha(s) ignorada(s).", "success" if imported else "warning")
    except Exception as exc:
        app.logger.exception("Falha ao importar entrada de materiais")
        flash(f"Não consegui importar a planilha. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_material_entries")


@app.get("/admin/material-entries/report")
@admin_required
@page_access_required("admin_stock")
def admin_material_entries_report():
    if not user_has_action_access(current_user(), "stock_material_entries"):
        flash("Seu tipo de acesso não pode acessar entrada de materiais.", "warning")
        return redirect(url_for("admin_stock"))
    denied = require_action_permission("stock_reports", "Seu tipo de acesso não pode gerar relatórios de entrada.", "admin_material_entries")
    if denied:
        return denied
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()
    if not start_raw or not end_raw:
        flash("Informe a data inicial e a data final para gerar o relatório de entradas.", "warning")
        return redirect_to_return("admin_material_entries")
    try:
        start_local = parse_report_date(start_raw, "Data inicial")
        end_base = parse_report_date(end_raw, "Data final")
        start_dt, end_dt = sao_paulo_report_bounds_to_utc(start_local, end_base)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect_to_return("admin_material_entries")
    if end_base < start_local:
        flash("A data final não pode ser menor que a data inicial.", "warning")
        return redirect_to_return("admin_material_entries")
    entries = list_material_entries(start_dt, end_dt)
    buffer = build_material_entries_report_pdf(entries, start_local, end_base, require_current_user())
    filename = f"relatorio_entrada_materiais_{start_local.strftime('%Y%m%d')}_{end_base.strftime('%Y%m%d')}.pdf"
    store_generated_file(storage_key("reports", "material_entries", filename), buffer, "application/pdf", {"type": "material_entries_report", "start_date": start_raw, "end_date": end_raw})
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/admin/assets")
@admin_required
@page_access_required("admin_stock")
def admin_assets():
    selected_base = (request.args.get("base") or "").strip()
    selected_franchise = (request.args.get("franchise") or "").strip()
    selected_regional = normalize_asset_regional(request.args.get("regional", ""))
    filtered_asset_base_options = base_unit_options_for_asset_regional(selected_regional)
    filtered_asset_franchise_options = franchise_unit_options_for_asset_regional(selected_regional)

    if selected_base and selected_base not in filtered_asset_base_options:
        selected_base = ""
    if selected_franchise and selected_franchise not in filtered_asset_franchise_options:
        selected_franchise = ""
    if selected_base and selected_franchise:
        flash("Selecione somente uma base ou uma franquia para filtrar.", "warning")
        selected_franchise = ""

    selected_unit = selected_base or selected_franchise
    assets = list_assets(base=selected_unit, regional=selected_regional)
    with db_connect() as conn:
        product_rows = conn.execute(
            "SELECT * FROM products WHERE active = 1 AND catalog_archived = 0 AND stock_tag = ? ORDER BY category ASC, name ASC",
            (ASSET_STOCK_TAG,),
        ).fetchall()
    asset_product_options = [
        {
            "id": product.id,
            "name": product.name,
            "category": product.category or "Sem categoria",
            "stock_quantity": product.stock_quantity,
            "unit_measure": product.unit_measure or "un",
        }
        for row in product_rows
        if (product := row_to_product(row)) is not None
    ]
    totals = {
        "assets": len(assets),
        "items": sum(sum(item.quantity for item in asset.items) for asset in assets),
        "bases": len({asset.base for asset in assets if asset.base}),
        "mg": sum(1 for asset in assets if asset.regional == "MG"),
        "spn": sum(1 for asset in assets if asset.regional == "SPN"),
        "matriz": sum(1 for asset in assets if asset.regional == "Matriz"),
        "sc": sum(1 for asset in assets if asset.regional in {"SC CGE", "SC RAO"}),
    }
    return render_template(
        "admin/assets.html",
        assets=assets,
        totals=totals,
        selected_base=selected_base,
        selected_franchise=selected_franchise,
        selected_regional=selected_regional,
        filtered_asset_base_options=filtered_asset_base_options,
        filtered_asset_franchise_options=filtered_asset_franchise_options,
        filtered_base_options=base_options_for_asset_regional(selected_regional),
        asset_product_options=asset_product_options,
    )


@app.post("/admin/assets/new")
@admin_required
@page_access_required("admin_stock")
def admin_asset_new():
    name = request.form.get("name", "").strip()
    base = request.form.get("base", "").strip()
    regional = request.form.get("regional", "").strip().upper()
    sector = request.form.get("sector", "").strip()
    manager = request.form.get("manager", "").strip()
    item_names = request.form.getlist("item_name")
    serial_numbers = request.form.getlist("serial_number")

    item_rows: list[tuple[str, str]] = []
    for index, raw_item_name in enumerate(item_names):
        item_name = (raw_item_name or "").strip()
        if not item_name:
            continue
        serial_number = (serial_numbers[index] if index < len(serial_numbers) else "").strip()
        item_rows.append((item_name[:180], serial_number[:120]))

    if not name or not base or not regional or not sector or not manager:
        flash("Preencha nome, base/franquia, regional, setor e gestor para adicionar o ativo.", "warning")
        return redirect(url_for("admin_assets"))
    if base not in BASE_FRANCHISE_OPTION_SET:
        flash("Selecione uma base ou franquia válida para o ativo.", "warning")
        return redirect(url_for("admin_assets"))
    if regional not in ASSET_REGIONAL_OPTION_SET:
        flash("Selecione uma regional válida para o ativo.", "warning")
        return redirect(url_for("admin_assets"))
    if asset_regional_for_base(base) != regional:
        flash("A base/franquia selecionada não pertence à regional informada.", "warning")
        return redirect(url_for("admin_assets", regional=regional))
    if not item_rows:
        flash("Adicione pelo menos um item ao ativo.", "warning")
        return redirect(url_for("admin_assets"))

    try:
        with db_connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO assets (name, base, regional, sector, manager, created_by_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (name[:180], base, regional, sector[:120], manager[:120], require_current_user().id, now_iso()),
            )
            asset_id = get_cursor_lastrowid(cursor)
            if asset_id is None:
                raise RuntimeError("Não foi possível identificar o ativo criado.")
            for item_name, serial_number in item_rows:
                conn.execute(
                    """
                    INSERT INTO asset_items (asset_id, item_name, serial_number, created_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (int(asset_id), item_name, serial_number, now_iso()),
                )
            conn.commit()
        flash("Ativo adicionado ao relatório.", "success")
    except Exception as exc:
        app.logger.exception("Falha ao adicionar ativo")
        flash(f"Não consegui adicionar o ativo. Erro: {type(exc).__name__}.", "danger")
    return redirect(url_for("admin_assets", base=base, regional=regional))


def admin_asset_new_with_stock():
    denied = require_action_permission("stock_assets_create", "Seu tipo de acesso não pode criar ativos ou baixar estoque.", "admin_assets")
    if denied:
        return denied
    name = request.form.get("name", "").strip()
    base_raw = request.form.get("base", "").strip()
    franchise_raw = request.form.get("franchise", "").strip()
    regional = normalize_asset_regional(request.form.get("regional", ""))
    is_special_regional = is_special_asset_regional(regional)
    try:
        base, selected_unit_kind = validate_unit_selection(base_raw, franchise_raw, required=not is_special_regional)
    except ValueError as exc:
        flash(str(exc), "warning")
        return redirect_to_return("admin_assets", regional=regional)
    sector = request.form.get("sector", "").strip()
    manager = request.form.get("manager", "").strip()
    item_names = request.form.getlist("item_name")
    product_ids = request.form.getlist("product_id")
    quantities = request.form.getlist("quantity")
    serial_numbers = request.form.getlist("serial_number")

    item_rows: list[dict[str, Any]] = []
    missing_product = False
    invalid_quantity = False
    for index, raw_item_name in enumerate(item_names):
        item_name = (raw_item_name or "").strip()
        product_id = parse_required_positive_int(product_ids[index] if index < len(product_ids) else "")
        quantity = parse_required_positive_int(quantities[index] if index < len(quantities) else "")
        if not item_name and product_id is None:
            continue
        if product_id is None:
            missing_product = True
            continue
        if quantity is None:
            invalid_quantity = True
            continue
        serial_number = (serial_numbers[index] if index < len(serial_numbers) else "").strip()
        item_rows.append({"product_id": product_id, "quantity": quantity, "serial_number": serial_number[:120]})

    if is_special_regional:
        base = regional
        selected_unit_kind = ""

    redirect_args = {"regional": regional}
    if not is_special_regional and base:
        redirect_args["franchise" if selected_unit_kind == "franchise" else "base"] = base

    if not name or not regional or not sector or not manager or (not is_special_regional and not base):
        flash("Preencha nome, base/franquia, regional, setor e gestor para adicionar o ativo.", "warning")
        return redirect_to_return("admin_assets")
    if not regional:
        flash("Selecione uma regional valida para o ativo.", "warning")
        return redirect_to_return("admin_assets")
    if not is_special_regional and base not in BASE_FRANCHISE_OPTION_SET:
        flash("Selecione uma base ou franquia valida para o ativo.", "warning")
        return redirect_to_return("admin_assets", regional=regional)
    if not is_special_regional and asset_regional_for_base(base) != regional:
        flash("A base/franquia selecionada nao pertence a regional informada.", "warning")
        return redirect_to_return("admin_assets", regional=regional)
    if missing_product:
        flash("Selecione cada item pela lista de produtos do portal.", "warning")
        return redirect_to_return("admin_assets", **redirect_args)
    if invalid_quantity:
        flash("Informe uma quantidade valida para cada item.", "warning")
        return redirect_to_return("admin_assets", **redirect_args)
    if not item_rows:
        flash("Adicione pelo menos um item ao ativo.", "warning")
        return redirect_to_return("admin_assets", **redirect_args)

    try:
        with db_connect() as conn:
            requested_by_product: dict[int, int] = {}
            for item in item_rows:
                requested_by_product[item["product_id"]] = requested_by_product.get(item["product_id"], 0) + item["quantity"]

            product_ids_unique = sorted(requested_by_product)
            placeholders = ", ".join(["?"] * len(product_ids_unique))
            product_rows = conn.execute(
                f"SELECT * FROM products WHERE catalog_archived = 0 AND stock_tag = ? AND id IN ({placeholders})",
                [ASSET_STOCK_TAG, *product_ids_unique],
            ).fetchall()
            product_map = {
                int(row["id"]): product
                for row in product_rows
                if (product := row_to_product(row)) is not None
            }
            missing_ids = [product_id for product_id in product_ids_unique if product_id not in product_map]
            inactive = [product.name for product in product_map.values() if not product.active]
            insufficient = [
                f"{product_map[product_id].name} (solicitado {quantity}, estoque {product_map[product_id].stock_quantity})"
                for product_id, quantity in requested_by_product.items()
                if product_id in product_map and product_map[product_id].stock_quantity < quantity
            ]
            if missing_ids:
                flash("Um ou mais produtos selecionados nao existem mais no cadastro.", "warning")
                return redirect_to_return("admin_assets", **redirect_args)
            if inactive:
                flash("Produto(s) inativo(s) nao podem ser vinculados a ativos: " + ", ".join(inactive), "warning")
                return redirect_to_return("admin_assets", **redirect_args)
            if insufficient:
                flash("Estoque insuficiente para: " + "; ".join(insufficient), "warning")
                return redirect_to_return("admin_assets", **redirect_args)

            current = require_current_user()
            cursor = conn.execute(
                """
                INSERT INTO assets (name, base, regional, sector, manager, created_by_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (name[:180], base, regional, sector[:120], manager[:120], current.id, now_iso()),
            )
            asset_id = get_cursor_lastrowid(cursor)
            if asset_id is None:
                raise RuntimeError("Nao foi possivel identificar o ativo criado.")
            for item in item_rows:
                product = product_map[item["product_id"]]
                quantity = item["quantity"]
                stock_before = product.stock_quantity
                stock_after = stock_before - quantity
                conn.execute(
                    "UPDATE products SET stock_quantity = ?, updated_at = ? WHERE id = ?",
                    (stock_after, now_iso(), product.id),
                )
                conn.execute(
                    """
                    INSERT INTO asset_items (asset_id, product_id, item_name, quantity, serial_number, created_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (int(asset_id), product.id, product.name, quantity, item["serial_number"], now_iso()),
                )
                record_stock_movement(
                    conn,
                    product.id,
                    -quantity,
                    stock_before,
                    stock_after,
                    "asset_allocation",
                    f"Saida para ativo #{asset_id} - {name}.",
                    created_by_id=current.id,
                )
                product.stock_quantity = stock_after
            conn.commit()
            try:
                asset_link_args = {"regional": regional, "_anchor": f"asset-{asset_id}"}
                if not is_special_regional and base:
                    asset_link_args["franchise" if selected_unit_kind == "franchise" else "base"] = base
                asset_link = public_url_for("admin_assets", **asset_link_args)
                notify_feishu_asset_created(
                    int(asset_id),
                    name,
                    base,
                    regional,
                    sector,
                    manager,
                    current,
                    item_rows,
                    product_map,
                    asset_link,
                )
            except Exception:
                app.logger.exception("Falha ao preparar notificacao Feishu do ativo")
        flash("Ativo adicionado ao relatorio e estoque baixado.", "success")
    except Exception as exc:
        app.logger.exception("Falha ao adicionar ativo")
        flash(f"Nao consegui adicionar o ativo. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_assets", **redirect_args)


app.view_functions["admin_asset_new"] = admin_required(page_access_required("admin_stock")(admin_asset_new_with_stock))


@app.route("/admin/stock")
@admin_required
@page_access_required("admin_stock")
def admin_stock():
    with db_connect() as conn:
        product_rows = conn.execute(
            """
            SELECT id, name, category, category_emoji, image_name, image_key, image_content_type,
                   unit_measure, is_kit, kit_quantity, description, stock_quantity, price_cents,
                   limit_base, limit_franchise, limit_block_days, min_order_quantity, min_stock, max_stock,
                   active, visible_base, visible_franchise, internal, catalog_archived, created_at, updated_at
             FROM products
             WHERE catalog_archived = 0 AND stock_tag = ?
             ORDER BY active DESC, category ASC, name ASC
             LIMIT ?
            """,
            (SUPPLY_STOCK_TAG, bounded_int(os.getenv("D1_STOCK_PRODUCT_LIMIT"), 250 if low_row_read_mode() else 1000, 50, 1000),),
        ).fetchall()
        movement_rows = conn.execute(
            """
            SELECT sm.*,
                   p.name AS product_name,
                   p.category AS product_category,
                   COALESCE(u.responsible_name, reviewer.responsible_name) AS created_by_name,
                   COALESCE(u.username, reviewer.username) AS created_by_username
              FROM stock_movements sm
              LEFT JOIN products p ON p.id = sm.product_id
              LEFT JOIN users u ON u.id = sm.created_by_id
             LEFT JOIN supply_requests sr ON sr.id = sm.request_id
             LEFT JOIN users reviewer ON reviewer.id = sr.reviewed_by_id
             WHERE COALESCE(p.stock_tag, ?) = ?
             ORDER BY sm.created_at DESC, sm.id DESC
             LIMIT ?
            """,
            (SUPPLY_STOCK_TAG, SUPPLY_STOCK_TAG, DEFAULT_TABLE_PAGE_SIZE),
        ).fetchall()
        if exact_counts_enabled():
            totals = {
                "products": conn.execute("SELECT COUNT(*) AS total FROM products WHERE catalog_archived = 0 AND stock_tag = ?", (SUPPLY_STOCK_TAG,)).fetchone()["total"],
                "stock_total": conn.execute("SELECT COALESCE(SUM(stock_quantity), 0) AS total FROM products WHERE catalog_archived = 0 AND stock_tag = ?", (SUPPLY_STOCK_TAG,)).fetchone()["total"],
                "critical": conn.execute("SELECT COUNT(*) AS total FROM products WHERE catalog_archived = 0 AND stock_tag = ? AND min_stock IS NOT NULL AND stock_quantity <= min_stock", (SUPPLY_STOCK_TAG,)).fetchone()["total"],
                "movements": conn.execute("SELECT COUNT(*) AS total FROM stock_movements sm LEFT JOIN products p ON p.id = sm.product_id WHERE COALESCE(p.stock_tag, ?) = ?", (SUPPLY_STOCK_TAG, SUPPLY_STOCK_TAG)).fetchone()["total"],
            }
        else:
            totals = {
                "products": len(product_rows),
                "stock_total": sum(int(row["stock_quantity"] or 0) for row in product_rows),
                "critical": sum(1 for row in product_rows if row["min_stock"] is not None and int(row["stock_quantity"] or 0) <= int(row["min_stock"] or 0)),
                "movements": len(movement_rows),
            }

    formatted_movement_rows: list[dict[str, Any]] = []
    for row in movement_rows:
        movement = {key: row[key] for key in row.keys()} if hasattr(row, "keys") else dict(row)
        movement["created_at"] = format_sao_paulo_datetime(movement.get("created_at"))
        formatted_movement_rows.append(movement)
    movement_rows = formatted_movement_rows

    products = [product for row in product_rows if (product := row_to_product(row)) is not None]

    stock_rows: list[dict[str, Any]] = []
    for product in products:
        status = stock_status_class(product)
        maximum = product.max_stock if product.max_stock and product.max_stock > 0 else None
        minimum = product.min_stock if product.min_stock and product.min_stock > 0 else None
        reference = maximum or max(product.stock_quantity, minimum or 0, 1)
        percent = int(min(100, max(0, (product.stock_quantity / reference) * 100)))
        stock_rows.append({
            "product": product,
            "status": status,
            "label": stock_status_label(product),
            "percent": percent,
            "min": product.min_stock,
            "max": product.max_stock,
        })

    totals["healthy"] = sum(1 for row in stock_rows if row["status"] == "good")
    totals["normal"] = sum(1 for row in stock_rows if row["status"] == "normal")

    risk_rows = [row for row in stock_rows if row["status"] == "critical"]
    if not risk_rows:
        risk_rows = stock_rows[:6]

    chart_products = []
    category_map: dict[str, dict[str, Any]] = {}
    for row in stock_rows:
        product = row["product"]
        category = product.category or "Sem categoria"
        chart_products.append({
            "id": product.id,
            "name": product.name,
            "category": category,
            "stock": product.stock_quantity,
            "min": product.min_stock or 0,
            "max": product.max_stock or 0,
            "status": row["status"],
            "label": row["label"],
        })
        bucket = category_map.setdefault(category, {
            "category": category,
            "stock": 0,
            "min": 0,
            "max": 0,
            "products": 0,
            "critical": 0,
            "good": 0,
            "normal": 0,
        })
        bucket["stock"] += product.stock_quantity
        bucket["min"] += product.min_stock or 0
        bucket["max"] += product.max_stock or 0
        bucket["products"] += 1
        bucket[row["status"]] += 1

    chart_categories = sorted(category_map.values(), key=lambda item: item["category"].lower())

    return render_template(
        "admin/stock.html",
        products=products,
        stock_rows=stock_rows,
        risk_rows=risk_rows[:8],
        movement_rows=movement_rows,
        totals=totals,
        chart_products=chart_products,
        chart_categories=chart_categories,
        movement_type_label=movement_type_label,
        base_unit_options=BASE_UNIT_OPTIONS,
        franchise_unit_options=FRANCHISE_UNIT_OPTIONS,
    )


@app.get("/admin/stock/requests-report")
@admin_required
@page_access_required("admin_stock")
def admin_stock_requests_report():
    denied = require_action_permission("stock_reports", "Seu tipo de acesso não pode gerar relatórios de estoque.", "admin_stock")
    if denied:
        return denied
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()
    base_raw = (request.args.get("base") or "").strip()
    franchise_raw = (request.args.get("franchise") or "").strip()
    all_units = (request.args.get("all_units") or "").strip().lower() in {"1", "true", "on", "all", "todos"}
    if not start_raw or not end_raw:
        flash("Informe a data inicial e a data final para gerar o relatório.", "warning")
        return redirect(url_for("admin_stock"))

    try:
        if all_units:
            selected_unit, selected_kind = "", "all"
        else:
            selected_unit, selected_kind = validate_unit_selection(base_raw, franchise_raw, required=True)
        start_local = parse_report_date(start_raw, "Data inicial")
        end_base = parse_report_date(end_raw, "Data final")
        start_dt, end_dt = sao_paulo_report_bounds_to_utc(start_local, end_base)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin_stock"))

    if end_base < start_local:
        flash("A data final não pode ser menor que a data inicial.", "warning")
        return redirect(url_for("admin_stock"))

    requests_list = list_supply_requests_between(start_dt, end_dt, selected_unit)
    buffer = build_supply_requests_period_report_pdf(requests_list, start_local, end_base, require_current_user(), selected_unit, selected_kind)
    unit_slug = "todas_unidades" if selected_kind == "all" else (re.sub(r"[^A-Za-z0-9]+", "_", selected_unit).strip("_").lower() or "unidade")
    filename = f"relatorio_solicitacoes_insumos_{unit_slug}_{start_local.strftime('%Y%m%d')}_{end_base.strftime('%Y%m%d')}.pdf"
    store_generated_file(
        storage_key("reports", "supply_requests", filename),
        buffer,
        "application/pdf",
        {"type": "supply_requests_period_report", "start_date": start_raw, "end_date": end_raw, "unit": selected_unit, "unit_kind": selected_kind},
    )
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.get("/admin/assets/period-report")
@admin_required
@page_access_required("admin_stock")
def admin_assets_period_report():
    denied = require_action_permission("stock_reports", "Seu tipo de acesso não pode gerar relatórios de ativos.", "admin_assets")
    if denied:
        return denied
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()
    base_raw = (request.args.get("base") or "").strip()
    franchise_raw = (request.args.get("franchise") or "").strip()
    all_units = (request.args.get("all_units") or "").strip().lower() in {"1", "true", "on", "all", "todos"}
    if not start_raw or not end_raw:
        flash("Informe a data inicial e a data final para gerar o relatório de ativos.", "warning")
        return redirect_to_return("admin_assets")

    try:
        if all_units:
            selected_unit, selected_kind = "", "all"
        else:
            selected_unit, selected_kind = validate_unit_selection(base_raw, franchise_raw, required=True)
        start_local = parse_report_date(start_raw, "Data inicial")
        end_base = parse_report_date(end_raw, "Data final")
        start_dt, end_dt = sao_paulo_report_bounds_to_utc(start_local, end_base)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect_to_return("admin_assets")

    if end_base < start_local:
        flash("A data final não pode ser menor que a data inicial.", "warning")
        return redirect_to_return("admin_assets")

    assets = list_assets_between(start_dt, end_dt, selected_unit)
    buffer = build_assets_period_report_pdf(assets, start_local, end_base, require_current_user(), selected_unit, selected_kind)
    unit_slug = "todas_unidades" if selected_kind == "all" else (re.sub(r"[^A-Za-z0-9]+", "_", selected_unit).strip("_").lower() or "unidade")
    filename = f"relatorio_ativos_{unit_slug}_{start_local.strftime('%Y%m%d')}_{end_base.strftime('%Y%m%d')}.pdf"
    store_generated_file(
        storage_key("reports", "assets", filename),
        buffer,
        "application/pdf",
        {"type": "assets_period_report", "start_date": start_raw, "end_date": end_raw, "unit": selected_unit, "unit_kind": selected_kind},
    )
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/admin/requests/<int:request_id>")
@admin_required
@page_access_any_required(["admin_requests", "admin_requests_attended"])
def admin_request_detail(request_id: int):
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    current = require_current_user()
    if not can_view_supply_request_by_assignment(supply_request, current):
        abort(403)
    default_back_endpoint = "admin_requests" if user_has_page_access(current, "admin_requests") else "admin_requests_attended"
    back_url = return_target(default_back_endpoint)
    return render_template("admin/request_detail.html", supply_request=supply_request, back_url=back_url)




@app.post("/admin/requests/<int:request_id>/items")
@admin_required
@page_access_required("admin_requests")
def admin_request_update_items(request_id: int):
    denied = require_action_permission("requests_edit_items", "Seu tipo de acesso não pode editar itens de solicitações.", "admin_requests")
    if denied:
        return denied
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    if not can_view_supply_request_by_assignment(supply_request, require_current_user()):
        abort(403)
    if supply_request.status != "pending":
        flash("Apenas solicitações pendentes podem ter quantidades editadas.", "warning")
        return redirect_to_return("admin_request_detail", request_id=request_id)

    updated = 0
    with db_connect() as conn:
        for item in supply_request.items:
            quantity = parse_required_positive_int(request.form.get(f"quantity_{item.id}"))
            if quantity is None:
                flash("Todas as quantidades precisam ser maiores que zero.", "warning")
                return redirect_to_return("admin_request_detail", request_id=request_id)
            if quantity != item.quantity:
                conn.execute("UPDATE request_items SET quantity = ? WHERE id = ? AND request_id = ?", (quantity, item.id, request_id))
                updated += 1
        if updated:
            record_request_action(conn, request_id, "items_updated", require_current_user().id, f"Quantidades editadas em {updated} item(ns).")
        conn.commit()

    flash("Quantidades atualizadas." if updated else "Nenhuma quantidade foi alterada.", "success")
    return redirect_to_return("admin_request_detail", request_id=request_id)


@app.post("/admin/requests/<int:request_id>/approve")
@admin_required
@page_access_required("admin_requests")
def admin_request_approve(request_id: int):
    denied = require_action_permission("requests_approve_reject", "Seu tipo de acesso não pode aprovar solicitações.", "admin_requests")
    if denied:
        return denied
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    if not can_view_supply_request_by_assignment(supply_request, require_current_user()):
        abort(403)
    if supply_request.status != "pending":
        flash("Apenas solicitações pendentes podem ser aprovadas.", "warning")
        return redirect_to_return("admin_request_detail", request_id=request_id)

    insufficient: list[str] = []
    for item in supply_request.items:
        product = get_product(item.product_id)
        if product is None or not product.active:
            insufficient.append(item.product_name_snapshot)
        elif item.quantity > product.stock_quantity:
            insufficient.append(f"{product.name} (solicitado {item.quantity}, estoque {product.stock_quantity})")

    if insufficient:
        flash("Estoque insuficiente para aprovar: " + "; ".join(insufficient), "danger")
        return redirect_to_return("admin_request_detail", request_id=request_id)

    admin_note = request.form.get("admin_note", "").strip()
    current = require_current_user()
    with db_connect() as conn:
        for item in supply_request.items:
            product = get_product(item.product_id)
            stock_before = product.stock_quantity if product else 0
            stock_after = stock_before - item.quantity
            conn.execute("UPDATE products SET stock_quantity = ?, updated_at = ? WHERE id = ?", (stock_after, now_iso(), item.product_id))
            record_stock_movement(
                conn,
                product_id=item.product_id,
                request_id=request_id,
                created_by_id=current.id,
                movement_type="request_approved",
                quantity_delta=-item.quantity,
                stock_before=stock_before,
                stock_after=stock_after,
                note=f"Solicitação #{request_id} aprovada.",
            )
        conn.execute(
            """
            UPDATE supply_requests
            SET status = 'approved', reviewed_at = ?, reviewed_by_id = ?, admin_note = ?
            WHERE id = ?
            """,
            (now_iso(), current.id, admin_note, request_id),
        )
        record_request_action(conn, request_id, "approved", current.id, admin_note or "Solicitação aprovada e estoque descontado.")
        conn.commit()

    try:
        updated_request = get_supply_request(request_id)
        if updated_request is not None:
            notify_feishu_supply_request_action(updated_request, "approved", current, public_url_for("admin_request_detail", request_id=request_id), admin_note)
    except Exception:
        app.logger.exception("Falha ao preparar notificacao Feishu da aprovacao da solicitacao")

    flash("Solicitação aprovada e estoque descontado.", "success")
    return redirect_to_return("admin_request_detail", request_id=request_id)


@app.post("/admin/requests/<int:request_id>/reject")
@admin_required
@page_access_required("admin_requests")
def admin_request_reject(request_id: int):
    denied = require_action_permission("requests_approve_reject", "Seu tipo de acesso não pode recusar solicitações.", "admin_requests")
    if denied:
        return denied
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    if not can_view_supply_request_by_assignment(supply_request, require_current_user()):
        abort(403)
    if supply_request.status != "pending":
        flash("Apenas solicitações pendentes podem ser recusadas.", "warning")
        return redirect_to_return("admin_request_detail", request_id=request_id)

    admin_note = request.form.get("admin_note", "").strip()
    current = require_current_user()
    with db_connect() as conn:
        conn.execute(
            """
            UPDATE supply_requests
            SET status = 'rejected', reviewed_at = ?, reviewed_by_id = ?, admin_note = ?
            WHERE id = ?
            """,
            (now_iso(), current.id, admin_note, request_id),
        )
        record_request_action(conn, request_id, "rejected", current.id, admin_note or "Solicitação recusada.")
        conn.commit()
    try:
        updated_request = get_supply_request(request_id)
        if updated_request is not None:
            notify_feishu_supply_request_action(updated_request, "rejected", current, public_url_for("admin_request_detail", request_id=request_id), admin_note)
    except Exception:
        app.logger.exception("Falha ao preparar notificacao Feishu da recusa da solicitacao")
    flash("Solicitação recusada.", "success")
    return redirect_to_return("admin_request_detail", request_id=request_id)


@app.post("/admin/requests/<int:request_id>/delete")
@admin_required
@page_access_any_required(["admin_requests", "admin_requests_attended"])
def admin_request_delete(request_id: int):
    denied = require_action_permission("requests_delete", "Seu tipo de acesso não pode excluir solicitações.", "admin_requests")
    if denied:
        return denied
    supply_request = get_supply_request(request_id)
    if supply_request is None:
        abort(404)
    if not can_view_supply_request_by_assignment(supply_request, require_current_user()):
        abort(403)
    if supply_request.status == "deleted":
        flash("Esta solicitação já está marcada como excluída.", "warning")
        return redirect_to_return("admin_requests")
    current = require_current_user()
    try:
        with db_connect() as conn:
            conn.execute(
                """
                UPDATE supply_requests
                   SET status = 'deleted', reviewed_at = ?, reviewed_by_id = ?
                 WHERE id = ?
                """,
                (now_iso(), current.id, request_id),
            )
            record_request_action(conn, request_id, "deleted", current.id, "Solicitação marcada como excluída.")
            conn.commit()
        updated_request = get_supply_request(request_id)
        if updated_request is not None:
            notify_feishu_supply_request_action(updated_request, "deleted", current, public_url_for("admin_request_detail", request_id=request_id), "Solicitação marcada como excluída.")
        flash("Solicitação marcada como excluída.", "success")
    except Exception as exc:
        app.logger.exception("Falha ao marcar solicitação como excluída")
        flash(f"Não consegui excluir a solicitação. Erro: {type(exc).__name__}.", "danger")
    return redirect_to_return("admin_requests")


@app.errorhandler(403)
def forbidden(_):
    return render_template("error.html", title="Acesso negado", message="Você não possui permissão para acessar esta página."), 403


@app.errorhandler(404)
def not_found(_):
    return render_template("error.html", title="Página não encontrada", message="A página solicitada não existe."), 404


@app.errorhandler(401)
def unauthorized(_):
    return redirect(url_for("login"))


setup_database()


def run_local_server() -> None:
    port = resolve_port()
    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    open_browser_when_ready(port)
    print("==============================================")
    print(" Portal de Insumos J&T Express")
    print(f" Endereço local: http://127.0.0.1:{port}")
    print(" Para encerrar, feche esta janela ou pressione CTRL+C.")
    print("==============================================")
    app.run(host="127.0.0.1", port=port, debug=debug_mode, use_reloader=False)


if __name__ == "__main__":
    try:
        run_local_server()
    except KeyboardInterrupt:
        print("\nServidor encerrado.")
    except Exception as exc:
        print("\nERRO AO ABRIR O PORTAL:")
        print(type(exc).__name__, "-", exc)
        print("\nTente rodar o arquivo ABRIR_PORTAL.bat para instalar dependências e abrir o sistema.")
        input("\nPressione Enter para fechar...")
        raise
